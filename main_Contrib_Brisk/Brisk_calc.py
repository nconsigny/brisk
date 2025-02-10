#########

import sys
import Environnement as EnvB
import Files_utils
import EC_tools_fires as Ec
from Environnement import (u_round, ingberg_temp_seuil)

import gc
import os
import glob
import shutil
import subprocess
import time
from datetime import timedelta, date
import xml.etree.ElementTree as Etr
import copy
# from typing import Any, Optional  # ,List
# from typing import Dict, Any
import pandas as pd
import numpy as np
import scipy as sc
# from typing import List
import sympy as sy
from sympy import sqrt
# import math  # pour nan_value = float('nan') math.isnan(nan_value)
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, Reference
from openpyxl.chart.series_factory import SeriesFactory  # as Series
# from openpyxl.chart import Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter  # , FORMULAE
from openpyxl.styles import Alignment
from openpyxl.styles import Font
import csv
from Files_utils import (lire_texte_xml, find_and_replace_value_in_xml,
                         remplacer_nombres_dans_xml, remplacer_str_dans_xml, remplacer_double_nombres_dans_xml,
                         modif_fire_description_in_xml, find_and_replace_value_varname, find_and_replace_item_varname,
                         lire_str_dans_xml, find_value_in_xml, find_attribute_value)
# from Files_utils import (lire_nombres_xml, lire_first_texte_arbre_xml) class method from_brisk...
from Files_utils import test_open_write
import re
from pathlib import Path


# end import subprocess


#########
# Time function definition

# from scipy import interpolate
# from scipy.integrate import quad

class TimeFunction:

    def __init__(self, abscissas, ordinates):
        self.abscissas = np.array(abscissas)
        self.ordinates = np.array(ordinates)
        self.f = sc.interpolate.interp1d(abscissas, ordinates)
        self.inverse_f = sc.interpolate.interp1d(ordinates, abscissas)

    def valeur(self, x):
        if x < self.abscissas[0] or x > self.abscissas[-1]:
            return 0
        else:
            return float(self.f(x))

    def abscisse(self, y):
        return float(self.inverse_f(y))

    def integral_total(self):  # return [0] integral: float  [1] abs-err: float
        y = sc.integrate.quad(self.valeur, float(self.abscissas[0]), float(self.abscissas[-1]),
                              full_output=1, limit=100)
        return y

    def integral(self, x1, x2):  # return [0] integral: float  [1] abs-err: float
        y = sc.integrate.quad(self.valeur, x1, x2,
                              full_output=1, limit=100)
        return y

    def cherche_abscisse(self, angle_inf):
        for i in range(1, len(self.abscissas) - 2):
            a = calculer_angle(self.abscissas[i - 1], self.abscissas[i], self.abscissas[i + 1],
                               self.ordinates[i - 1], self.ordinates[i], self.ordinates[i + 1])
            if a < angle_inf:
                return float(self.abscissas[i])
        return None

    # def integral(self, a, b):  # return [0] integral: float  [1] abs-err: float
    #     if a < self.abscissas[0] or b > self.abscissas[-1]:
    #         return 0
    #     else:
    #         return sc.integrate.quad(self.valeur, a, b)


class HrrFunction(TimeFunction):
    def __init__(self, x_coord, y_coord):
        super().__init__(x_coord, y_coord)
        self.t_fin = float(self.abscissas[-1])
        self.t_max = self.cherche_abscisse(-20)
        # Filtrer les abscissas et ordinates pour f1 (si listes)
        # abscissas_f1 = [x for x in self.abscissas if x <= abs1]
        # ordinates_f1 = [y for x, y in zip(self.abscissas, self.ordinates) if x <= abs1]
        # self.f1 = TimeFunction(abscissas_f1, ordinates_f1)
        if self.t_max is not None:
            self.f_growth = TimeFunction(self.abscissas[self.abscissas <= self.t_max],
                                         self.ordinates[self.abscissas <= self.t_max])
            f_reste = TimeFunction(self.abscissas[self.abscissas >= self.t_max],
                                   self.ordinates[self.abscissas >= self.t_max])
            f_reste_temp = TimeFunction(self.abscissas[self.abscissas > self.t_max],
                                        self.ordinates[self.abscissas > self.t_max])
            self.t_dec = f_reste_temp.cherche_abscisse(-10)
            if self.t_dec is not None:
                self.f_fully_developed = TimeFunction(
                    self.abscissas[(self.abscissas >= self.t_max) & (self.abscissas <= self.t_dec)],
                    self.ordinates[(self.abscissas >= self.t_max) & (self.abscissas <= self.t_dec)])
                self.f_decay = TimeFunction(self.abscissas[self.abscissas >= self.t_dec],
                                            self.ordinates[self.abscissas >= self.t_dec])
            else:
                self.f_fully_developed = f_reste
                self.f_decay = TimeFunction([self.abscissas[-1], self.abscissas[-1]],
                                            [self.ordinates[-1], self.ordinates[-1]])
                self.t_dec = float(self.abscissas[-1])
        else:
            self.f_growth = TimeFunction(x_coord, y_coord)
            self.f_fully_developed = TimeFunction([self.abscissas[-1], self.abscissas[-1]],
                                                  [self.ordinates[-1], self.ordinates[-1]])
            self.f_decay = self.f_fully_developed
            self.t_max = float(self.abscissas[-1])
            self.t_dec = float(self.t_max)

    def integral_total(self):
        integral1 = self.f_growth.integral_total()[0]
        integral2 = self.f_fully_developed.integral_total()[0]
        integral3 = self.f_decay.integral_total()[0]
        total_integral = integral1 + integral2 + integral3
        return total_integral

    def to_cfast(self, ident, height, yco, ys):
        # Fonction Area pour calculer la valeur de la colonne AREA
        def area(y):
            rho_air = 1.195
            cp_air = 1.01
            t_air = 293
            g = 9.81
            # Cfast User Guide, note 1 p28, Froude number =1
            d_star = (y / (rho_air * cp_air * t_air * sqrt(g))) ** (2 / 5)  # [m]
            return max(1e-3, sy.pi * d_star ** 2 / 4)  # [m²]

        header = (f"&TABL ID = '{ident}' LABELS = 'TIME', 'HRR', 'HEIGHT', 'AREA', 'CO_YIELD', 'SOOT_YIELD', "
                  f"'HCN_YIELD', 'HCL_YIELD', 'TRACE_YIELD' /\n")
        data_lines = []
        # TODO Check for Cfast Maximum number of data points for a single fire definition (199)
        for t, h in zip(self.abscissas, self.ordinates):
            f_area = area(h)
            data_line = (f"&TABL ID = '{ident}', DATA = {t}, {u_round(h, '[kW]')}, {height}, "
                         f"{u_round(f_area, '[m²]')}, {yco}, {ys}"
                         f"0, 0, 0 /\n")
            data_lines.append(data_line)
        return header + ''.join(data_lines)


class Fuel:
    def __init__(self, fuel_type: str, carbon: float, chlorine: float, hydrogen: float, nitrogen: float, oxygen: float,
                 hc: float, rf: float, yco: float, ys: float):
        self.fuel_type = fuel_type
        self.carbon = carbon
        self.chlorine = chlorine
        self.hydrogen = hydrogen
        self.nitrogen = nitrogen
        self.oxygen = oxygen
        self.heat_of_combustion = hc  # [kJ/g]
        self.radiative_fraction = rf
        self.co_yield = yco
        self.soot_yield = ys

    def __repr__(self):
        return (f"Fuel(fuel_type='{self.fuel_type}', Carbon={self.carbon}, "
                f"Chlorine={self.chlorine}, Hydrogen={self.hydrogen}, "
                f"Nitrogen={self.nitrogen}, Oxygen={self.oxygen}, CO_Yield={self.co_yield}, "
                f"SOOT_Yield={self.soot_yield}"
                f"Heat of combustion={self.heat_of_combustion}, Radiative loss fraction ={self.radiative_fraction})")

    def copy(self):
        return Fuel(self.fuel_type, self.carbon, self.chlorine, self.hydrogen,
                    self.nitrogen, self.oxygen, self.heat_of_combustion, self.radiative_fraction,
                    self.co_yield, self.soot_yield)


# Définir les constantes globales see 241028 Predictive_timber_charring_V3 sheet "Tempo HRR Initial"
Pine = Fuel(fuel_type='Pine SFPE', carbon=1, chlorine=0, hydrogen=1.7, nitrogen=0, oxygen=0.83,
            hc=17.5, rf=0.298, yco=0.005, ys=0.01)
Pine_Brisk = Fuel(fuel_type='Pine Brisk', carbon=1, chlorine=0, hydrogen=1.7, nitrogen=0, oxygen=0.83,
                  hc=12.4, rf=0.298, yco=0.005, ys=0.01)
Propane = Fuel(fuel_type='Propane', carbon=3, chlorine=0, hydrogen=8, nitrogen=0, oxygen=0, hc=43.7, rf=0.3,
               yco=0.005, ys=0.024)


class Fire:
    def __init__(self, fd, ident, x, y, z, hrr_pua=250, dim_x=0.3, dim_y=0.3, dim_h=0, h=0, mc=0):
        self.id = ident + '_hrr'
        self.hrr_pua = hrr_pua  # [kW/m²]
        self.fd = fd  # [MJ/m²]
        self.x = x
        self.y = y
        self.z = z
        self.dim_x = dim_x
        self.dim_y = dim_y
        self.dim_h = dim_h
        self.elevation = h
        self.mass_of_combustion = mc
        self.wind = [0, 0, 1]
        # self.fuel_type = ft -> Fuel
        self.fuel = None
        self.fuel_ingberg_equivalency = 0  # [mn]
        self.total_fire_ingberg_equivalency = 0  # [mn]

    def to_cfast(self, ident):
        return (f"&FIRE ID = '{self.id}' COMP_ID = '{ident}', FIRE_ID = '{self.id}' LOCATION = {self.x}, {self.y} /\n"
                f"&CHEM ID = '{self.id}' CARBON = {self.fuel.carbon} CHLORINE = {self.fuel.chlorine} "
                f"HYDROGEN = {self.fuel.hydrogen} NITROGEN = {self.fuel.nitrogen} OXYGEN = {self.fuel.oxygen} "
                f"HEAT_OF_COMBUSTION = {self.fuel.heat_of_combustion * 1000} "
                f"RADIATIVE_FRACTION = {self.fuel.radiative_fraction} /\n")


def calculer_angle(x1, x2, x3, y1, y2, y3):
    # Vecteurs P1P2 et P2P3
    vecteur1 = sy.Matrix([x2 - x1, y2 - y1])
    vecteur2 = sy.Matrix([x3 - x2, y3 - y2])

    # Produit scalaire des vecteurs
    produit_scalaire = vecteur1.dot(vecteur2)

    # Normes des vecteurs
    norme_vecteur1 = vecteur1.norm()
    norme_vecteur2 = vecteur2.norm()

    # Produit vectoriel (déterminant 2D)
    produit_vectoriel = vecteur1[0] * vecteur2[1] - vecteur1[1] * vecteur2[0]
    # Calcul de l'angle et signe en radians
    angle_radians = sy.sign(produit_vectoriel) * sy.acos(produit_scalaire / (norme_vecteur1 * norme_vecteur2))

    # Conversion de l'angle en degrés
    angle_degrees = sy.deg(angle_radians)
    if angle_degrees.is_real:
        angle_degrees = float(angle_degrees)
    else:
        angle_degrees = float(angle_degrees.as_real_imag()[0])

    return angle_degrees


# End Time function definition
#########


#########
# define Compartment Class
class Vent:
    def __init__(self, num, ident='V', room1='1', room2='outside', width=0, height=0,
                 sill=0, offset=0, discharge_coefficient=0.68, face='Front', typ='WALL'):
        self.num = num
        if ident == 'V':
            self.id = ident + str(num)
        else:
            self.id = ident
        self.Room1 = room1
        self.Room2 = room2
        self.width = width
        self.height = height
        self.sill = sill
        self.offset = offset
        self.discharge_coefficient = discharge_coefficient
        self.face = face
        self.type = typ

    def to_cfast(self):
        return (f"&VENT TYPE = '{self.type}' ID = '{self.id}' COMP_IDS = '{self.Room1}' '{self.Room2}' , "
                f"BOTTOM = {self.sill} HEIGHT = {self.height}, WIDTH = {self.width}\n"
                f"      FACE = '{self.face.upper()}'  OFFSET = {self.offset} /\n")


class MaterialComposant:
    def __init__(self, thk, num, material, conductivity, specific_heat, density, emissivity,
                 tfs=0, fsp=0, cdf='null.txt', hc=0, comments='', ys=0, yco=0, yh=0, cf=1, comments2=''):
        self.description = material
        self.thickness = thk  # [mm]
        self.conductivity = conductivity  # [W/mK]
        self.specific_heat = specific_heat  # [J/kgK]
        self.density = density  # [kg/m3]
        self.emissivity = emissivity  # []
        self.cone_file = cdf  # [.txt]
        self.min_temp_spread = tfs  # [C]
        self.flame_spread_parameter = fsp  # []
        self.eff_heat_of_combustion = hc  # [kJ/kg]
        self.soot_yield = ys  # [g/g]
        self.CO2_yield = yco  # [g/g]
        self.H20_yield = yh  # [g/g]
        self.HCN_yield = 0  # [g/g]
        self.id = num
        self.comments = comments
        self.comments2 = comments2
        self.calibration_factor = cf

    @classmethod
    def from_thermal(cls, thk, row):
        return cls(thk=thk, num=row.iloc[0], material=row.iloc[1], conductivity=row.iloc[2],
                   specific_heat=row.iloc[3], density=row.iloc[4], emissivity=row.iloc[5], tfs=row.iloc[6],
                   fsp=row.iloc[7], cdf=row.iloc[8], hc=row.iloc[9], comments=row.iloc[10], ys=row.iloc[11],
                   yco=row.iloc[12], yh=row.iloc[13], cf=row.iloc[14], comments2=row.iloc[15])

    def __str__(self):
        return f"Material {self.id}: {self.description} d'épaisseur {self.thickness}"

    def thermal_absorptivity(self):
        return float(sqrt(self.conductivity * self.specific_heat * self.density))

    def return_element(self, ty):
        substrate = Etr.Element(ty, present='True')
        Etr.SubElement(substrate, 'description').text = self.description
        Etr.SubElement(substrate, 'thickness').text = str(self.thickness)
        Etr.SubElement(substrate, 'conductivity').text = str(self.conductivity)
        Etr.SubElement(substrate, 'specific_heat').text = str(self.specific_heat)
        Etr.SubElement(substrate, 'density').text = str(self.density)
        return substrate

    # TODO replace by is_in_words, words in Environnement_Brisk ?
    def is_wood(self) -> bool:
        mots = ["wood", "CLT", "oak", "pine", "timber"]
        for mot in mots:
            if re.search(mot, self.description, re.IGNORECASE):
                return True
        return False

    def is_protection(self) -> bool:
        mots = ['board', 'GYP']
        for mot in mots:
            if re.search(mot, self.description, re.IGNORECASE):
                return True
        return False

    def wood_thick(self) -> float:
        if self.is_wood():
            return self.thickness
        else:
            return 0

    def wood_density(self) -> float:
        if self.is_wood():
            return self.density
        else:
            return 0

    def replace_brisk_mat(self, root, section):
        def replace(st, val):
            for elem in root.findall(st):
                elem.text = str(val)

        replace(section + '/description', self.description)
        replace(section + '/thickness', self.thickness)
        replace(section + '/conductivity', self.conductivity)
        replace(section + '/specific_heat', self.specific_heat)
        replace(section + '/density', self.density)
        replace(section + '/emissivity', self.emissivity)
        replace(section + '/cone_file', self.cone_file)
        replace(section + '/min_temp_spread', self.min_temp_spread)
        replace(section + '/flame_spread_parameter', self.flame_spread_parameter)
        replace(section + '/eff_heat_of_combustion', self.eff_heat_of_combustion)
        replace(section + '/soot_yield', self.soot_yield)
        replace(section + '/CO2_yield', self.CO2_yield)
        replace(section + '/H20_yield', self.H20_yield)
        replace(section + '/HCN_yield', self.HCN_yield)

    # noinspection SpellCheckingInspection
    def cfast_id(self, loc, typo):
        return (f"&MATL ID = '{loc + typo}' MATERIAL = '{self.description}',\n"
                f"      CONDUCTIVITY = {self.conductivity} DENSITY = {self.density} SPECIFIC_HEAT = "
                f"{self.specific_heat / 1000}, THICKNESS = {self.thickness / 1000} EMISSIVITY = {self.emissivity} /\n")


class Material:
    def __init__(self, location, lining, substrate1=None, substrate2=None):
        self.location = location
        self.lining: MaterialComposant = lining
        self.substrate1: MaterialComposant = substrate1
        self.substrate2: MaterialComposant = substrate2
        self.wood_consumed = False

    def __str__(self):
        if self.substrate2 is not None:
            return f"{self.location} lining " + str(self.lining) + " + " + str(self.substrate1) + " + " \
                + str(self.substrate2)
        else:
            if self.substrate1 is not None:
                return f"{self.location} lining " + str(self.lining) + " + " + str(self.substrate1)
            else:
                return f"{self.location} lining " + str(self.lining)

    # def return_element(self):
    #     ty = str(self.location)
    #     lining = Etr.Element(ty + '_lining')
    #     Etr.SubElement(lining, 'description').text = self.description
    #     Etr.SubElement(lining, 'thickness').text = str(self.thickness)
    #     Etr.SubElement(lining, 'conductivity').text = str(self.conductivity)
    #     Etr.SubElement(lining, 'specific_heat').text = str(self.specific_heat)
    #     Etr.SubElement(lining, 'density').text = str(self.density)
    #     if self.substrate1 is not None:
    #         Etr.SubElement(lining, self.substrate1(ty))

    def thermal_absorptivity(self, tm):
        # prEN 1991-1-2:2021(E) A.3(3)
        b1 = self.lining.thermal_absorptivity()
        s1 = self.lining.thickness
        if self.substrate1 is not None:
            b2 = self.substrate1.thermal_absorptivity()
            if b1 < b2:
                return b1
            else:
                slim = Ec.s_lim(self.lining.density, self.lining.specific_heat, self.lining.conductivity, tm)
                if s1 > slim:
                    return b1
                else:
                    return s1 * b1 / slim + (1 - s1 / slim) * b2
        else:
            return b1

    def is_wood_protected(self) -> bool:
        # if self.substrate1 is not None:
        if self.lining.is_protection():
            if (self.substrate1 is not None and self.substrate1.is_wood()) or \
                    (self.substrate2 is not None and self.substrate2.is_wood()):
                return True
        return False

    def wood_thick(self) -> float:
        if self.substrate1 is not None:
            s1 = self.substrate1.wood_thick()
        else:
            s1 = 0
        if self.substrate2 is not None:
            s2 = self.substrate2.wood_thick()
        else:
            s2 = 0
        return self.lining.wood_thick() + s1 + s2

    def wood_density(self) -> float:
        if self.substrate1 is not None:
            d1 = self.substrate1.wood_density()
        else:
            d1 = 0
        if self.substrate2 is not None:
            d2 = self.substrate2.wood_density()
        else:
            d2 = 0
        d = max(self.lining.wood_density(), d1, d2)
        if d == 0:
            d = EnvB.default_wood_density
        return d

    def write_brisk_mat(self, root):
        section = './/rooms/room/' + str(self.location) + '_lining'
        self.lining.replace_brisk_mat(root, section)
        substrate_section = './/rooms/room/' + str(self.location) + '_substrate'
        # Trouver et remplacer la section _substrate
        for elem in root.findall(substrate_section):
            root.find('.//rooms/room').remove(elem)  # Supprimer l'ancienne section
            break  # Sortir après la première occurrence
        # Insérer la nouvelle section après lining
        parent = root.find('.//rooms/room')  # Obtenir le parent de lining
        lining = root.find(section)
        index = list(parent).index(lining)  # Trouver l'index de wall_lining dans son parent
        if self.substrate1 is not None:
            nouvelle_section = self.substrate1.return_element(str(self.location) + '_substrate')
        else:
            nouvelle_section = Etr.Element(str(self.location) + '_substrate', present='False')
        parent.insert(index + 1, nouvelle_section)  # Insérer la nouvelle section

    # noinspection SpellCheckingInspection
    def cfast_mat(self, loc):
        mat_id = [self.lining.cfast_id(loc, '_lining')]
        if self.substrate1 is not None:
            mat_id.append(self.substrate1.cfast_id(loc, '_s1'))
            if self.substrate2 is not None:
                mat_id.append(self.substrate2.cfast_id(loc, '_s2'))
                descr = (f"{loc}_MATL_ID = '{loc}_lining', '{loc}_s1', '{loc}_s2' "
                         f"{loc}_THICKNESS = {self.lining.thickness / 1000}, {self.substrate1.thickness / 1000}, "
                         f"{self.substrate2.thickness / 1000}")
            else:
                descr = (f"{loc}_MATL_ID = '{loc}_lining', '{loc}_s1' "
                         f"{loc}_THICKNESS = {self.lining.thickness / 1000}, {self.substrate1.thickness / 1000}")
        else:
            descr = f"{loc}_MATL_ID = '{loc}_lining' {loc}_THICKNESS = {self.lining.thickness / 1000}"
        return mat_id, descr


class Compartment:
    def __init__(self, nom, origin=(0.0, 0.0, 0.0), length=0.0, width=0.0, height=0.0,
                 hrr_fuel=HrrFunction([0, 600], [0, 0]), wall=0, ceiling=0, column=0, beam=0, floor=0,
                 max_thickness_char_ceiling=150, max_thickness_char_wall=150, interior=17, exterior=17, humid=0.5,
                 time_step=1, t_char_fin=240, beam_thick=0, column_thick=0, wwt=0, cwt=0, fwt=0,
                 wwd=0, cwd=0, fwd=0, bwd=0, co_wd=0, error_control=0.1, error_vent=1e-3,
                 excel_interval=EnvB.default_excel_interval, pressure=EnvB.default_pressure, fire=None):
        # DO créer les méthodes (et modifier celles d'initialisation) pour mettre à jour les max_thickness
        #  en fonction des épaisseurs des parois dans le compartiment
        self.id = nom  # ''
        self.origin = origin  # ([m],[m],[m])
        self.length = length  # [m]
        self.width = width  # [m]
        self.height = height  # [m]
        if isinstance(wall, (int, float)):
            self.a_wall_exposed = wall  # [m²]
        else:
            self.a_wall_exposed = 0
        if isinstance(ceiling, (int, float)):
            self.a_ceiling_exposed = ceiling  # [m²]
        else:
            self.a_ceiling_exposed = 0
        if isinstance(column, (int, float)):
            self.a_column_exposed = column  # [m²]
        else:
            self.a_column_exposed = 0
        if isinstance(beam, (int, float)):
            self.a_beam_exposed = beam  # [m²]
        else:
            self.a_beam_exposed = 0
        if isinstance(floor, (int, float)):
            self.a_floor_exposed = floor  # [m²]
        else:
            self.a_floor_exposed = 0
        self.hrr_fuel = hrr_fuel  # Hrr Function ([s],[kW])
        # self.exposed_time = exposed_time # [s] remplacé par une methode lisant hrr_fuel
        self.max_thickness_char_ceiling = max_thickness_char_ceiling  # [mm]
        self.max_thickness_char_wall = max_thickness_char_wall  # [mm]
        self.vents = []
        self.wall_mat = None
        self.wall_protected = False
        self.wall_time_fo = 0
        self.ceiling_mat = None
        self.ceiling_protected = False
        self.ceiling_time_fo = 0
        self.floor_mat = None
        self.floor_protected = False
        self.floor_time_fo = 0
        self.temp_interior = 273 + interior + 1  # [K] Brisk, make Text and Tint slightly different (Collen Wade 10/24)
        self.temp_exterior = 273 + exterior  # [K]
        self.rel_humidity = humid  # [%]
        self.time_step = time_step  # [s]
        self.error_control = error_control  # []
        self.error_vent_control = error_vent  # []
        self.excel_interval = excel_interval  # []
        self.time_char_fin = t_char_fin  # [s]
        self.beam_thick = beam_thick * 1000  # [mm]
        self.column_thick = column_thick * 1000  # [mm]
        if fire is None:
            self.fire = Fire(0, self.id, self.origin[0] + self.length / 2, self.origin[1] + self.width / 2,
                             self.origin[2])
            # self.heat_of_combustion = Pine.heat_of_combustion  # 12.4 [kJ/g] -> Fire.Fuel
        else:
            self.fire = fire
            # self.heat_of_combustion = self.fire.fuel.heat_of_combustion -> Fire.Fuel
        self.wall_wood_thick = wwt  # [mm]
        self.ceiling_wood_thick = cwt  # [mm]
        self.floor_wood_thick = fwt  # [mm]
        self.wall_wood_density = int(wwd)  # [kg/m3]
        self.ceiling_wood_density = int(cwd)  # [kg/m3]
        self.floor_wood_density = int(fwd)  # [kg/m3]
        if bwd != 0:
            self.beam_wood_density = int(bwd)  # [kg/m3]
        else:
            self.beam_wood_density = EnvB.default_wood_density
        if co_wd != 0:
            self.column_wood_density = int(co_wd)  # [kg/m3]
        else:
            self.column_wood_density = EnvB.default_wood_density
        # DO column & beam density & material
        self.wood_consumed = False
        self.auto_extinction = True
        self.lining_involved_only = False
        self.floor_bloc = 0
        self.a_f = self.length * self.width
        self.a_wall = 2 * (self.length + self.width) * self.height
        self.a_t = 2 * self.a_f + self.a_wall
        self.a_v = 0  # sum((item.width * item.height) for item in self.vents)
        self.h_v = 0  # sum((item.height * item.width * item.height) for item in self.vents) / a_v
        self.opening_factor = 0  # a_v * sqrt(h_v) / a_t
        self.calculated_ok = '1'  # Not calculated or error: 1 calculated and converge 0 see output1.xml flag stop
        self.init_pressure = pressure
        self.warning = ''
        self.Gamma = 0  # EC1-2 A.3(1) Γ = [O/b]**2/(0,04/1 160)**2 []
        self.thermal_absorptivity = 0  # EC1-2 A.3(3)  [J/m2s1/2K]
        self.t_max = 0  # EC1-2 A.3(5) (A.7)
        self.t_end = 0  # EC1-2 A.3(5) (A.7) from Background
        #                 Brandon Parametric fire design – zero-strength- layers and charring rates (9) to (11)
        self.parametric_fire_control = ''  # 'Fuel' or 'Ventilation' EC1-2 A.3(5) NOTE 1
        self.t_constant_char = 0  # t0 EC5-1-2 (A.11)
        self.beta_par_surf = self.beta_par_lin = 0  # EC5-1-2 (A.9)
        self.d_sect_eff_surf = self.d_sect_eff_lin = 0  # EC5-1-2 (A.14)
        # or Brandon Fire Safety Challenges of Tall Wood Buildings (8)

    def __str__(self):
        # écriture en Bleu
        st = f"Compartment \033[1;34m{self.id}\033[0m at origin {self.origin} "  # noqa: E231, E702
        return st + f"with dimensions {self.length}x{self.width}x" \
                    f"{self.height} and {self.a_wall_exposed}m², {self.a_ceiling_exposed}m², " \
                    f"{self.a_beam_exposed}m², {self.a_column_exposed}m², {self.a_floor_exposed}m², " \
                    f"{self.a_floor_exposed}m² of wall ceiling beam column and floor exposed, HRR Fuel: " \
                    f"{self.hrr_fuel.abscissas[-1]} sec, {len(self.vents)} openings, " \
                    f"opening factor: {u_round(self.opening_factor, '[m1/²]')} [m1/2]"

    def calculate_ec_curves_parameters(self, q_st=0):
        # Brandon Annex values (for test 1-5 i.e. NRC&NIST 5)
        # self.a_f = 4.6 * 9.1
        # self.a_t = 2 * self.a_f + 2 * (4.6 + 9.1) * 2.7
        # self.a_v = 1.8 * 2
        # self.opening_factor = float(self.a_v * sqrt(2) / self.a_t)
        # qd_tot = q_st + 550 * self.a_f / self.a_t
        # End Brandon Annex values
        qd_tot = q_st + self.fire.fd * self.a_f / self.a_t  # q𝑑,𝑡ot,𝑡 =⋅ q𝑑,𝑓i,𝑡 + q𝑑,𝑠𝑡,𝑡 (A.12)
        print(qd_tot)

        self.t_max = Ec.t_max(qd_tot, self.opening_factor)
        tm = self.t_max
        if tm == Ec.t_lim():
            self.parametric_fire_control = 'Fuel'
        else:
            self.parametric_fire_control = 'Ventilation'
        # bf = self.floor_mat.thermal_absorptivity(tm)
        # bc = self.ceiling_mat.thermal_absorptivity(tm)
        # bw = self.wall_mat.thermal_absorptivity(tm)
        # b = (bf * self.a_f + bc * self.a_f + bw * self.a_wall) / (self.a_t - self.a_v)
        # Brandon Annex values (for test 1-5 i.e. NRC&NIST 5)
        # b = 770
        b = 600
        self.thermal_absorptivity = b
        if EnvB.current_zone_model[1] == 'PC':
            if not (100 <= b <= 2200):
                if not re.search("absorptivity", self.warning, re.IGNORECASE):
                    self.warning = self.warning + f"EC1-2 A.3(1) thermal absorptivity b {round(b, 0)} out of range \n"
            if not (50 <= qd_tot <= 1000):
                if not re.search("qt,d", self.warning, re.IGNORECASE):
                    self.warning = self.warning + f"EC1-2 A.3(5) qt,d {round(qd_tot, 0)} out of range \n"
        self.Gamma = (self.opening_factor / self.thermal_absorptivity) ** 2 / (0.04 / 1160) ** 2  # []

    def add_parametric_temperature_time_curve(self, df, q_st=0):
        # Brandon Annex values (for test 1-5 i.e. NRC&NIST 5)
        # qd_tot = q_st + 550 * self.a_f / self.a_t
        qd_tot = q_st + self.fire.fd * self.a_f / self.a_t
        print(qd_tot)
        o = self.opening_factor
        b = self.thermal_absorptivity  # A.3(1) & (A.5)
        t_max = self.t_max / 60  # [h]
        t_lim = Ec.t_lim() / 60  # [h]
        # temp_max = Ec.heating_phase_temperature(t_max, self.Gamma)  # (A.6)
        # tm_star = (0.2e-3 * qd_tot / o) * self.Gamma
        if self.parametric_fire_control == 'Ventilation' or EnvB.parametric_curves == 'B':
            gamma = self.Gamma
        else:  # A.3(6) 'Fuel' ie t_max==t_lim
            o_lim = 0.1e-3 * qd_tot / t_lim
            if o > 0.04 and b < 1600 and qd_tot < 75:  # (A.10)
                k = 1 + ((o - 0.04) / 0.04) * ((qd_tot - 75) / 75) * ((1160 - b) / 1160)
            else:
                k = 1
            gamma = k * (o_lim / b) ** 2 / (0.04 / 1160) ** 2  # []
        temp_max = Ec.heating_phase_temperature(t_max, gamma)  # (A.6)
        tm_star = (0.2e-3 * qd_tot / o) * gamma
        if t_max > t_lim:
            x = 1
        else:
            # x = t_lim * self.Gamma / tm_star
            x = t_lim * gamma / tm_star
        self.t_end = Ec.cooling_phase_end(gamma, tm_star, temp_max, x) * 60
        df['Time (h)'] = df['Time (min)'] / 60
        df['Temp EC param curve (C)'] = df['Time (h)'].apply(lambda ti: Ec.heating_phase_temperature(
            ti, gamma) if ti <= t_max else Ec.cooling_phase_temperature(ti, gamma, tm_star, temp_max, x))

    def calculated(self):
        if self.calculated_ok == '0':
            return 'OK'
        elif self.calculated_ok == '1':
            return '/!\\ NO'
        else:
            return '?'

    def model_extinction(self):
        if self.wood_consumed:
            return 'wood consumed'
        else:
            if not self.auto_extinction:
                return f"iteration > {EnvB.max_iter}"
            else:
                return 'VRAI'

    def wood_thickness(self):
        if self.lining_involved_only:  # Exception DO modify code or excel for combustible lining (column s_cond)
            wwt = self.wall_wood_thick = self.wall_mat.lining.wood_thick()
            cwt = self.ceiling_wood_thick = self.ceiling_mat.lining.wood_thick()
            fwt = self.floor_wood_thick = self.floor_mat.lining.wood_thick()
        else:
            if self.floor_bloc != 0:
                wwt = self.wall_wood_thick = self.wall_mat.wood_thick()
                cwt = self.ceiling_wood_thick = self.ceiling_mat.wood_thick()
                fwt = self.floor_wood_thick = self.floor_bloc
            else:
                wwt = self.wall_wood_thick = self.wall_mat.wood_thick()
                cwt = self.ceiling_wood_thick = self.ceiling_mat.wood_thick()
                fwt = self.floor_wood_thick = self.floor_mat.wood_thick()
        return wwt, cwt, fwt

    def wood_zero_strength_layer_depth(self):
        wwt, cwt, fwt = self.wood_thickness()
        # TODO check all EC5-1-2 7.2.2 to 7.2.4 to insure exact calculation of zero strength layer
        #  for example when 7.2.3(10) separate calculation for each direction x & y is relevant?...
        #  implementation of side difference regarding CLT (not GL) orientation for beams & columns 7.2.2(10)...
        # ceiling: (7.15) d0=8+h/55 <=10
        d0c = min((8 + cwt / 55), 10)
        # floor: (7.14) d0=9+h/20 <=14
        d0f = min((9 + fwt / 20), 14)
        # wall: (7.14) d0=9+h/20 <=14
        d0w = min((9 + wwt / 20), 14)
        # Beam 7.2.2(9)
        d0b = 10
        # Column 7.2.2(8)
        d0co = 14
        return d0w, d0c, d0f, d0b, d0co

    def wood_density(self):
        if self.wall_wood_density == 0:
            wwd = self.wall_wood_density = self.wall_mat.wood_density()
        else:
            wwd = self.wall_wood_density
        if self.ceiling_wood_density == 0:
            cwd = self.ceiling_wood_density = self.ceiling_mat.wood_density()
        else:
            cwd = self.ceiling_wood_density
        if self.floor_wood_density == 0:
            fwd = self.floor_wood_density = self.floor_mat.wood_density()
        else:
            fwd = self.floor_wood_density
        return wwd, cwd, fwd, self.beam_wood_density, self.column_wood_density

    def exposed_time(self):
        return float(self.hrr_fuel.abscissas[-1])  # [s]

    def calc_opening_factor(self):
        # a_f = self.length * self.width
        # a_t = 2 * a_f + 2 * (self.length + self.width) * self.height
        # a_v = sum((item.width * item.height) for item in self.vents)
        # h_v = sum((item.height * item.width * item.height) for item in self.vents) / a_v
        # return a_v * sqrt(h_v) / a_t
        self.a_f = self.length * self.width
        self.a_t = 2 * self.a_f + 2 * (self.length + self.width) * self.height
        self.a_v = sum((item.width * item.height) for item in self.vents)
        self.h_v = sum((item.height * item.width * item.height) for item in self.vents) / self.a_v
        self.opening_factor = float(self.a_v * sqrt(self.h_v) / self.a_t)
        if EnvB.current_zone_model[1] == 'PC' and self.opening_factor > 0.1:
            if self.opening_factor <= 0.2:
                if not re.search("opening", self.warning, re.IGNORECASE):
                    self.warning = self.warning + ("EC1-2 A.3(5) O, opening factor "
                                                   "0,10<=O<=0,20 set to 0.1 see EC5-1-2 A.4.4.1(5) \n")
            else:
                if not re.search("opening", self.warning, re.IGNORECASE):
                    self.warning = self.warning + ("EC1-2 A.3(5) O, opening factor out of range, EC5 Annex A"
                                                   "O>0,20 (set to 0.1) see EC5-1-2 A.4.4.1(5) \n")
                self.calculated_ok = '1'
            self.opening_factor = 0.1

    def set_fuel(self, ft, hc=17.5):
        if ft is not None and ft == 'Propane':
            self.fire.fuel = Propane.copy()
        else:
            if EnvB.current_zone_model[1] == 'Brisk':
                # DO check other method (than hc Pine = 12.4 instead of 17.5 [MJ/m²] )
                #  to have conservative value (temperature) with BRisk ?
                # 250203: Don't change 12.4 is the value used by Brisk for crib (and 47.3 for Propane
                # see dbase/fire.mdb
                self.fire.fuel = Pine_Brisk.copy()
            else:
                # self.fire.fuel = Pine.copy()
                self.fire.fuel.heat_of_combustion = u_round(hc, '[kJ/g]')

    def calculate_brandon_hrr_fuel(self, f_d, t_fin, v_wind, d_wind,
                                   fd_factor: float = 1, hrr_factor: float = 1) -> bool:
        def concat_list(g, dev, dec):
            if not dev.size == 0:
                dev = np.delete(dev, 0)
            if not dec.size == 0:
                dec = np.delete(dec, 0)
            return np.concatenate((g, dev, dec))

        def factor(v_dir):
            #  w_dir = np.pi/180 * d_wind
            #  return np.cos(w_dir)
            # include factor Cpe,1 according to EN 1991-1-4 Table 7.1 with h/d=5 (maximisation of wind effect)
            if v_dir >= 315 or v_dir <= 45:
                return 1  # Case D of Table 7.1
            else:
                if 135 <= v_dir <= 225:
                    return -0.7  # Case E of Table 7.1
                else:
                    return -1.4  # Case A of Table 7.1

        def cherche_valeur(equation, variable, valeurs_initiales):
            def ok(sol):
                if abs(sol) < max_sol:
                    return True
                else:
                    return False

            tol = 1e-6
            max_sol = 1e7
            for xd in valeurs_initiales:
                solution = sy.nsolve(equation, variable, xd, verify=False)
                if solution.is_real:
                    if ok(solution):
                        return float(solution)
                else:
                    real_part, image_part = solution.as_real_imag()
                    if abs(image_part) < tol and ok(real_part):
                        return float(real_part)
                    solution = sy.nsolve(equation, variable, equation.subs(variable, xd).evalf(), verify=False)
                    if solution.is_real:
                        if ok(solution):
                            return float(solution)
                    else:
                        real_part, image_part = solution.as_real_imag()
                        if abs(image_part) < tol and ok(real_part):
                            return float(real_part)
            return np.nan

        y = 0
        x = 1
        alpha_f0 = 0
        # new_fire = Fire(f_d)
        self.fire.fd = f_d * fd_factor
        a_f = self.length * self.width
        a_v = sum((item.width * item.height) for item in self.vents)
        h_v = sum((item.height * item.width * item.height) for item in self.vents) / a_v
        # alpha = EnvB.alphas
        alpha0 = 0.047
        alpha1 = 0.4
        alpha4 = 0.1
        alpha5 = 0.8
        alpha6 = 0.5
        self.fire.hrr_pua = u_round(320 * hrr_factor, '[kW/m²]')  # [kW/m²] Brandon Rise Report 2021:63 Annex B
        q_cf = self.fire.hrr_pua * a_f
        q_c_max_tot = alpha1 * 3.01E+06 * a_v * sqrt(h_v) * (1 + alpha4) / 1000
        q_peak = min(q_cf, q_c_max_tot)

        # include Wind Rèf:
        #   Wind effect on internal and external compartment fire exposure
        #      Daniel Brandon Johan Anderson  RISE Report 2018:72
        if EnvB.allow_wind:
            w_speed = v_wind
            wind_factor = 1 + 0.47 * w_speed * factor(d_wind) / (1.7 * sqrt(9.81 * self.h_v))
            self.fire.wind[2] = u_round(wind_factor, '[]')
            q_peak = q_peak * wind_factor

        t_g = float(sqrt(q_peak / alpha0))
        f_tot = a_f * self.fire.fd
        int_0_t_dec = f_tot * alpha5 * alpha6
        int_0_tg = alpha0 * t_g ** 3 / 3
        t_dec = t_g + (int_0_t_dec * 1000 - int_0_tg) / q_peak
        q_dec = q_peak

        try:
            if t_fin < t_g:
                t_g = t_dec = t_fin
            else:
                if t_fin <= t_dec:
                    t_dec = t_fin
                else:
                    if t_dec <= t_g:
                        t_dec = t_g
                    else:
                        yc = sy.Symbol('yc', real=True)
                        equation2 = (t_dec - yc) * (sy.log(t_fin - yc) - sy.log(t_dec - yc)) - (
                                alpha5 * (1 - alpha6) * f_tot * 1000 / q_peak)
                        y = cherche_valeur(equation2, yc, [0, 1500, -1500, 750, -750])
                        t_fin_temp = t_fin
                        for i in range(1, 11):
                            if not np.isnan(y):
                                break
                            if t_fin_temp <= t_dec:
                                t_fin_temp = 2 * t_dec
                            else:
                                t_fin_temp = t_dec + 2 * (t_fin_temp - t_dec)
                            equation2 = ((t_dec - yc) * (sy.log(t_fin_temp - yc) - sy.log(t_dec - yc)) -
                                         (alpha5 * (1 - alpha6) * f_tot * 1000 / q_peak))
                            y = cherche_valeur(equation2, yc, [0, 1500, -1500, 750, -750])
                        if not np.isnan(y):
                            x = 1 / (q_dec * (t_dec - y))
                        else:
                            x = 0
            t_g = u_round(t_g, '[s]')
            t_dec = u_round(t_dec, '[s]')
            t_fin = u_round(t_fin, '[s]')

            if EnvB.current_zone_model[1] == 'Cfast':
                inc = t_fin / EnvB.default_hrr_interval
                n_g = max(int(t_g // inc), 1)
                n_d = max(int((t_dec - t_g) // inc), 1)
                n_f = max(int((t_fin - t_dec) // inc), 1)
            else:
                n_g = n_d = n_f = EnvB.default_hrr_interval
            a_g = np.linspace(0, t_g, n_g)
            a_g = np.vectorize(u_round)(a_g, '[s]')
            a_g = np.array(a_g, float)
            o_g = [u_round(alpha0 * t ** 2, '[kW]') for t in a_g]
            o_g = np.array(o_g, float)
            if t_dec > t_g:
                a_dev = np.linspace(t_g, t_dec, n_d)
                a_dev = np.vectorize(u_round)(a_dev, '[s]')
                a_dev = np.array(a_dev, float)
                o_dev = [u_round(alpha_f0 * (t_dec - t) ** 2 + q_dec, '[kW]') for t in a_dev]
                o_dev = np.array(o_dev, float)
                # continuité des fonctions :
                o_dev[0] = o_g[-1]
            else:
                a_dev = np.array([])
                o_dev = np.array([])
            if t_fin > t_dec:
                a_dec = np.linspace(t_dec, t_fin, n_f)
                a_dec = np.vectorize(u_round)(a_dec, '[s]')
                a_dec = np.array(a_dec, float)
                if not np.isnan(y):
                    o_dec = [u_round(1 / (x * (t - y)), '[kW]') for t in a_dec]
                else:
                    o_dec = [u_round(q_dec * (t - t_fin) / (t_dec - t_fin), '[kW]') for t in a_dec]
                o_dec = np.array(o_dec, float)
                # continuité des fonctions :
                if t_dec > t_g:
                    o_dec[0] = o_dev[-1]
                else:
                    o_dec[0] = o_g[-1]
            else:
                a_dec = np.array([])
                o_dec = np.array([])
            abscissas = concat_list(a_g, a_dev, a_dec)
            ordinates = concat_list(o_g, o_dev, o_dec)  # * hrr_factor
            self.hrr_fuel = HrrFunction(abscissas, ordinates)
            self.fire.fd = u_round(self.hrr_fuel.integral_total() / alpha5 / a_f / 1000, '[MJ/m²]')

        except Exception as e:
            print(f"Erreur lors de la création de Hrr Fuel: {e}")
            return False
        return True

    def calculate_hrr_fuel(self, f_d, t_fin, v_wind, d_wind, fd_factor: float = 1, hrr_pua_factor: float = 1) -> bool:
        def concat_list(g, dev, dec):
            if not dev.size == 0:
                dev = np.delete(dev, 0)
            if not dec.size == 0:
                dec = np.delete(dec, 0)
            return np.concatenate((g, dev, dec))

        def factor(v_dir):
            #  w_dir = np.pi/180 * d_wind
            #  return np.cos(w_dir)
            # include factor Cpe,1 according to EN 1991-1-4 Table 7.1 with h/d=5 (maximisation of wind effect)
            if v_dir >= 315 or v_dir <= 45:
                return 1  # Case D of Table 7.1
            else:
                if 135 <= v_dir <= 225:
                    return -0.7  # Case E of Table 7.1
                else:
                    return -1.4  # Case A of Table 7.1

        def cherche_valeur(equation, variable, valeurs_initiales):
            def ok(sol):
                if abs(sol) < max_sol:
                    return True
                else:
                    return False

            tol = 1e-6
            max_sol = 1e7
            for xd in valeurs_initiales:
                solution = sy.nsolve(equation, variable, xd, verify=False)
                if solution.is_real:
                    if ok(solution):
                        return float(solution)
                else:
                    real_part, image_part = solution.as_real_imag()
                    if abs(image_part) < tol and ok(real_part):
                        return float(real_part)
                    solution = sy.nsolve(equation, variable, equation.subs(variable, xd).evalf(), verify=False)
                    if solution.is_real:
                        if ok(solution):
                            return float(solution)
                    else:
                        real_part, image_part = solution.as_real_imag()
                        if abs(image_part) < tol and ok(real_part):
                            return float(real_part)
            return np.nan

        y = 0
        x = 1
        alpha_f0 = 0
        # new_fire = Fire(f_d)
        self.fire.fd = f_d * fd_factor
        a_f = self.length * self.width
        a_v = sum((item.width * item.height) for item in self.vents)
        h_v = sum((item.height * item.width * item.height) for item in self.vents) / a_v
        alpha = EnvB.alphas
        self.fire.hrr_pua = self.fire.hrr_pua * hrr_pua_factor
        q_c = self.fire.hrr_pua * a_f
        q_v = 1205 * a_v * sqrt(h_v)

        # Modif ContribBrisk2 suivant calcul HRR 241028_predictive_Timber_charring_V5
        # q_c_max_int = alpha[1] * alpha[2] * a_v * sqrt(h_v) / 1000
        # q_c_max_tot = q_c_max_int * (1 + alpha[4])
        # et email L.Girompaire du 28/10/2024
        # q_peak = min(q_c, q_c_max_tot)
        # if q_c > q_c_max_tot:
        #     # controlled = 'Ventilation'
        #     q_dec = q_v
        # else:
        #     # controlled = 'Fuel'
        #     q_dec = 0.67 * q_c
        if q_c > q_v:
            # controlled = 'Ventilation'
            q_peak = 1.5 * q_v
            q_dec = q_v
        else:
            # controlled = 'Fuel'
            q_dec = 0.67 * q_c
            q_peak = q_c

        # include Wind Rèf:
        #   Wind effect on internal and external compartment fire exposure
        #      Daniel Brandon Johan Anderson  RISE Report 2018:72
        if EnvB.allow_wind:
            w_speed = v_wind
            wind_factor = 1 + 0.47 * w_speed * factor(d_wind) / (1.7 * sqrt(9.81 * self.h_v))
            self.fire.wind[2] = u_round(wind_factor, '[]')
            q_peak = q_peak * wind_factor

        t_g = float(sqrt(q_peak / alpha[0]))
        f_tot = a_f * self.fire.fd
        int_0_t_dec = f_tot * alpha[5] * alpha[6]
        int_0_tg = alpha[0] * t_g ** 3 / 3

        try:
            # Définir la variable inconnue
            tdc = sy.Symbol('tdc', real=True)
            # Équation
            equation1 = ((tdc ** 3 / 3 - tdc ** 2 * t_g + tdc * t_g ** 2 - t_g ** 3 / 3) * (q_peak - q_dec) /
                         ((tdc - t_g) ** 2) + q_dec * (tdc - t_g) + int_0_tg) - int_0_t_dec * 1000
            # Résoudre l'équation
            # test1 = sy.solve(equation1, tdc)[0]
            # test = sy.#nsolve(equation1, (t_fin - t_g) / 2, verify=False)
            # if test.is_real:
            #     t_dec = float(test)
            # else:
            #     t_dec = float(test.as_real_#imag()[0])
            # # int_tg_t_dec = q_peak * (t_dec - t_g)
            t_dec = cherche_valeur(equation1, tdc, [(t_fin - t_g) / 2, (t_fin - t_g) / 8, 7 * (t_fin - t_g) / 8])
            if t_fin < t_g:
                t_g = t_dec = t_fin
            else:
                if t_fin <= t_dec:
                    t_dec = t_fin
                else:
                    if t_dec <= t_g:
                        t_dec = t_g
                    else:
                        yc = sy.Symbol('yc', real=True)
                        equation2 = (t_dec - yc) * (sy.log(t_fin - yc) - sy.log(t_dec - yc)) - (
                                alpha[5] * (1 - alpha[6]) * f_tot * 1000 / q_peak)
                        # # test = equation2.subs(yc, 1500).#evalf()
                        # test = sy.#nsolve(equation2, yc, 0, verify=False)
                        # if test.is_real:
                        #     y = float(test)
                        # else:
                        #     y = float(test.as_real_#imag()[0])
                        y = cherche_valeur(equation2, yc, [0, 1500, -1500, 750, -750])
                        t_fin_temp = t_fin
                        for i in range(1, 11):
                            if not np.isnan(y):
                                break
                            if t_fin_temp <= t_dec:
                                t_fin_temp = 2 * t_dec
                            else:
                                t_fin_temp = t_dec + 2 * (t_fin_temp - t_dec)
                            equation2 = ((t_dec - yc) * (sy.log(t_fin_temp - yc) - sy.log(t_dec - yc)) -
                                         (alpha[5] * (1 - alpha[6]) * f_tot * 1000 / q_peak))
                            y = cherche_valeur(equation2, yc, [0, 1500, -1500, 750, -750])
                        if not np.isnan(y):
                            x = 1 / (q_dec * (t_dec - y))
                        else:
                            x = 0
                        alpha_f0 = (q_peak - q_dec) / ((t_dec - t_g) ** 2)
            t_g = u_round(t_g, '[s]')
            t_dec = u_round(t_dec, '[s]')
            t_fin = u_round(t_fin, '[s]')
            # inc = 75
            # n_g = int(t_g // inc)
            # n_d = int((t_dec - t_g) // inc)
            # n_f = int((t_fin - t_dec) // inc)
            if EnvB.current_zone_model[1] == 'Cfast':
                inc = t_fin / EnvB.default_hrr_interval
                n_g = max(int(t_g // inc), 1)
                n_d = max(int((t_dec - t_g) // inc), 1)
                n_f = max(int((t_fin - t_dec) // inc), 1)
            else:
                n_g = n_d = n_f = EnvB.default_hrr_interval
            a_g = np.linspace(0, t_g, n_g)
            a_g = np.vectorize(u_round)(a_g, '[s]')
            a_g = np.array(a_g, float)
            o_g = [u_round(alpha[0] * t ** 2, '[kW]') for t in a_g]
            o_g = np.array(o_g, float)
            if t_dec > t_g:
                a_dev = np.linspace(t_g, t_dec, n_d)
                a_dev = np.vectorize(u_round)(a_dev, '[s]')
                a_dev = np.array(a_dev, float)
                o_dev = [u_round(alpha_f0 * (t_dec - t) ** 2 + q_dec, '[kW]') for t in a_dev]
                o_dev = np.array(o_dev, float)
                # continuité des fonctions :
                o_dev[0] = o_g[-1]
            else:
                a_dev = np.array([])
                o_dev = np.array([])
            if t_fin > t_dec:
                a_dec = np.linspace(t_dec, t_fin, n_f)
                a_dec = np.vectorize(u_round)(a_dec, '[s]')
                a_dec = np.array(a_dec, float)
                if not np.isnan(y):
                    o_dec = [u_round(1 / (x * (t - y)), '[kW]') for t in a_dec]
                else:
                    o_dec = [u_round(q_dec * (t - t_fin) / (t_dec - t_fin), '[kW]') for t in a_dec]
                o_dec = np.array(o_dec, float)
                # continuité des fonctions :
                if t_dec > t_g:
                    o_dec[0] = o_dev[-1]
                else:
                    o_dec[0] = o_g[-1]
            else:
                a_dec = np.array([])
                o_dec = np.array([])
            abscissas = concat_list(a_g, a_dev, a_dec)
            ordinates = concat_list(o_g, o_dev, o_dec)
            self.hrr_fuel = HrrFunction(abscissas, ordinates)

        except Exception as e:
            print(f"Erreur lors de la création de Hrr Fuel: {e}")
            return False
        return True

    # Not Used TODO to complete & correct (fire, heat-of_of combustion...)
    # noinspection SpellCheckingInspection
    """#  @classmethod
    # def from_brisk(cls, file_path):
    #     # Parse the XML file
    #     input_file = file_path + '\\' + 'input1.xml'
    #     item_file = file_path + '\\' + 'items.xml'
    #     tree = Etr.parse(input_file)
    #     root = tree.getroot()
    #
    #     # Find the room element
    #     room = root.find(".//room")
    #
    #     # Extract the values
    #     # nom = room.get("id")
    #     nom = lire_first_texte_arbre_xml(room, 'description')
    #     origin = (float(room.find("abs_X").text), float(room.find("abs_Y").text),
    #               float(room.find("floor_elevation").text))
    #     length = float(room.find("length").text)
    #     width = float(room.find("width").text)
    #     height = float(room.find("max_height").text)
    #     abscissa = lire_nombres_xml(input_file, 'time')
    #     ordinate = lire_nombres_xml(input_file, 'HRR')
    #     # f = TimeFunction(abscissa, ordinate)
    #     f_hrr = HrrFunction(abscissa, ordinate)
    #     hrr_fuel = f_hrr
    #     # max_time = f.abscissas[-1]
    #     heat_of_combustion = find_and_replace_value_in_xml(item_file, 'heat of combustion', None)
    #     # Create a Compartment object and return it
    #     return cls(nom=nom, origin=origin, length=length, width=width, height=height, hrr_fuel=hrr_fuel,
    #                heat_of_combustion=heat_of_combustion)
    #
    # @classmethod
    # def from_cfast(cls, text_file):
    #     nom = "vide"
    #     length = 0
    #     width = 0
    #     height = 0
    #     origin = (0, 0, 0)
    #     with open(text_file, 'r') as file:
    #         lines = file.readlines()
    #
    #     for line in lines:
    #         if line.startswith('&COMP'):
    #             data = line.split()
    #             nom = data[3].replace("'", "")
    #         elif line.startswith('      DEPTH'):
    #             data = line.split()
    #             length = float(data[data.index('DEPTH') + 2])
    #             width = float(data[data.index('WIDTH') + 2])
    #             height = float(data[data.index('HEIGHT') + 2])
    #         elif line.startswith('      ORIGIN'):
    #             data = line.split()
    #             origin = (float(data[data.index('ORIGIN') + 2].replace(",", "")),
    #                       float(data[data.index('ORIGIN') + 3].replace(",", "")),
    #                       float(data[data.index('ORIGIN') + 4]))
    #     return cls(nom=nom, origin=origin, length=length, width=width, height=height)
    #     # TO DO Read HHR Fuel & Heat of Combustion

    # def to_brisk(self, xml_file):  # TODO mettre à jour (compléter) la méthode en concaténant le reste de l'arbre xml
    #     # nécessaire pour un fichier input Brisk complet
    #     root = Etr.Element('room', id=self.id)
    #
    #     Etr.SubElement(root, 'width').text = str(self.width)
    #     Etr.SubElement(root, 'length').text = str(self.length)
    #     Etr.SubElement(root, 'max_height').text = str(self.height)
    #     Etr.SubElement(root, 'abs_X').text = str(self.origin[0])
    #     Etr.SubElement(root, 'abs_Y').text = str(self.origin[1])
    #     Etr.SubElement(root, 'floor_elevation').text = str(self.origin[2])
    #
    #     tree = Etr.ElementTree(root)
    #     tree.write(xml_file)
    """

    # noinspection SpellCheckingInspection
    def write_brisk_base_model(self):
        def replace(st, val):
            for item in root.findall(st):
                item.text = str(val)

        # noinspection SpellCheckingInspection
        def modif_vent(xml, cur_vent: Vent, ind):
            if ind == 0:
                modif = xml
            else:
                modif = copy.deepcopy(xml)
                # modif = Etr.Element('Vent')
                # modif.extend(xml)
                # for child in xml:
                #     new_child = Etr.SubElement(modif, child.tag)
                #     new_child.text = child.text
            tc = modif.find('id')
            tc.text = str(ind + 1)
            modif.find('description').text = self.id + '_O' + str(ind + 1)
            modif.find('offset').text = str(cur_vent.offset)
            modif.find('face').text = str(location.index(cur_vent.face))
            modif.find('sillheight').text = str(cur_vent.sill)
            if location.index(cur_vent.face) in [0, 3]:
                modif.find('walllength1').text = str(self.length)
            else:
                modif.find('walllength1').text = str(self.width)
            # DO Wath is walllength2 = 0 ? From CW:
            # The walllength inputs are used in the door mixing flows in multiroom models.
            # They correspond to the width of wall containing the vent for each of the rooms to which
            # the vent is connected. For a single room model, the walllength2 would typically just be set at 0.
            find_and_replace_value_varname(modif, 'height', cur_vent.height)
            find_and_replace_value_varname(modif, 'width', cur_vent.width)
            if ind > 0:
                root.append(modif)

        def modif_h_vent(xml, cur_vent: Vent, ind):
            if ind == 0:
                modif = xml
            else:
                modif = copy.deepcopy(xml)
                # modif = Etr.Element('Vent')
                # modif.extend(xml)
                # for child in xml:
                #     new_child = Etr.SubElement(modif, child.tag)
                #     new_child.text = child.text
            tc = modif.find('id')
            tc.text = str(ind + 1)
            modif.find('height').text = str(cur_vent.height)
            modif.find('width').text = str(cur_vent.width)
            modif.find('sill_height').text = str(cur_vent.sill)
            modif.find('offset').text = str(cur_vent.offset)
            modif.find('face').text = str(location.index(cur_vent.face))
            if location.index(cur_vent.face) in [0, 2]:
                modif.find('wall_length_1').text = str(self.length)
            else:
                modif.find('wall_length_1').text = str(self.width)
            if ind > 0:
                parent.append(modif)

        def modif_input(duration, interval):
            # tf = self.exposed_time()
            # dt = min(u_round(tf / 100, 's'), EnvB.default_excel_interval)
            replace('.//general_settings/simulation_duration', duration)
            replace('.//general_settings/display_interval', interval)
            # DO test précision sur temps calcul et résultats (peu d'impact)
            replace('.//general_settings/ceiling_nodes', EnvB.default_ceiling_nodes)
            replace('.//general_settings/wall_nodes', EnvB.default_wall_nodes)
            replace('.//general_settings/floor_nodes', EnvB.default_floor_nodes)
            replace('.//general_settings/excel_interval', interval)
            replace('.//general_settings/time_step', self.time_step)
            replace('.//general_settings/error_control', self.error_control)
            replace('.//general_settings/error_vent_control', self.error_vent_control)
            # noinspection SpellCheckingInspection
            replace('.//general_settings/error_control_ventflows', self.error_vent_control)
            replace('.//rooms/room/width', self.width)
            replace('.//rooms/room/length', self.length)
            replace('.//rooms/room/max_height', self.height)
            replace('.//rooms/room/description', self.id)
            replace('.//rooms/room/min_height', self.height)
            replace('.//rooms/room/floor_elevation', self.origin[2])
            replace('.//rooms/room/abs_X', self.origin[0])
            replace('.//rooms/room/abs_Y', self.origin[1])
            self.wall_mat.write_brisk_mat(root)
            self.ceiling_mat.write_brisk_mat(root)
            self.floor_mat.write_brisk_mat(root)
            # noinspection SpellCheckingInspection
            replace('.//chemistry/fueltype', self.fire.fuel.fuel_type)

        EnvB.ModelName = f"{self.id}"
        dest_dir = Files_utils.init_brisk_basemodel(EnvB.CurPath + r'\input\basemodel_default_gen')
        os.rename(dest_dir + '\\' + 'basemodel_default.xml', dest_dir + '\\' + f"{EnvB.BaseModel}0.xml")

        # Charger le fichier XML Base_model
        fichier_xml = dest_dir + '\\' + f"{EnvB.BaseModel}0.xml"
        test_open_write(fichier_xml)
        tree = Etr.parse(fichier_xml)
        root = tree.getroot()
        tf = self.exposed_time()
        dt = min(u_round(tf / 100, 's'), self.excel_interval)
        modif_input(tf, dt)
        # TODO Verif that output_interval as no impact on calculation & can be set egal to display & excel_interval
        replace('.//general_settings/output_interval', dt)
        replace('.//general_settings/base_name', f"{EnvB.BaseModel}0")
        # xnoinspection SpellCheckingInspection
        # replace('.//general_settings/error_control_ventflows', self.error_vent_control)
        tree.write(fichier_xml)

        # Changer le fichier XML vents
        fichier_xml = dest_dir + '\\' + 'vents.xml'
        test_open_write(fichier_xml)
        tree = Etr.parse(fichier_xml)
        root = tree.getroot()
        location = ['Front', 'Right', 'Rear', 'Left']

        # Trouver et supprimer le Vent avec id 2
        for vent in root.findall('Vent'):
            if vent.find('id').text == '2':
                root.remove(vent)
        # Trouver le Vent avec id 1, modifier ses champs et ajouter les autres vents
        for xml_vent in root.findall('Vent'):
            if xml_vent.find('id').text == '1':
                for index, elem in enumerate(self.vents):
                    modif_vent(xml_vent, elem, index)
        # Sauvegarder le fichier XML modifié
        tree.write(fichier_xml)

        # Changer le fichier XML items
        fichier_xml_entree = dest_dir + '\\' + 'items.xml'
        # NEW EC5: test_open_write(fichier_xml)
        test_open_write(fichier_xml_entree)
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
                               'description', self.id + '_hrr')
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
                               'detaileddescription', self.id + '_hrr')
        remplacer_double_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "hrr",
                                          self.hrr_fuel.abscissas, self.hrr_fuel.ordinates)
        # DO intégrer mass comme champs de Fire
        self.fire.mass_of_combustion = u_round(self.hrr_fuel.integral_total()
                                               / (self.fire.fuel.heat_of_combustion * 1000), '[kg]')
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'length', str(self.fire.dim_x))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'width', str(self.fire.dim_y))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'height', str(self.fire.dim_h))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'mass', str(self.fire.mass_of_combustion))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'xleft',
        #                        str(u_round(self.length / 2 - self.fire.dim_x, '[m]')))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'ybottom',
        #                        str(u_round(self.width / 2 - self.fire.dim_y, '[m]')))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'xleft',
                               str(u_round(self.fire.x - self.fire.dim_x, '[m]')))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'ybottom',
                               str(u_round(self.fire.y - self.fire.dim_y, '[m]')))
        # find_and_replace_item_varname(root, 'Interior Temperature', self.temp_interior, 'varvalue')
        find_and_replace_value_in_xml(fichier_xml_entree, 'Elevation', str(self.fire.elevation))
        # New EC5 TODO heat_of_combustion stored in Fire ?
        find_and_replace_value_in_xml(fichier_xml_entree, 'heat of combustion', str(self.fire.fuel.heat_of_combustion))

        # Changer le fichier input
        fichier_xml_entree = dest_dir + '\\' + 'input1.xml'
        test_open_write(fichier_xml_entree)
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'temp_interior', str(self.temp_interior))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'temp_exterior', str(self.temp_exterior))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'rel_humidity', str(self.rel_humidity))
        remplacer_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "time", self.hrr_fuel.abscissas)
        remplacer_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "HRR", self.hrr_fuel.ordinates)
        modif_fire_description_in_xml(fichier_xml_entree, self.id + '_hrr')
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'fire_height', str(self.fire.elevation))
        # New EC5
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'heat_of_combustion',
                               str(self.fire.fuel.heat_of_combustion))
        # self.fire.fuel.fuel_type = lire_str_dans_xml(fichier_xml_entree, 'fueltype')
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'fueltype',
                               str(self.fire.fuel.fuel_type))
        # TODO modification possible de la localisation du feu (<!-- fire location, corner=2, wall=1, centre=0 -->)
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'fire_location', str(0))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'data_points',
                               str(len(self.hrr_fuel.abscissas)))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_length', str(self.fire.dim_x))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_width', str(self.fire.dim_y))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_height', str(self.fire.dim_h))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_x',
                               str(u_round(self.length / 2 - self.fire.dim_x, '[m]')))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_y',
                               str(u_round(self.width / 2 - self.fire.dim_y, '[m]')))

        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'HRRUA', str(self.fire.hrr_pua))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'x1',
                               str(u_round(self.length / 2 - self.fire.dim_x, '[m]')))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'x2', str(self.length / 2))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'y1',
                               str(u_round(self.width / 2 - self.fire.dim_y, '[m]')))
        remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'y2', str(self.width / 2))
        # DO alphaT peakHRR: From CW:
        # The alphaT and peakHRR inputs are only used when you have selected the powerlaw  type of design fire e.g. t2 ,
        # not if you are manually entering the time vs HRR values.
        fichier_xml = fichier_xml_entree
        tree = Etr.parse(fichier_xml)
        root = tree.getroot()
        modif_input(tf, dt)
        # xnoinspection SpellCheckingInspection
        # replace('.//general_settings/error_control_ventflow', self.error_vent_control)
        # Trouver et supprimer le Vent avec id 2
        parent = root.find('.//hvents')  # Obtenir le parent de hvent
        # index = list(parent).index(wall_lining)  # Trouver l'index de wall_lining dans son parent
        for vent in root.findall('.//hvent'):
            if vent.find('id').text == '2':
                parent.remove(vent)
        # Trouver le Vent avec id 1, modifier ses champs et ajouter les autres vents
        for xml_vent in root.findall('.//hvent'):
            if xml_vent.find('id').text == '1':
                for index, elem in enumerate(self.vents):
                    modif_h_vent(xml_vent, elem, index)
        # Sauvegarder le fichier XML modifié
        tree.write(fichier_xml)

        # Changer le fichier room
        fichier_xml = dest_dir + '\\' + 'rooms.xml'
        test_open_write(fichier_xml)
        tree = Etr.parse(fichier_xml)
        root = tree.getroot()
        replace('.//room/room_length', self.length)
        replace('.//room/room_width', self.width)
        replace('.//room/room_minheight', self.height)
        replace('.//room/room_maxheight', self.height)
        replace('.//room/room_elevation', self.origin[2])
        replace('.//room/room_absx', self.origin[0])
        replace('.//room/room_absy', self.origin[1])
        replace('.//room/room_description', self.id)
        find_and_replace_value_varname(root, 'length', self.length)
        find_and_replace_value_varname(root, 'width', self.width)
        tree.write(fichier_xml)

        # Changer le fichier distribution
        fichier_xml = dest_dir + '\\' + 'distributions.xml'
        test_open_write(fichier_xml)
        tree = Etr.parse(fichier_xml)
        root = tree.getroot()
        find_and_replace_item_varname(root, 'Interior Temperature', self.temp_interior, 'varvalue')
        find_and_replace_item_varname(root, 'Exterior Temperature', self.temp_exterior, 'varvalue')
        find_and_replace_item_varname(root, 'Relative Humidity', self.rel_humidity, 'varvalue')
        find_and_replace_item_varname(root, 'Fire Load Energy Density', self.fire.fd, 'varvalue')
        tree.write(fichier_xml)

        print(f"Base model {EnvB.ModelName} écrit")

    # New EC5:
    def write_brisk_default_values(self):
        EnvB.ModelName = f"{self.id}"
        EnvB.ModelPaths = EnvB.CurPath + '\\' + EnvB.ModelName
        dest_dir = EnvB.ModelPaths + '\\' + f"{EnvB.BaseModel}0" + '\\' + f"{EnvB.BaseModel}0"

        # Changer le fichier XML items
        fichier_xml_entree = dest_dir + '\\' + 'items.xml'
        test_open_write(fichier_xml_entree)
        # TODO lire HRR ?
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
        #                        'description', self.id + '_hrr')
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
        #                        'detailed#description', self.id + '_hrr')
        # remplacer_double_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "hrr",
        #                                   self.hrr_fuel.abscissas, self.hrr_fuel.ordinates)
        # self.fire.mass_of_combustion = u_round(self.hrr_fuel.integral_total()
        #                                        / (self.heat_of_combustion * 1000), '[kg]')
        self.fire.dim_x = float(lire_str_dans_xml(fichier_xml_entree, 'length'))
        self.fire.dim_y = float(lire_str_dans_xml(fichier_xml_entree, 'width'))
        self.fire.dim_h = float(lire_str_dans_xml(fichier_xml_entree, 'height'))
        self.fire.mass_of_combustion = float(lire_str_dans_xml(fichier_xml_entree, 'mass'))
        self.fire.elevation = find_value_in_xml(fichier_xml_entree, 'Elevation')
        self.fire.fuel.heat_of_combustion = find_value_in_xml(fichier_xml_entree, 'heat of combustion')

        # Changer le fichier input
        fichier_xml_entree = dest_dir + '\\' + 'input1.xml'
        test_open_write(fichier_xml_entree)
        self.temp_interior = float(lire_str_dans_xml(fichier_xml_entree, 'height'))
        self.temp_exterior = float(lire_str_dans_xml(fichier_xml_entree, 'temp_exterior'))
        self.rel_humidity = float(lire_str_dans_xml(fichier_xml_entree, 'rel_humidity'))
        # remplacer_nombres_dans_xml (fichier_xml_entree, fichier_xml_entree, "time", self.hrr_fuel.abscissas)
        # remplacer_nombres_dans_xml (fichier_xml_entree, fichier_xml_entree, "HRR", self.hrr_fuel.ordinates)
        # modif_fire_description_in_xml (fichier_xml_entree, self.id + '_hrr')
        # noinspection SpellCheckingInspection
        self.fire.fuel.fuel_type = lire_str_dans_xml(fichier_xml_entree, 'fueltype')
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'fire_location', str(0))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'data_points',
        #                        str(len(self.hrr_fuel.abscissas)))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_length', str(self.fire.dim_x))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_width', str(self.fire.dim_y))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_height', str(self.fire.dim_h))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_x',
        #                        str(u_round(self.length / 2 - self.fire.dim_x, '[m]')))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'obj_y',
        #                        str(u_round(self.width / 2 - self.fire.dim_y, '[m]')))

        # noinspection SpellCheckingInspection
        self.fire.hrr_pua = float(lire_str_dans_xml(fichier_xml_entree, 'HRRUA'))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'x1',
        #                        str(u_round(self.length / 2 - self.fire.dim_x, '[m]')))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'x2', str(self.length / 2))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'y1',
        #                        str(u_round(self.width / 2 - self.fire.dim_y, '[m]')))
        # remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'y2', str(self.width / 2))

        self.excel_interval = float(lire_str_dans_xml(fichier_xml_entree, 'excel_interval'))
        self.time_step = float(lire_str_dans_xml(fichier_xml_entree, 'time_step'))
        self.error_control = float(lire_str_dans_xml(fichier_xml_entree, 'time_step'))

        print(f"Base model {EnvB.ModelName} lu")

    def to_cfast(self, dest_dir, n_iter):
        def head():
            return f"&HEAD VERSION = 7700, TITLE = 'CFAST Simulation {file_name}' /\n\n"

        # noinspection SpellCheckingInspection
        def scenario():
            tf = self.exposed_time()
            # dt = min(u_round(tf / 100, 's'), self.excel_interval)
            return (f"!! Scenario Configuration\n"
                    f"&TIME SIMULATION = {tf} PRINT = {self.excel_interval} SMOKEVIEW = {EnvB.default_smokeview_step} "
                    f"SPREADSHEET = {self.excel_interval} /\n"
                    f"&INIT PRESSURE = {self.init_pressure} RELATIVE_HUMIDITY = {self.rel_humidity * 100} "
                    f"INTERIOR_TEMPERATURE = {self.temp_interior - 273} "  # Temp in ° not in K for Cfast
                    f"EXTERIOR_TEMPERATURE = {self.temp_exterior - 273} /\n\n")

        # EnvB.ModelName = f"{self.id}"
        self.fire.id = self.id + '_' + EnvB.default_method + f"_iter{n_iter}_hrr"
        text_file = dest_dir + '\\' + f"{EnvB.BaseModel}{n_iter}.in"
        # Récupérer le chemin du fichier, son nom et son extension
        file_path = Path(text_file)
        file_name = file_path.stem  # Nom du fichier sans extension
        # File_extension = file_path.suffix # Extension du fichier
        # full_name = file_path.name # Nom complet avec l'extension
        # directory = file_path. parent # Dossier contenant le fichier
        file = open(text_file, 'w')
        # with open(text_file, 'w') as file:
        file.write(head())
        file.write(scenario())

        wall_mat_id, wall_mat_descr = self.wall_mat.cfast_mat('WALL')
        ceil_mat_id, ceil_mat_descr = self.ceiling_mat.cfast_mat('CEILING')
        floor_mat_id, floor_mat_descr = self.floor_mat.cfast_mat('FLOOR')
        file.write("!! Material Properties\n")
        for item in wall_mat_id:
            file.write(item)
        for item in ceil_mat_id:
            file.write(item)
        for item in floor_mat_id:
            file.write(item)

        file.write("\n!! Compartments \n")
        file.write(f"&COMP ID = '{self.id}'\n")
        # /!\ Cfast Depth = Brisk width (Y) & Cfast Width = Brisk length (X)
        file.write(f"      DEPTH = {self.width} HEIGHT = {self.height} WIDTH = {self.length}\n")
        mat = '      ' + ceil_mat_descr + ' ' + wall_mat_descr + ' ' + floor_mat_descr + f"\n"
        file.write(mat)
        file.write(f"      ORIGIN = {self.origin[0]}, {self.origin[1]}, {self.origin[2]} GRID = 50, 50, 50 "
                   f"LEAK_AREA_RATIO = {EnvB.default_leak_area_ratio}, {EnvB.default_leak_area_ratio} /\n")

        file.write("\n!! Wall vents \n")
        for elem in self.vents:
            file.write(elem.to_cfast())

        file.write("\n!! Fires \n")
        file.write(self.fire.to_cfast(self.id))
        file.write(self.hrr_fuel.to_cfast(self.fire.id, self.fire.elevation,
                                          self.fire.fuel.co_yield, self.fire.fuel.soot_yield))

        file.write("\n&TAIL /\n")
        file.close()

    # TODO 'flag stop'
    def to_lists(self):
        parametric_curves = EnvB.current_zone_model[1] == 'PC'
        # noinspection SpellCheckingInspection
        liste_champs = ['Experiment ', 'Directory ', 'Generate by program: ', 'Contribution Method', 'Full calculation',
                        'allow char energy storage (O2% reduction)' if EnvB.default_method == 'EC5' else '',
                        'protection fall off implemented',
                        'wind effect implemented' if not parametric_curves else 'char calculation', 'Calculation Time',
                        'length',
                        'width',
                        'height',
                        'wall + ceiling + floor area',
                        'opening area',
                        'opening factor',
                        'thermal absorbtivity (EC1-1-2 A.3)',
                        'gamma (EC1-1-2 A.3)',
                        'growth rate (EC1-1-2)',
                        'parametric curve control by (EC1-1-2):',
                        'exposed wall',
                        'exposed ceiling',
                        'exposed beam',
                        'exposed column',
                        'exposed floor',
                        'growth time',
                        'decay time',
                        'fully developed phase duration',
                        'decay phase duration',
                        'total calculation time',
                        'fire density',
                        'HRRPUA',
                        'initial fire load Ingberg equivalency time',
                        'final total fire load Ingberg equivalency time',
                        'heat of combustion',
                        'fuel type' if not parametric_curves else 't0 (time of constant char)',
                        'exterior wind' if not parametric_curves else '3 * t0',
                        'constant surface element char rate' if parametric_curves else '',
                        'constant linear element char rate' if parametric_curves else '',
                        'total surface element char  dchar,t (A.10)' if parametric_curves else '',
                        'total linear element char dchar,t (A.10)' if parametric_curves else '',
                        'surface effective cross-section  def (A.14)' if parametric_curves else '',
                        'linear effective cross-section  def (A.14)' if parametric_curves else '',
                        'calculation warning' if parametric_curves else '']
        tt = self.exposed_time() / 60
        parametric_method = ''
        if parametric_curves:
            t_max = self.t_max
            t_dec = t_max
            td = self.t_end - t_max
            if EnvB.parametric_curves == 'E':
                parametric_method = 'Eurocode'
            else:
                parametric_method = ("Fire Safety Challenges of Tall Wood Buildings – Phase 2: "
                                     "Task 4 Engineering Methods (D.Brandon)")
                EnvB.long_line.append(parametric_method)
        else:
            t_max = self.hrr_fuel.t_max / 60
            t_dec = self.hrr_fuel.t_dec / 60
            td = tt - t_dec
        tf = t_dec - t_max
        of = u_round(self.opening_factor, '[m1/2]')
        c_ok = self.calculated()
        liste_valeurs = [self.id, EnvB.ModelPaths, EnvB.ProgName, EnvB.default_method, c_ok,
                         EnvB.allow_char_energy_storage if EnvB.default_method == 'EC5' else '',
                         EnvB.allow_contribution_protected,
                         EnvB.allow_wind if not parametric_curves else parametric_method,
                         EnvB.CalculationTime, self.length, self.width, self.height, u_round(self.a_t, '[m²]'),
                         u_round(self.a_v, '[m²]'), of, u_round(self.thermal_absorptivity, '[J/m²Ks1/2]'),
                         u_round(self.Gamma, '[]'), Ec.fire_growth, self.parametric_fire_control, self.a_wall_exposed,
                         self.a_ceiling_exposed, self.a_beam_exposed, self.a_column_exposed, self.a_floor_exposed,
                         u_round(t_max, '[mn]'), u_round(t_dec, '[mn]'), u_round(tf, '[mn]'),
                         u_round(td, '[mn]'), u_round(tt, '[mn]'),
                         u_round(self.fire.fd, '[MJ/m²]'), u_round(self.fire.hrr_pua, '[kW/m²]'),
                         self.fire.fuel_ingberg_equivalency, self.fire.total_fire_ingberg_equivalency,
                         self.fire.fuel.heat_of_combustion,
                         self.fire.fuel.fuel_type if not parametric_curves else u_round(self.t_constant_char, '[mn]'),
                         str(self.fire.wind[0]) + ' ; ' + str(self.fire.wind[1]) + ' ; ' + str(self.fire.wind[2])
                         if not parametric_curves else u_round(3 * self.t_constant_char, '[mn]'),
                         u_round(self.beta_par_surf, '[mm/min]') if parametric_curves else '',
                         u_round(self.beta_par_lin, '[mm/min]') if parametric_curves else '',
                         u_round(2 * self.beta_par_surf * self.t_constant_char, '[mm]') if parametric_curves else '',
                         u_round(2 * self.beta_par_lin * self.t_constant_char, '[mm]') if parametric_curves else '',
                         u_round(self.d_sect_eff_surf, '[mm]') if parametric_curves else '',
                         u_round(self.d_sect_eff_lin, '[mm]') if parametric_curves else '',
                         self.warning]
        liste_unit = ['', '', '', '', '', '', '', '', '[s]',
                      '[m]',
                      '[m]',
                      '[m]',
                      '[m²]',
                      '[m²]',
                      '[m1/2]',
                      '[J/m²Ks1/2]',
                      '[]',
                      '[]',
                      '[]',
                      '[m²]',
                      '[m²]',
                      '[m²]',
                      '[m²]',
                      '[m²]',
                      '[mn]',
                      '[mn]',
                      '[mn]',
                      '[mn]',
                      '[mm]',
                      '[MJ/m²]',
                      '[kW/m²]',
                      '[mn]',
                      '[mn]',
                      '[kJ/g]',
                      '[]' if not parametric_curves else '[mn]',
                      '[m/s ; ° ; HRR_factor]' if not parametric_curves else '[mn]',
                      '' if not parametric_curves else '[mm/mn]',
                      '' if not parametric_curves else '[mm/mn]',
                      '' if not parametric_curves else '[mm]',
                      '' if not parametric_curves else '[mm]',
                      '' if not parametric_curves else '[mm]',
                      '' if not parametric_curves else '[mm]',
                      '']
        return [liste_champs, liste_valeurs, liste_unit]

    def description(self):
        # Créer un nouveau DataFrame avec une nouvelle colonne
        list_description = self.to_lists()
        df = pd.DataFrame({'Description': list_description[0], 'Value': list_description[1],
                           'Units': list_description[2]})
        return df

    def dump_to_first_sheet_xlsx(self, hrr):
        filename = EnvB.CurPath + '\\' + self.id + '_' + EnvB.default_method + '.xlsx'
        wb = load_workbook(filename)
        first_sheet = wb.worksheets[0]
        first_sheet.title = 'Experiment'
        df_descr = self.description()
        # Définir les polices de couleur
        red_font = Font(bold=True, color="FF0000")
        orange_font = Font(bold=True, color="FFA500")
        for r in dataframe_to_rows(df_descr, index=False, header=True):
            first_sheet.append(r)
            last_row = first_sheet.max_row
            # Vérifier la condition et appliquer la couleur de police
            if r[0] == 'calculation warning' and r[1]:
                for cell in first_sheet[last_row]:  # Appliquer à la dernière ligne ajoutée
                    cell.font = orange_font
            if (r[0] == 'linear effective cross-section  def (A.14)' and
                    self.d_sect_eff_lin < 2 * self.beta_par_lin * self.t_constant_char):
                self.warning = self.warning + f"/!\\ linear effective cross-section < char depth \n"
                for cell in first_sheet[last_row]:  # Appliquer à la dernière ligne ajoutée
                    cell.font = red_font
            if (r[0] == 'surface effective cross-section  def (A.14)' and
                    self.d_sect_eff_surf < 2 * self.beta_par_surf * self.t_constant_char):
                self.warning = self.warning + f"/!\\ surface effective cross-section < char depth \n"
                for cell in first_sheet[last_row]:  # Appliquer à la dernière ligne ajoutée
                    cell.font = red_font

        free_mem(df_descr)
        # Parcourir les colonnes et ajuster leur largeur
        adjust_column_width(first_sheet)
        # Justifier à gauche le texte dans la colonne C et centre la ligne 1
        for cell in first_sheet['C']:
            cell.alignment = Alignment(horizontal='left')
        for cell in first_sheet[1]:
            cell.alignment = Alignment(horizontal='center')

        if not EnvB.current_zone_model[1] == 'PC':
            chart = ScatterChart()
            if hrr is None:
                hrr = self.hrr_fuel
            # Écrire les données dans les colonnes U et V.
            for i in range(len(hrr.abscissas)):
                first_sheet[f'U{i + 1}'] = hrr.abscissas[i].tolist()
                first_sheet[f'V{i + 1}'] = hrr.ordinates[i].tolist()
            # Utiliser Reference pour extraire les valeurs
            # values = [ws[f'V{i + 1}'].value for i in range(len(a))]
            # x_values = [ws[f'U{i + 1}'].value for i in range(len(a))]
            values = Reference(first_sheet, min_col=22, min_row=1, max_row=len(hrr.abscissas))  # Colonne V
            x_values = Reference(first_sheet, min_col=21, min_row=1, max_row=len(hrr.abscissas))  # Colonne U
            nom_series = 'HRR Fuel'
            chart.title = nom_series + 'Initial'
            # values = self.hrr_fuel.ordinates
            # values = np.array([v.decode('utf-8') if isinstance(v, bytes) else v for v in values])
            # x_values = np.array([x.decode('utf-8') if isinstance(x, bytes) else x for x in x_values])
            series = SeriesFactory(values, x_values, title='')
            chart.series.append(series)
            chart.x_axis.title = 'Time (s)'
            loc = f"{column_name(4)}{8}"
            first_sheet.add_chart(chart, loc)  # Ajustez la référence de cellule au besoin

        test_sauvegarde(wb, filename)
        wb.close()


# End define compartment class
#########

#########
# Read an Excel / csv file

# import os
# import signal
# import psutil
#
# def kill_excel_process():
#     # Parcourir tous les processus en cours
#     for proc in psutil.process_iter(['pid', 'name']):
#         # Vérifier si le processus est Excel
#         if proc.info['name'] == 'EXCEL.EXE':
#             os.kill(proc.info['pid'], signal.SIGTERM)

def convert_xls_to_csv(xls_file):
    if os.path.exists(xls_file):
        xls = pd.ExcelFile(xls_file)

        # Convertir chaque feuille en fichier CSV
        for sheet_name in xls.sheet_names:
            d_fr = pd.read_excel(xls_file, sheet_name=sheet_name)
            csv_file = f'{sheet_name}.csv'
            d_fr.to_csv(csv_file, index=False)
            print(f"Feuille '{sheet_name}' convertie en '{csv_file}'")
    print("Conversion terminée !")


def convert_csv_to_xlsx(csv_file, xlsx_file):
    # Créer un classeur
    workbook = Workbook()
    sheet = workbook.active

    # Lire le fichier CSV et écrire les données dans la feuille de calcul
    with open(csv_file, 'r', newline='', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            sheet.append(row)

    # Enregistrer le fichier Excel
    workbook.save(xlsx_file)


# # Exemple d'utilisation
# csv_file = 'votre_fichier.csv'
# xlsx_file = 'votre_fichier.xlsx'
# csv_to_xlsx(csv_file, xlsx_file)
# print(f"Le fichier {csv_file} a été transformé en {xlsx_file}.")


# Old version (sans contrôles d'erreurs, à supprimer
# def extract_dataframe(xls_file, sheet_to_extract):
#     if os.path.exists(xls_file):
#         xls = pd.ExcelFile(xls_file)
#
#         # Convertir chaque feuille en fichier CSV
#         for sheet_name in xls.sheet_names:
#             if sheet_name == sheet_to_extract:
#                 print(f"'{sheet_name}' extraite")
#                 return pd.read_excel(xls_file, sheet_name=sheet_name)


def csv_to_dataframe(csv_file):
    # Lire le fichier CSV et le transformer en DataFrame sans la première ligne
    dataframe = pd.read_csv(csv_file, skiprows=1)
    return dataframe


# Exemple d'utilisation
# csv_file = 'votre_fichier.csv'
# dataframe = csv_to_dataframe(csv_file)

def extract_dataframe(xls_file, sheet_to_extract):
    while True:
        if os.path.exists(xls_file):
            try:
                wb = load_workbook(xls_file, read_only=True, data_only=True)
                sheet = wb[sheet_to_extract]
                data = sheet.values
                columns = next(data)
                df = pd.DataFrame(data, columns=columns)
                wb.close()  # Fermer le classeur Excel
                return df
            except Exception as e:
                print(f"Erreur lors de l'extraction du DataFrame: {e}")
                # return None
                input("Appuyez sur une touche pour réessayer...")
        else:
            print(f"Le fichier '{xls_file}' n'existe pas.")
            # return None
            input("Appuyez sur une touche pour réessayer...")


# End Read an Excel / csv file
#########

#########
# Write an Excel / csv file

def find_chart_by_title(ws_graph, title):
    # Parcourez tous les graphiques dans la feuille de calcul
    # noinspection PyProtectedMember
    for chart in ws_graph._charts:
        # Vérifiez si le graphique a un titre et s'il correspond au titre recherché
        if chart.title and chart.title.text.rich and chart.title.text.rich.p:
            # s_title = chart.title.text.rich.p[0].r[0].t
            s_title = ''.join(run.t for para in chart.title.text.rich.p for run in para.r)
            if s_title == title:
                return chart
    # Si aucun graphique correspondant n'est trouvé, retournez None
    return None


# win32com
# def find_chart_by_title(ws, title):
#     for chart in ws.ChartObjects():
#         if chart.Chart.HasTitle and chart.Chart.ChartTitle.Text == title:
#             return chart.Chart
#     return None

def column_name(i):
    if i <= 0:
        raise ValueError("L'indice de la colonne doit être positif")
    result = ""
    while i > 0:
        i -= 1
        result = chr(ord('A') + i % 26) + result
        i //= 26
    return result


def write_list_line_r(ws, liste, r):
    for col, val in enumerate(liste, start=1):
        ws.cell(row=r, column=col, value=val)


def adjust_column_width(ws_sheet):
    # Parcourir les colonnes et ajuster leur largeur
    for col in ws_sheet.columns:
        max_length = 0
        column = col[0].column  # Obtenir le numéro de la colonne
        for cell in col:
            try:
                c = str(cell.value)
                if not (c.startswith('=') or c in EnvB.long_line) and len(c) > max_length:
                    max_length = len(str(cell.value))
                if c == '/!\\ NO':
                    cell.font = Font(color="FF0000")
            except ValueError as e:
                # Gestion spécifique pour ValueError
                print(f"Une erreur ValueError s'est produite: {e}")
                pass
            except TypeError as e:
                # Gestion spécifique pour TypeError
                print(f"Une erreur TypeError s'est produite: {e}")
                pass
            except Exception as e:
                # Clause de secours pour les autres exceptions
                print(f"Une erreur inattendue s'est produite: {e}")
                pass
        adjusted_width = (max_length + 2)  # Ajouter un peu plus d'espace
        ws_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


def define_formule(header, df, niter, li, unit, exposed, mean):
    # Trouver l'index de la colonne avec le titre spécifié
    if exposed:
        a = '=ARRONDI(SI(D2>=0,B'
        b = '+C'
        c = '*D2,RECHERCHEV(B2+D2,'
        d = '=SI(D2>=0,ARRONDI(B'
        end = '*D2,'
    else:
        a = '=ARRONDI(SI(J2>=0,H'
        b = '+I'
        c = '*J2,RECHERCHEV(H2+J2,'
        d = '=SI(J2>=0,ARRONDI(H'
        end = '*J2,'
    try:
        if mean:
            formule = d + str(li) + b + str(li) + end + str(EnvB.unit_round(unit)) + '), "To calculate") '
        else:
            col_index = df.columns.get_loc(header) + 1
            col_letter = get_column_letter(col_index)
            # plage = f"iter_{niter}!$A:${col_letter}"
            plage = 'iter_' + str(niter) + '!$A:$' + str(col_letter)
            # noinspection SpellCheckingInspection
            formule = (a + str(li) + b + str(li) + c + plage + ',' + str(col_index) + ')),' +
                       str(EnvB.unit_round(unit)) + ')')
    except KeyError:
        formule = d + str(li) + b + str(li) + end + str(EnvB.unit_round(unit)) + '), "To calculate") '
    return formule


def write_xls_simplified_summary_char(ws, df_new, comp):  # t_char_fin, wood_consumed:
    # écriture des résultats d'épaisseur de carbonisation

    t_char_fin = comp.time_char_fin / 60
    wood_consumed = comp.wood_consumed
    # calculated = comp.calculated()
    (wall_min_f, wall_max_f, wall_mean_f, ceiling_min_f, ceiling_max_f, ceiling_mean_f, beam_min_f, beam_max_f,
     beam_mean_f, column_min_f, column_max_f, column_mean_f, p_wall_min_f, p_wall_max_f, p_wall_mean_f,
     p_ceiling_min_f, p_ceiling_max_f, p_ceiling_mean_f) = calcul_final_char(df_new, t_char_fin, True)
    list_results = [['Summary of char calculations:', 'exposed', '', '', '', '', '', 'protected', '', '', '', '', '',
                     't_max', 't_end'],
                    ['@end = ', df_new['Time (min)'].iloc[-1], '[mn]',
                     '', '', '', '',
                     df_new['Time (min)'].iloc[-1],
                     '', '', '', '', '',
                     u_round(comp.t_max, '[mn]'),
                     u_round(comp.t_end, '[mn]'), '[mn]'],
                    ['location', 'thickness [mm]', '', '', '', '', '', '', '', 'thickness [mm]', '', '', '',
                     u_round(comp.t_max * 60, '[s]'),
                     u_round(comp.t_end * 60, '[s]'), '[s]'],
                    ['wall ', wall_max_f, '', '', '', '', '', '', '', p_wall_max_f],
                    ['ceiling', ceiling_max_f, '', '', '', '', '', '', '', p_ceiling_max_f],
                    ['Column', column_max_f],
                    ['Beam', beam_max_f],
                    ['floor', u_round(df_new['Cumul e_char floor (mm)'].max(), '[mm]')],
                    # ['Full calculation', calculated],
                    ['One Wood material totally consumed', wood_consumed]]
    if wood_consumed:
        list_results.append(['floor, ceiling, wall:', str(comp.floor_mat.wood_consumed)
                             + ', ' + str(comp.ceiling_mat.wood_consumed) + ', ' + str(comp.wall_mat.wood_consumed)])
    red_font = Font(color="FF0000")
    for i in list_results:
        write_list_line_r(ws, i, list_results.index(i) + 1)
        if wood_consumed and i[0] == 'One Wood material totally consumed':
            ws.cell(row=list_results.index(i) + 1, column=2).font = red_font
        # if calculated == '1' and i[0] == 1:
        #     ws.cell(row=list_results.index(i) + 1, column=2).font = red_font


def write_xls_summary_char(ws, df_new, niter, hrr, comp):  # t_char_fin, wood_consumed:
    # écriture des résultats d'épaisseur de carbonisation
    # L2=['@end of simulation :', compartment1.exposed_time(), 'mn     @', df_new['Time (min)'].#iloc[-2], 'mn']
    # #iloc[i] indexation négative commence à partir de la fin du DataFrame,
    # donc -1 sera la dernière valeur et -2 l’avant-dernière.

    # Trouver l'index de la colonne avec le titre spécifié
    # col_index = df_new.columns.get_loc('Cumul e_char ceiling (mm)') + 1
    # col_letter = get_column_letter(col_index)
    # plage = f 'iter_{niter} ! A : {col_letter}'
    # formule = f'=SI (D2>0,B7+C7*D2,#RECHERCHEV(B2+D2,{plage},{col_index}))'

    t_char_fin = comp.time_char_fin / 60
    wood_consumed = comp.wood_consumed
    # calculated = comp.calculated()
    # Convertir l'index en lettre de colonne Excel
    hll = df_new['Layer (m)'].iloc[-1]
    hul = df_new['h_UL (m)'].iloc[-1]
    v_wall_min = df_new['Beta_LL (mm/min)'].iloc[-1]
    v_wall_max = df_new['Beta_UL (mm/min)'].iloc[-1]
    v_wall_mean = u_round((v_wall_min * hll + v_wall_max * hul) / (hll + hul), '[mm/min]')
    wall_min = df_new['Cumul e_char LL (mm)'].max()  # i.e. i loc[-1]
    wall_max = df_new['Cumul e_char UL (mm)'].max()
    wall_mean = u_round((wall_min * hll + wall_max * hul) / (hll + hul), '[mm]')
    vp_wall_min = df_new['Beta_protected_LL (mm/min)'].iloc[-1]
    vp_wall_max = df_new['Beta_protected_UL (mm/min)'].iloc[-1]
    vp_wall_mean = u_round((vp_wall_min * hll + vp_wall_max * hul) / (hll + hul), '[mm/min]')
    p_wall_min = df_new['Cumul e_char protected_LL (mm)'].max()  # i.e. i loc[-1]
    p_wall_max = df_new['Cumul e_char protected_UL (mm)'].max()
    p_wall_mean = u_round((p_wall_min * hll + p_wall_max * hul) / (hll + hul), '[mm]')
    v_column_min = df_new['Beta_column_LL (mm/min)'].iloc[-1]  # i.e. i loc[-1]
    v_column_max = df_new['Beta_column_UL (mm/min)'].iloc[-1]
    v_column_mean = u_round((v_column_min * hll + v_column_max * hul) / (hll + hul), '[mm/min]')
    column_min = df_new['Cumul e_char column LL (mm)'].max()
    column_max = df_new['Cumul e_char column UL (mm)'].max()
    column_mean = u_round((column_min * hll + column_max * hul) / (hll + hul), '[mm]')
    # hll_r = u_round(hll,'[m]')
    # hul_r = u_round(hul,'[m]')
    # formule_column = '=(D8*'+str(hll_r)+'+D9*'+str(hul_r)+')/('+str(hll_r)+'+'+str(hul_r)+')'
    (wall_min_f, wall_max_f, wall_mean_f, ceiling_min_f, ceiling_max_f, ceiling_mean_f, beam_min_f, beam_max_f,
     beam_mean_f, column_min_f, column_max_f, column_mean_f, p_wall_min_f, p_wall_max_f, p_wall_mean_f,
     p_ceiling_min_f, p_ceiling_max_f, p_ceiling_mean_f) = calcul_final_char(df_new, t_char_fin)
    list_results = [['Summary of char calculations:', 'exposed', '', '', '', '', '', 'protected', '', '', '', '', '',
                     't_max', 't_dec', 't_fd'],
                    ['@end = ', df_new['Time (min)'].iloc[-1], '[mn]    @end +',
                     t_char_fin - df_new['Time (min)'].iloc[-1], '[mn]',
                     '@char_fin = ' + str(u_round(t_char_fin, '[mn]')), '[mn] @end =',
                     df_new['Time (min)'].iloc[-1], '[mn]    @end +',
                     t_char_fin - df_new['Time (min)'].iloc[-1], '[mn]',
                     '@char_fin = ' + str(u_round(t_char_fin, '[mn]')), '[mn]',
                     u_round(hrr.t_max / 60, '[mn]'),
                     u_round(hrr.t_dec / 60, '[mn]'), u_round((hrr.t_dec - hrr.t_max) / 60, '[mn]'), '[mn]'],
                    ['location', 'thickness [mm]', 'speed [mm/mn]', 'thickness [mm]', '', '', '',
                     'thickness [mm]', 'speed [mm/mn]', 'thickness [mm]', '', '', '',
                     u_round(hrr.t_max, '[s]'),
                     u_round(hrr.t_dec, '[s]'), u_round((hrr.t_dec - hrr.t_max), '[s]'), '[s]'],
                    ['wall min', u_round(wall_min, '[mm]'),
                     u_round(v_wall_min, '[mm/min]'),
                     define_formule('Cumul e_char LL (mm)', df_new, niter, 4, '[mm]', True, False), 'LL value',
                     wall_min_f, 'min of e_char(z_wall)',
                     u_round(p_wall_min, '[mm]'),
                     u_round(vp_wall_min, '[mm/min]'),
                     define_formule('Cumul e_char protected_LL (mm)', df_new, niter, 4, '[mm]', False, False),
                     'LL value',
                     p_wall_min_f, 'min of protected e_char(z_wall)'],
                    ['wall max', u_round(wall_max, '[mm]'),
                     u_round(v_wall_max, '[mm/min]'),
                     define_formule('Cumul e_char UL (mm)', df_new, niter, 5, '[mm]', True, False), 'UL value',
                     wall_max_f, 'max of e_char(z_wall)',
                     u_round(p_wall_max, '[mm]'),
                     u_round(vp_wall_max, '[mm/min]'),
                     define_formule('Cumul e_char protected_UL (mm)', df_new, niter, 5, '[mm]', False, False),
                     'UL value',
                     p_wall_max_f, 'max of protected e_char(z_wall)'],
                    ['wall mean', wall_mean, v_wall_mean,
                     define_formule('Cumul e_char UL (mm)', df_new, niter, 6, '[mm]', True, True), 'Based on z(LL)',
                     wall_mean_f, 'mean of e_char(z_wall)',
                     p_wall_mean, vp_wall_mean,
                     define_formule('Cumul e_char protected_UL (mm)', df_new, niter, 6, '[mm]', False, True),
                     'Based on z(LL)',
                     p_wall_mean_f, 'mean of protected e_char(z_wall)'],
                    ['ceiling', u_round(df_new['Cumul e_char ceiling (mm)'].max(), '[mm]'),
                     u_round(df_new['Beta_ceiling (mm/min)'].iloc[-1], '[mm/min]'),
                     define_formule('Cumul e_char ceiling (mm)', df_new, niter, 7, '[mm]', True, False),
                     '', ceiling_mean_f, '',
                     u_round(df_new['Cumul e_char protected_ceiling (mm)'].max(), '[mm]'),
                     u_round(df_new['Beta_protected_ceiling (mm/min)'].iloc[-1], '[mm/min]'),
                     define_formule('Cumul e_char protected_ceiling (mm)', df_new, niter, 7, '[mm]', False, False),
                     '', p_ceiling_mean_f],
                    ['Column min', u_round(column_min, '[mm]'),
                     u_round(v_column_min, '[mm/min]'),
                     define_formule('Cumul e_char column LL (mm)', df_new, niter, 8, '[mm]', True, False), 'LL value',
                     column_min_f],
                    ['Column max', u_round(column_max, '[mm]'),
                     u_round(v_column_max, '[mm/min]'),
                     define_formule('Cumul e_char column UL (mm)', df_new, niter, 9, '[mm]', True, False), 'UL value',
                     column_max_f],
                    ['Column mean', column_mean, v_column_mean,
                     define_formule('Cumul e_char column UL (mm)', df_new, niter, 10, '[mm]', True, True),
                     'Based on z(LL)', column_mean_f],
                    ['Beam', u_round(df_new['Cumul e_char beam (mm)'].max(), '[mm]'),
                     u_round(df_new['Beta_beam (mm/min)'].iloc[-1], '[mm/min]'),
                     define_formule('Cumul e_char beam (mm)', df_new, niter, 11, '[mm]', True, False),
                     '', beam_mean_f],
                    ['floor', u_round(df_new['Cumul e_char floor (mm)'].max(), '[mm]'),
                     u_round(df_new['Beta_floor (mm/min)'].iloc[-1], '[mm/min]'),
                     define_formule('Cumul e_char floor (mm)', df_new, niter, 12, '[mm]', True, False)],
                    # ['Full calculation', calculated],
                    ['One Wood material totally consumed', wood_consumed]]
    if wood_consumed:
        list_results.append(['floor, ceiling, wall:', str(comp.floor_mat.wood_consumed)
                             + ', ' + str(comp.ceiling_mat.wood_consumed) + ', ' + str(comp.wall_mat.wood_consumed)])
    red_font = Font(color="FF0000")
    for i in list_results:
        write_list_line_r(ws, i, list_results.index(i) + 1)
        if wood_consumed and i[0] == 'One Wood material totally consumed':
            ws.cell(row=list_results.index(i) + 1, column=2).font = red_font
        # if calculated == '1' and i[0] == 1:
        #     ws.cell(row=list_results.index(i) + 1, column=2).font = red_font


def test_sauvegarde(wb_t, filename):
    # Vérifiez si le fichier est ouvert
    while True:
        try:
            wb_t.save(filename)
            print(f"Le fichier {filename} a été enregistré avec succès.")
            break  # Sortir de la boucle si le code réussit
        except PermissionError:
            print(f"Le fichier {filename} est actuellement ouvert. Veuillez le fermer et réessayer.")
            input("Appuyez sur une touche pour réessayer...")
        except Exception as e:
            print(f"Erreur lors de l'extraction du DataFrame: {e}")
            # del wb
            # del ws_graph
            # del ws_data
            # free_mem(compartment1)
            sys.exit(1)


def write_xls_from_dataframe(df_selected: pd.DataFrame, niter: int, list_series: list, compartment1: Compartment,
                             one_zone_model=False) -> None:
    # Créez un nouveau classeur et ajoutez une feuille de calcul
    wb = Workbook()
    # Nom du fichier
    filename = EnvB.CurPath + '\\' + compartment1.id + '_' + EnvB.default_method + '.xlsx'

    # Ajoutez une feuille de calcul pour le graphique avec le nom "graph"
    ws_graph = wb.create_sheet("Results")
    # Ajoutez une feuille de calcul pour les données
    ws_data = wb.create_sheet(f"iter_{niter}")
    # Écrivez le DataFrame dans la feuille de calcul
    for r in dataframe_to_rows(df_selected, index=False, header=True):
        ws_data.append(r)
    if one_zone_model:
        write_xls_simplified_summary_char(ws_graph, df_selected, compartment1)
    else:
        write_xls_summary_char(ws_graph, df_selected, niter, compartment1.hrr_fuel, compartment1)
    # compartment1.time_char_fin / 60, compartment1.wood_consumed
    adjust_column_width(ws_graph)
    # Obtenez les indices des colonnes pour les colonnes avec des noms dans list_graph
    list_graph = [element[1] for element in list_series]
    list_axes = [element[0] for element in list_series]
    column_indices = [i for i, col_name in enumerate(df_selected.columns, start=1) if col_name in list_graph]
    # Définissez la référence pour l'axe des abscisses (colonne A)
    # x_values = Reference (ws_data, min_col=x_axis_indice, min_row=2, max_row=ws_data.max_row)
    i = 0
    n_col = 3  # Nombre de graphiques sur la largeur de la feuille
    # Ajoutez des données pour chaque colonne
    for col_index in column_indices:  # Colonnes dont l'intitulé est dans list_graph
        # Créez un graphique de nuage de points
        pos_x = (i % n_col) * 5 + 1
        # if not (i % n_col) == 0:
        #     pos_x = pos_x - 6
        pos_y = (i // n_col) * 18 + 17
        chart = ScatterChart()
        nom_series = ws_data.cell(row=1, column=col_index).value
        chart.title = "Evolution " + nom_series
        if not (df_selected.iloc[:, col_index - 1] == 0).all():
            axe_x = list_axes[list_graph.index(df_selected.columns[col_index - 1])]
            x_index = df_selected.columns.get_loc(axe_x)
            x_values = Reference(ws_data, min_col=x_index + 1, min_row=2, max_row=ws_data.max_row)
            # chart.style = 13
            values = Reference(ws_data, min_col=col_index, min_row=2, max_row=ws_data.max_row)
            series = SeriesFactory(values, x_values, title='Iter_0')
            # series = Series(values, x_values, title=nom_series)
            # ou title=ws_data.cell(row=1, column=i).value) title_from_data=True
            chart.series.append(series)
            # Configuration et graduations des axes
            # chart.x_axis.majorTickMark = 'in'
            # chart.y_axis.majorTickMark = 'in'
            chart.x_axis.title = ws_data.cell(row=1, column=x_index + 1).value
        # Placez le graphique dans la feuille de calcul
        loc = f"{column_name(pos_x)}{pos_y}"
        ws_graph.add_chart(chart, loc)  # Ajustez la référence de cellule au besoin
        i = i + 1

    # Fermez le fichier
    test_sauvegarde(wb, filename)
    wb.close()
    # del wb
    # del ws_graph
    # del ws_data


def add_data_and_series(df_new: pd.DataFrame, niter: int, list_series: list,
                        compartment1: Compartment, one_zone_model=False) -> None:
    filename = EnvB.CurPath + '\\' + compartment1.id + '_' + EnvB.default_method + '.xlsx'
    # Rouvrir le fichier
    # Vérifiez si le fichier est ouvert
    if os.path.isfile(filename):
        try:
            wb = load_workbook(filename)
            # print(f"Le fichier {filename} a été ouvert avec succès.")
            # Accéder à la feuille de graphique existante
            ws_graph = wb["Results"]
            # ws_experiment = wb[0]
            # chart = ws_graph["E5"].chart # chart = ws_graph[chart_location].chart
            # Ajoutez une nouvelle feuille de calcul pour les données supplémentaires
            ws_new = wb.create_sheet(f"iter_{niter}")
            # Écrivez le nouveau DataFrame dans la nouvelle feuille de calcul
            for r in dataframe_to_rows(df_new, index=False, header=True):
                ws_new.append(r)
            # Obtenez les indices des colonnes pour les colonnes avec des noms dans list_graph
            list_graph = [element[1] for element in list_series]
            list_axes = [element[0] for element in list_series]
            column_indices = [i for i, col_name in enumerate(df_new.columns, start=1) if col_name in list_graph]
            # Définissez la référence pour l'axe des abscisses (colonne A)
            # x_values = Reference (ws_new, min_col=1, min_row=3, max_row=ws_new.max_row)
            # Ajoutez des données pour chaque colonne
            for col_index in column_indices:  # Colonnes dont l'intitulé est dans list_graph
                # Accéder au graphique existant
                nom_series = "Evolution " + ws_new.cell(row=1, column=col_index).value
                # noinspection PyTypeChecker
                chart = find_chart_by_title(ws_graph, nom_series)
                if chart is not None and not (df_new.iloc[:, col_index - 1] == 0).all():
                    axe_x = list_axes[list_graph.index(df_new.columns[col_index - 1])]
                    x_index = df_new.columns.get_loc(axe_x)
                    x_values = Reference(ws_new, min_col=x_index + 1, min_row=2, max_row=ws_new.max_row)
                    values = Reference(ws_new, min_col=col_index, min_row=2, max_row=ws_new.max_row)
                    series = SeriesFactory(values, x_values, title=f"iter_{niter}")
                    # ou title=ws_data.cell(row=1, column=i).value) title_from_data=True
                    chart.series.append(series)

            # écriture des résultats d'épaisseur de carbonisation
            if one_zone_model:
                write_xls_simplified_summary_char(ws_graph, df_new, compartment1)
            else:
                hrr = HrrFunction(df_new['Time (sec)'].tolist(), df_new['Unconstrained HRR (kW)'].tolist())
                write_xls_summary_char(ws_graph, df_new, niter, hrr, compartment1)  # compartment1.time_char_fin / 60,
            # compartment1.wood_consumed
            # free mem hrr?

            # Sauvegarder le fichier
            wb.save(filename)
            # Fermer le fichier
            wb.close()
            # del wb
            # del ws_graph
            # del ws_new
            # del ws_experiment
        except PermissionError:
            print(f"Le fichier {filename} est actuellement ouvert. Veuillez le fermer et réessayer.")
    else:
        print(f"Le fichier {filename} n'existe pas.")


def find_ligne_excel(sheet, nom):
    # Trouver la ligne avec le nom dans la colonne A
    for row in range(1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == nom:
            return row
    return None


def find_colonne_excel(sheet, nom):
    # Trouver la colonne avec le nom dans la ligne 1
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == nom:
            return col
    return None


def calcul_final_char(df_new, t_char_fin, one_zone_model=False):
    t_left = t_char_fin - df_new['Time (min)'].iloc[-1]
    if t_left < 0:
        index = (df_new['Time (min)'] - t_char_fin).abs().idxmin()
        # DO calcul wall mean,max,min f(z,t_char_fin) -> char_thickness
        if one_zone_model:
            wall_mean = wall_min = wall_max = u_round(df_new['Cumul e_char wall (mm)']
                                                      .iloc[index], '[mm]')
            p_wall_mean = p_wall_min = p_wall_max = u_round(df_new['Cumul e_char protected_wall (mm)']
                                                            .iloc[index], '[mm]')
            column_mean = column_min = column_max = u_round(df_new['Cumul e_char protected_wall (mm)']
                                                            .iloc[index], '[mm]')
        else:
            wall_min = u_round(df_new['e_char(z_wall) (mm)'].min(), '[mm]')
            wall_max = u_round(df_new['e_char(z_wall) (mm)'].max(), '[mm]')
            wall_mean = u_round(df_new['e_char(z_wall) (mm)'].mean(), '[mm]')
            p_wall_min = u_round(df_new['e_char protected(z_wall) (mm)'].min(), '[mm]')
            p_wall_max = u_round(df_new['e_char protected(z_wall) (mm)'].max(), '[mm]')
            p_wall_mean = u_round(df_new['e_char protected(z_wall) (mm)'].mean(), '[mm]')
            column_min = df_new['Cumul e_char column LL (mm)'].iloc[index]
            column_max = df_new['Cumul e_char column UL (mm)'].iloc[index]
            hll = df_new['Layer (m)'].iloc[index]
            hul = df_new['h_UL (m)'].iloc[index]
            column_mean = u_round((column_min * hll + column_max * hul) / (hll + hul), '[mm]')
        ceiling_mean = ceiling_min = ceiling_max = u_round(df_new['Cumul e_char ceiling (mm)']
                                                           .iloc[index], '[mm]')
        p_ceiling_mean = p_ceiling_min = p_ceiling_max = u_round(df_new['Cumul e_char protected_ceiling (mm)']
                                                                 .iloc[index], '[mm]')
        beam_min = beam_max = beam_mean = u_round(df_new['Cumul e_char beam (mm)'].iloc[index], '[mm]')
    else:
        if t_left > 0:
            if one_zone_model:
                wall_mean = wall_min = wall_max = u_round(df_new['Cumul e_char wall (mm)'].max(), '[mm]')
                p_wall_mean = p_wall_min = p_wall_max = u_round(df_new['Cumul e_char protected_wall (mm)'].max(),
                                                                '[mm]')
                column_mean = column_min = column_max = u_round(df_new['Cumul e_char column (mm)'].max(), '[mm]')
                p_ceiling_mean = p_ceiling_min = p_ceiling_max = u_round(df_new['Cumul e_char protected_ceiling (mm)']
                                                                         .max(), '[mm]')
                ceiling_mean = ceiling_min = ceiling_max = u_round(df_new['Cumul e_char ceiling (mm)'].max(),
                                                                   '[mm]')
                beam_min = beam_max = beam_mean = u_round(df_new['Cumul e_char beam (mm)'].max(), '[mm]')
                # Not relevant, total element char given in sheet Experiment
                # wall_mean = wall_min = wall_max = u_round(df_new['Cumul e_char wall (mm)'].max() +
                #                                           t_left * df_new['Beta_wall (mm/min)'].iloc[-1], '[mm]')
                # p_wall_mean = p_wall_min = p_wall_max = u_round(df_new['Cumul e_char protected_wall (mm)'].max() +
                #                                                 t_left * df_new['Beta_protected_wall (mm/min)']
                #                                                 .iloc[-1], '[mm]')
                # column_mean = column_min = column_max = u_round(df_new['Cumul e_char column (mm)'].max() +
                #                                                 t_left * df_new['Beta_column (mm/min)'].iloc[-1],
                #                                                 '[mm]')
            else:
                wall_min = u_round(df_new['e_char(z_wall) (mm)'].min() +
                                   t_left * df_new['Beta_LL (mm/min)'].iloc[-1], '[mm]')
                wall_max = u_round(df_new['e_char(z_wall) (mm)'].max() +
                                   t_left * df_new['Beta_UL (mm/min)'].iloc[-1], '[mm]')
                wall_mean = u_round(df_new['e_char(z_wall) (mm)'].mean() +
                                    t_left * df_new['Beta_UL (mm/min)'].iloc[-1], '[mm]')
                p_wall_min = u_round(df_new['e_char protected(z_wall) (mm)'].min() +
                                     t_left * df_new['Beta_protected_LL (mm/min)'].iloc[-1], '[mm]')
                p_wall_max = u_round(df_new['e_char protected(z_wall) (mm)'].max() +
                                     t_left * df_new['Beta_protected_UL (mm/min)'].iloc[-1], '[mm]')
                p_wall_mean = u_round(df_new['e_char protected(z_wall) (mm)'].mean() +
                                      t_left * df_new['Beta_protected_UL (mm/min)'].iloc[-1], '[mm]')
                column_min = (df_new['Cumul e_char column LL (mm)'].max() + t_left *
                              df_new['Beta_column_LL (mm/min)'].iloc[-1])
                column_max = (df_new['Cumul e_char column UL (mm)'].max() + t_left *
                              df_new['Beta_column_UL (mm/min)'].iloc[-1])
                ceiling_mean = ceiling_min = ceiling_max = \
                    u_round(df_new['Cumul e_char ceiling (mm)'].max() +
                            t_left * df_new['Beta_ceiling (mm/min)'].iloc[-1], '[mm]')
                p_ceiling_mean = p_ceiling_min = p_ceiling_max = \
                    u_round(df_new['Cumul e_char protected_ceiling (mm)'].max() +
                            t_left * df_new['Beta_protected_ceiling (mm/min)'].iloc[-1], '[mm]')
                beam_min = beam_max = beam_mean = \
                    u_round(df_new['Cumul e_char beam (mm)'].max() +
                            t_left * df_new['Beta_beam (mm/min)'].iloc[-1], '[mm]')
        else:  # t_left==0
            if one_zone_model:
                wall_mean = wall_min = wall_max = u_round(df_new['Cumul e_char wall (mm)'].max(), '[mm]')
                p_wall_mean = p_wall_min = p_wall_max = u_round(df_new['Cumul e_char protected_wall (mm)'].max(),
                                                                '[mm]')
                column_mean = column_min = column_max = u_round(df_new['Cumul e_char column (mm)'].max(), '[mm]')
            else:
                wall_min = u_round(df_new['e_char(z_wall) (mm)'].min(), '[mm]')
                wall_max = u_round(df_new['e_char(z_wall) (mm)'].max(), '[mm]')
                wall_mean = u_round(df_new['e_char(z_wall) (mm)'].mean(), '[mm]')
                p_wall_min = u_round(df_new['e_char protected(z_wall) (mm)'].min(), '[mm]')
                p_wall_max = u_round(df_new['e_char protected(z_wall) (mm)'].max(), '[mm]')
                p_wall_mean = u_round(df_new['e_char protected(z_wall) (mm)'].mean(), '[mm]')
                column_min = df_new['Cumul e_char column LL (mm)'].max()
                column_max = df_new['Cumul e_char column UL (mm)'].max()
            p_ceiling_mean = p_ceiling_min = p_ceiling_max = u_round(df_new['Cumul e_char protected_ceiling (mm)']
                                                                     .max(), '[mm]')
            ceiling_mean = ceiling_min = ceiling_max = u_round(df_new['Cumul e_char ceiling (mm)'].max(),
                                                               '[mm]')
            beam_min = beam_max = beam_mean = u_round(df_new['Cumul e_char beam (mm)'].max(), '[mm]')
        if not one_zone_model:
            hll = df_new['Layer (m)'].iloc[-1]
            hul = df_new['h_UL (m)'].iloc[-1]
            column_mean = u_round((column_min * hll + column_max * hul) / (hll + hul), '[mm]')
    column_min = u_round(column_min, '[mm]')
    column_max = u_round(column_max, '[mm]')
    # TODO Floor char
    # noinspection PyUnboundLocalVariable
    return (wall_min, wall_max, wall_mean, ceiling_min, ceiling_max, ceiling_mean, beam_min, beam_max, beam_mean,
            column_min, column_max, column_mean, p_wall_min, p_wall_max, p_wall_mean, p_ceiling_min, p_ceiling_max,
            p_ceiling_mean)


# TODO write status from 'flag stop'
def write_xls_char_results(df_new, nom, t_char_fin, auto_extinction, fail=False, warning=''):
    while True:
        if os.path.isfile(EnvB.current_experiment_excel_file):
            try:
                wb = load_workbook(EnvB.current_experiment_excel_file)
                sheet = wb['Char Depth Graph']
                # ! Ouvrir un seul workbook,
                # # wb_data = load_workbook(EnvB.experiment_excel_file, data_only=True)
                # # sheet_data = wb_data['Char Depth Graph']
                # Vérifier si les indices de ligne et de colonne ont été trouvés
                # # row_num = find_ligne_excel(sheet_data, nom)
                if nom in EnvB.result_list_char:
                    row_num = EnvB.result_list_char.index(nom) + 2
                else:
                    row_num = None
                (wall_min, wall_max, wall_mean, ceiling_min, ceiling_max, ceiling_mean, beam_min, beam_max, beam_mean,
                 column_min, column_max, column_mean, p_wall_min, p_wall_max, p_wall_mean, p_ceiling_min, p_ceiling_max,
                 p_ceiling_mean) = calcul_final_char(df_new, t_char_fin, EnvB.current_zone_model[1] == 'PC')
                list_results = [['Model Wall min', max(wall_min, p_wall_min), max(wall_min, p_wall_min) == p_wall_min],
                                ['Model Wall max', max(wall_max, p_wall_max), max(wall_max, p_wall_max) == p_wall_max],
                                ['Model Wall mean', max(wall_mean, p_wall_mean),
                                 max(wall_mean, p_wall_mean) == p_wall_mean],
                                ['Model Column min', column_min, False], ['Model Column max', column_max, False],
                                ['Model Column mean', column_mean, False],
                                ['Model Ceiling min', max(ceiling_min, p_ceiling_min),
                                 max(ceiling_min, p_ceiling_min) == p_ceiling_min],
                                ['Model Ceiling max', max(ceiling_max, p_ceiling_max),
                                 max(ceiling_max, p_ceiling_max) == p_ceiling_max],
                                ['Model Ceiling mean', max(ceiling_mean, p_ceiling_mean),
                                 max(ceiling_mean, p_ceiling_mean) == p_ceiling_mean],
                                ['Model Beam min', beam_min, False],
                                ['Model Beam max', beam_max, False], ['Model Beam mean', beam_mean, False],
                                ['Model Auto Extinction', auto_extinction, not auto_extinction == 'VRAI'],
                                ['Results write by', EnvB.ProgFile, False],
                                ['Calculation Time', EnvB.CalculationTime, False]]
                if fail:
                    list_results.append(['Calculation Fail', 'YES' + warning, True])
                elif warning:
                    list_results.append(['Calculation Fail', warning, True])
                for i in list_results:
                    red_font = Font(color="FF0000")
                    orange_font = Font(bold=True, color="FFA500")
                    col_num = find_colonne_excel(sheet, i[0])
                    if row_num is not None and col_num is not None:
                        # Écrire la valeur x dans la cellule trouvée
                        sheet.cell(row=row_num, column=col_num).value = i[1]
                        # Appliquer la couleur rouge au texte
                        if i[2]:
                            sheet.cell(row=row_num, column=col_num).font = red_font
                        if fail:
                            # noinspection SpellCheckingInspection
                            sheet.cell(row=row_num, column=col_num).font = Font(bold=True, italic=True,
                                                                                color="FFFF0000")
                        elif warning and i[0] == 'Calculation Fail':
                            sheet.cell(row=row_num, column=col_num).font = orange_font
                    else:
                        print(f"Ligne ou colonne {i} non trouvée")
                # Sauvegarder et fermer le fichier Excel
                wb.save(EnvB.current_experiment_excel_file)
                wb.close()
                # # wb_data.close()
                # # del wb
                # # del wb_data
                break
            except Exception as e:
                print(f"Erreur {e} lors de l'extraction du DataFrame: '{EnvB.current_experiment_excel_file}'")
                # return None
                input("Appuyez sur une touche pour réessayer...")
        else:
            print(f"Le fichier '{EnvB.current_experiment_excel_file}' n'existe pas.")
            # return None
            input("Appuyez sur une touche pour réessayer...")


# #def write_list_to_csv(lst, csv_file):
#     """
#     Écrit une liste de chaînes de caractères dans un fichier CSV.
#     Args :
#         lst (list) : Liste de chaînes de caractères
#         csv_file (str) : Nom du fichier CSV.
#     """
#     df = pd. DataFrame (lst, columns=['Titres de colonnes'])
#     if os.path.isfile(csv_file):
#         try:
#             df.to_csv (csv_file, index=False)
#             # print(f"Les titres de colonnes ont été écrits dans {csv_file}")
#             # Ferme le fichier
#             with open(csv_file, 'a') as f:
#                 f.close()
#         except PermissionError:
#             print(f"Le fichier {csv_file} est actuellement ouvert. Veuillez le fermer et réessayer.")


# End Write an Excel /csv file
#########


#########
# Operation on DataFrame

def find_cell(df, cherche):
    for i in range(df.shape[0]):
        for j in range(df.shape[1]):
            if df.iat[i, j] == cherche:
                return i, j
    print(f"{cherche} non trouvé")
    return None, None


#  Trouver la première ligne qui contient cherche
def find_ligne(df, cherche):
    for i in range(len(df)):
        if cherche in df.iloc[i].values:
            return i
    return None


#  Trouver l'index des lignes contenant cherche
def find_lignes(df, cherche) -> list:
    row_index = df[df.isin([cherche]).any(axis=1)].index
    if not row_index.empty:
        print(cherche + ' trouvé dans les lignes:')
        print(row_index)
    else:
        print(f"{cherche} n'a pas été trouvé dans le DataFrame")
    return row_index


def remove_column(df, header):
    # Trouver toutes les colonnes dont le nom correspond à 'header'
    columns_to_drop = [col for col in df.columns if col == header]
    # Supprimer les colonnes trouvées
    df.drop(columns=columns_to_drop, inplace=True)


def remove_column_none(df):
    # Trouver toutes les colonnes dont le nom correspond à 'header'
    columns_to_drop = [col for col in df.columns if col is None]
    # Supprimer les colonnes trouvées
    df.drop(columns=columns_to_drop, inplace=True)


def arrondir_colonne(df, colonne):
    # Vérifier si la colonne existe dans le DataFrame
    if colonne in df.columns:
        # Appliquer la fonction d'arrondi uniquement aux cellules qui sont des nombres
        # n = EnvB.unit_round(df[colonne].iat[1])
        # current_data['Af'].iat[1]
        # print(n)
        df[colonne] = df[colonne].apply(lambda x: u_round(x, df[colonne].iat[1]) if isinstance(x, float) else x)
    else:
        print(f"La colonne '{colonne}' n'existe pas dans le DataFrame.")
    return df


def filter_columns_by_keywords(df, keywords):
    """
    Filtre les colonnes d'un DataFrame en fonction de la présence de l'un des mots-clés spécifiés dans leur titre.

    Args :
        df (pd. DataFrame) : Le DataFrame d'entrée.
        Keywords (list) : Liste de chaînes de caractères pour filtrer les colonnes.

    Returns :
        pd. DataFrame : DataFrame avec uniquement les colonnes contenant l'un des mots-clés spécifiés.
    """
    df_cleaned = df.dropna(axis=1, how='all')
    # Crée une liste des colonnes à conserver
    columns_to_keep = [col for col in df_cleaned.columns if any(keyword in col for keyword in keywords)]

    # Filtre le DataFrame
    filtered_df = df_cleaned[columns_to_keep]

    return filtered_df


# #def columns_to_list(df):
#     """
#     Convertit les titres des colonnes d'un DataFrame en une liste de chaînes de caractères.
#
#     Args :
#         df (pd. DataFrame) : Le DataFrame d'entrée.
#
#     Returns :
#         list : Liste des titres de colonnes.
#     """
#     return df.columns.tolist()


def ajouter_colonne_modif(df, header1, header2, modif):
    # Vérifier si la colonne header1 existe dans le DataFrame
    if header1 in df.columns:
        # Créer une nouvelle colonne header2 en appliquant la fonction modif à chaque élément de la colonne header1
        df[header2] = df[header1].apply(modif)
    else:
        print(f"La colonne '{header1}' n'existe pas dans le DataFrame.")
    return df


def ajouter_colonne_multiplication(df, header1, header2, header3, factor):
    # Vérifier si les colonnes header1 et header2 existent dans le DataFrame
    if factor is None:
        factor = 1
    if header1 in df.columns and header2 in df.columns:
        # Créer une nouvelle colonne header3 en multipliant les colonnes header1 et header2
        df[header3] = df[header1] * df[header2] * factor
    else:
        print(f"Les colonnes '{header1}' et/ou '{header2}' n'existent pas dans le DataFrame.")
    return df


def colonne_multiplication(df, header1, header2, factor):
    # Vérifier si les colonnes header1 et header2 existent dans le DataFrame
    if factor is None:
        factor = 1
    if header1 in df.columns and header2 in df.columns:
        # multiplie les colonnes header1 et header2
        df[header2] = df[header1] * df[header2] * factor
    else:
        print(f"Les colonnes '{header1}' et/ou '{header2}' n'existent pas dans le DataFrame.")
    return df


def ajouter_colonne_z_wall(df, new_header, z_min, z_max):
    # new_header = 'Z Wall (m)'
    num_rows = len(df)
    interval = (z_max - z_min) / (num_rows - 1)
    # Créer une nouvelle colonne avec les valeurs z de z_min à z_max
    df[new_header] = [z_min + i * interval for i in range(num_rows)]
    df[new_header] = df[new_header].round(decimals=EnvB.unit_round('[m]'))


def char_thickness(z, df, protected, maxtime=0):
    # Initialiser la somme totale
    total_sum = 0
    header1 = 'h_UL (m)'
    if protected:
        header2 = 'Cumul e_char protected_LL (mm)'
        header3 = 'Cumul e_char protected_UL (mm)'
    else:
        header2 = 'Cumul e_char LL (mm)'
        header3 = 'Cumul e_char UL (mm)'

    # Itérer sur chaque ligne du dataframe
    if maxtime == 0:
        maxtime = len(df)
    for i in range(1, maxtime):
        if z < df.loc[i, header1]:
            total_sum += df.loc[i, header2] - df.loc[i - 1, header2]
        else:
            total_sum += df.loc[i, header3] - df.loc[i - 1, header3]
    return total_sum


def ajouter_colonne_char_z_wall(df, hz_wall, he_char, compartment1, protected, maxtime=0):
    # z_wall = 'Z Wall (m)'
    # e_char = 'e_char(z_wall) (mm)'
    if not protected:
        ajouter_colonne_z_wall(df, hz_wall, compartment1.origin[2], compartment1.height)
    num_rows = len(df)
    # Créer une nouvelle colonne avec les valeurs z de z_min à z_max
    df[he_char] = [char_thickness(df.loc[i, hz_wall], df, protected, maxtime) for i in range(num_rows)]


# Plafonne le calcul si max_char est atteint à l'instant t1 alors les valeurs de contribution (colonnes beta et val)
# pour t>t1 sont mises à zéro et la valeur de cumul reste à la valeur de t1
def max_charring_update(c_data, max_char, cumul, beta, val):
    filtered_data = c_data[c_data[cumul] > max_char]
    if not filtered_data.empty:
        index_sup = filtered_data.index[0]
        c_data.loc[index_sup + 1:, beta] = 0
        c_data.loc[index_sup + 1:, val] = 0
        c_data.loc[index_sup + 1:, cumul] = c_data.loc[index_sup, cumul]
        return True
    else:
        return False


def converge(c_iter, c_data, prev_data, racine_dir, comp):
    # New EC5 max_diff_sum_hrr = 1000  # (kW)
    # New EC5 max_diff_sum_hrr_percent = 0.01
    if c_iter == 0:  # or EnvB.current_zone_model[1] == 'PC':
        return False
    else:
        # New EC5 A.3(3) convergence criteria:
        # final design charring depth from one to the next iteration is not more than 0,5 mm.
        max_diff_char = 0.5
        # e1 = c_data['Cumul e_char column UL (mm)'].max() - c_data['Cumul e_char column UL (mm)'].max()
        # Trouver les colonnes contenant 'Cumul'
        cumul_columns = [col for col in c_data.columns if 'Cumul' in col]

        # Calculer la différence maximale
        if prev_data.empty:
            max_diff = max(c_data[col].max() for col in cumul_columns)
        else:
            max_diff = max(c_data[col].max() - prev_data[col].max() for col in cumul_columns)

        if not EnvB.current_zone_model[1] == 'PC':
            if c_iter == 1:  # Rmque idem prev_data.empty
                c_data['diff Q iter'] = c_data['New HRR (kW)'] - c_data['Unconstrained HRR (kW)']
            else:
                c_data['diff Q iter'] = c_data['New HRR (kW)'] - prev_data['New HRR (kW)']
            sum_diff = c_data['diff Q iter'].sum()
            sum_percent = sum_diff / c_data['New HRR (kW)'].sum()
            # print(f"Itération n° {c_iter - 1} différence total HRR {int(sum_diff)} kW
            # soit: {round (sum_percent * 100, 2)}%")
            # if (sum_diff < max_diff_sum_hrr) or (sum_percent < max_diff_sum_hrr_percent) \
            #         or c_iter > EnvB.max_iter :
            #     if sum_diff < max_diff_sum_hrr:
            #         print(f"convergence différence total HRR {int(sum_diff)} kW < {max_diff_sum_hrr} kW")
            #     if sum_percent < max_diff_sum_hrr_percent:
            #         print(f"convergence différence HRR total < {round (max_diff_sum_hrr_percent * 100, 2)}%")
            #     if c_iter > EnvB.max_iter:
            #         print(f"STOP ITERATION Nombre d'itérations supérieur à {EnvB.max_iter}")
            #         comp.auto_extinction = False
            #     if os.path.exists(racine_dir):
            #         shutil.rmtree(racine_dir)  # vide le répertoire créé pour la prochaine itération s'il existe
            #     return True
            print(f"Itération n° {c_iter - 1} différence total HRR {int(sum_diff)} "
                  f"kW soit: {round(sum_percent * 100, 2)}%")
        print(f"différence max épaisseur de carbonisation {round(max_diff, 2)}")
        if (max_diff < max_diff_char) or c_iter > EnvB.max_iter:
            if max_diff < max_diff_char:
                print(f"convergence différence {round(max_diff, 2)} < {max_diff_char}")
            if c_iter > EnvB.max_iter:
                print(f"STOP ITERATION Nombre d'itération supérieure à {EnvB.max_iter}")
                comp.auto_extinction = False
            if os.path.exists(racine_dir):
                shutil.rmtree(racine_dir)  # vide le répertoire créé pour la prochaine itération s'il existe
            return True
        else:
            return False


def free_mem(c_data):
    # Vider current_data et libérer la mémoire
    c_data.drop(c_data.index, inplace=True)
    del c_data
    gc.collect()
    # c_data = pd.DataFrame()


def fill_none_with_previous(df, header):
    """
    Remplit les valeurs None dans une colonne spécifiée avec la valeur de la première ligne précédente non None.
    :param df : DataFrame pandas
    :param header : Nom de la colonne à vérifier et remplir
    :return : DataFrame modifié
    """
    if header in df.columns:
        # Parcourir chaque ligne de la colonne spécifiée
        for i in range(1, len(df)):
            if pd.isna(df.iloc[i][header]):
                # Trouver la première valeur non None précédente
                for j in range(i - 1, -1, -1):
                    if not pd.isna(df.iloc[j][header]):
                        df.iloc[i, df.columns.get_loc(header)] = df.iloc[j][header]
                        break
    else:
        print(f"La colonne '{header}' n'existe pas dans le DataFrame.")

    return df


# End Operation on DataFrame
#########


#########
# Algorithm functions for wood contribution

def flux_from_ul_temp(x):
    sigma = 5.67e-11  # the Stefan-Boltzmann constant [kW/m2K4]
    epsilon = 0.8
    return epsilon * sigma * (x + 273) ** 4


def h_ul(x, h):
    return max(h - x, 1e-3)


def view_factor(h, w):
    tn = w * sy.atan(1 / w) + h * sy.atan(1 / h) - sqrt(w ** 2 + h ** 2) * sy.atan(1 / sqrt(w ** 2 + h ** 2))
    l1 = (((1 + h ** 2) * (1 + w ** 2)) / (1 + w ** 2 + h ** 2))
    l2 = (((h ** 2) * (1 + h ** 2 + w ** 2)) / ((1 + h ** 2) * (w ** 2 + h ** 2))) ** (h ** 2)
    l3 = (((w ** 2) * (1 + h ** 2 + w ** 2)) / ((1 + w ** 2) * (w ** 2 + h ** 2))) ** (w ** 2)
    ln = sy.log(l1 * l2 * l3)
    return (1 / (sy.pi * w)) * (tn + ln / 4)


def floor_view_factor(comp, hul):
    if hul <= 0:
        return 1
    x = comp.length / hul
    y = comp.width / hul
    a1 = sy.log(sqrt((1 + x ** 2) * (1 + y ** 2) / (1 + x ** 2 + y ** 2)))
    a2 = x * sqrt(1 + y ** 2) * sy.atan(x / sqrt(1 + y ** 2))
    a3 = y * sqrt(1 + x ** 2) * sy.atan(y / (1 + x ** 2)) - x * sy.atan(x) - y * sy.atan(y)
    return float(2 / (sy.pi * x * y) * (a1 + a2 + a3))


def phi_factor(comp, hul):
    f12 = view_factor(hul / comp.length, comp.width / comp.length)
    f21 = view_factor(hul / comp.width, comp.length / comp.width)
    return float(2 * (f12 + f21))


def charring_rate_law(q_incident, ro_w0):
    if q_incident < 0:
        return 0
    min_flux = 15
    valeur_min_flux = 0.374
    pente_origine = 5 * min_flux / ro_w0 + valeur_min_flux
    if q_incident > min_flux:
        return 5 * q_incident / ro_w0 + valeur_min_flux
    else:
        return pente_origine * q_incident / min_flux


# def charring_rate_ec5_parametric_curves_law(t, t0, d_char_t, coef):
#     if t <= round(t0):
#         return coef * d_char_t * t / round(t0)
#     else:
#         return 0
def charring_rate_ec5_law(t, t_max, t_end, d_char_t, coef):
    if t <= round(t_max):
        return coef * d_char_t
    elif t <= t_end:
        return coef * d_char_t * (t - t_end) / (t_max - t_end)
    else:
        return 0


def charring_ec5_law(t, beta_par, t0, coef):
    if t <= t0:
        return coef * beta_par * t
    elif t <= 3 * t0:
        return coef * beta_par * (1.5 * t - t ** 2 / (4 * t0) - t0 / 4)
    else:
        return coef * 2 * beta_par * t0


def charring_eff_ec5_law(t, beta_par, t_end, t_max, d0, coef):
    if t <= t_max:
        return coef * (beta_par * t + d0)
    elif t <= t_end and not t_end == t_max:
        r = beta_par * t_max + 0.5 * (-beta_par / (t_end - t_max)) * (t - t_max) ** 2 + beta_par * (t - t_max) + d0
        return coef * r
    else:
        return coef * (0.5 * beta_par * (t_end + t_max) + d0)


def ajoute_mlr_law(current, a, q, e, mp, m, v, mlr_f, ro_w0):
    # Constante c
    c = 1 / (ro_w0 * 0.75)  # Vous pouvez ajuster cette valeur selon vos besoins
    # Ajouter et initialiser les nouvelles colonnes
    current[mp] = 0.0
    current[m] = 0.0
    current[v] = 0.0
    current[e] = 0.0
    mlr_factor = 1

    # Calculer les valeurs des nouvelles colonnes
    for i in range(len(current)):
        if i == 0:
            current.at[i, e] = 0  # Initialisation à 0 pour la première ligne
        else:
            current.at[i, e] = (current.at[i - 1, e] + current.at[i - 1, v] / 60 *
                                (current.at[i, a] - current.at[i - 1, a]))
        current.at[i, mp] = mlr_f.valeur(current.at[i, e]) * mlr_factor
        current.at[i, m] = current.at[i, mp] * max(0, current.at[i, q])
        current.at[i, v] = current.at[i, m] * c * 60


def oxygen_reduction_factor(o_percent):
    min_percent = 11.5
    if o_percent < min_percent:
        return o_percent * 2.609 / 100 + 0.5
    else:
        return o_percent * 1.739 / 100 + 0.6


# Fonction pour calculer le produit de la cellule actuelle et de la cellule précédente
# Create char thickness from char speed and time (e= beta * t) beta=header2 t=header1
def e_char(row, df, header1, header2):
    if row.name == 0:
        return 0  # ou une autre valeur si vous préférez
    else:
        return (row[header1] / 60 - df.at[row.name - 1, header1] / 60) * (
                row[header2] + df.at[row.name - 1, header2]) / 2


def s_ingberg(row, df, header1, header2):
    if row.name == 0:
        return 0  # ou une autre valeur si vous préférez
    else:
        temp = ((row[header2] + df.at[row.name - 1, header2]) / 2 - ingberg_temp_seuil)
        return max((row[header1] - df.at[row.name - 1, header1]) * temp, 0)


# New EC5-1-2 Annex A Create char thickness from compartment temperature and  time (Te / t) Te=header2 t=header1
# def e_char_ec5(row, df, header1, header2):
#     if row.name == 0:
#         return 0  # ou une autre valeur si vous préférez
#     else:
#         t_mean = ((row[header2]+273)**2 + (df.at[row.name - 1, header2]+273)**2)/2
#         t2_integral = t_mean * (row[header1] / 60 - df.at[row.name - 1, header1] / 60)
#         return (t2_integral/1.35E5)**(1/1.6)
# !! Faux (a+b)**x différent a**x+b**x


# New EC5-1-2 Annex A Create char speed from char thickness and time (beta= e / t) e=header2 t=header1
def beta_char(row, df, header1, header2):
    if row.name == 0:
        return 0  # ou une autre valeur si vous préférez
    else:
        return (1 / (row[header1] / 60 - df.at[row.name - 1, header1] / 60)
                * (row[header2] + df.at[row.name - 1, header2]) / 2)


# Fonction kevin_2(T) = (T + 273)^2
def kelvin_2(t):
    return (t + 273) ** 2


# End Algorithm functions for wood contribution
#########


#########
# Calculation of effective cross-section (EC5-1-2 Annex A)

# noinspection SpellCheckingInspection
"""
# Construction du DataFrame courbe_ISO
# time_iso = np.arange(0, 121, 1)
# temp_Iso = 20 + 345 * np.log10(time_iso * 8 + 1)
# courbe_ISO = pd.DataFrame({'Time (min)': time_iso, 'Temp_Iso (C)': temp_Iso})
"""


def add_effective_char_depth_column(c_data, temp, char, zone, name, d0):
    d_ef = 'Effective charring depth ' + name + f" d0={u_round(d0, '[mm.]')}"
    c_rate = 'cooling rate ' + zone
    alpha_fi = 'alpha ' + zone
    # Trouver le temps pour lequel temp ('Upper Layer Temp (C)' par exemple) > 500 (flashover)
    flashover = c_data.loc[c_data[temp] > 500, 'Time (min)'].min()

    # Trouver le temps pour lequel temp est maximale (t_max)
    t_max = c_data.loc[c_data[temp].idxmax(), 'Time (min)']

    # Ajouter la colonne 'cooling rate' à c_data si elle n'a pas déjà été créée
    if c_rate not in c_data.columns:
        cooling_rate = []
        for i in range(len(c_data)):
            if i == 0 or c_data['Time (min)'][i] < t_max + 20:
                cooling_rate.append(np.nan)
            else:
                vi = abs((c_data[temp][i] - c_data[temp][i - 1]) /
                         (c_data['Time (min)'][i] - c_data['Time (min)'][i - 1]))
                cooling_rate.append(max(5, vi))
        c_data[c_rate] = cooling_rate

    # Ajouter la colonne 'alpha' à c_data si elle n'a pas déjà été créée
    if alpha_fi not in c_data.columns:
        alpha = []
        alpha_t_max = 1
        for i in range(len(c_data)):
            temp_i = c_data[temp][i]
            ti = c_data['Time (min)'][i]
            if ti <= t_max:
                temp_iso = 20 + 345 * np.log10((ti - flashover) * 8 + 1) if ti > flashover else 0
                # courbe_ISO.loc[courbe_ISO['Time (min)'] == ti, 'Temp_Iso (C)'].values[0]
                alpha_value = 2 if temp_i < temp_iso else 1
                alpha.append(alpha_value)
                if ti == t_max:
                    alpha_t_max = alpha_value
            elif ti >= t_max + 20:
                vi = c_data[c_rate][i]
                alpha.append(u_round(7 / (vi ** (1 / 3)), '[]'))
            else:
                # Interpolation linéaire entre t_max et t_max+20 pour la colonne 'alpha'
                if c_data['Time (min)'].iloc[-1] >= t_max + 20:  # i.e. if decay phase in calculation not > 20mn
                    t_fin = t_max + 20
                else:
                    t_fin = c_data['Time (min)'].iloc[-1]
                vi_t_max20 = c_data.loc[c_data['Time (min)'] >= t_fin, c_rate].values[0]
                alpha_interp = np.interp(ti, [t_max, t_fin], [alpha_t_max, 7 / (vi_t_max20 ** (1 / 3))])
                alpha.append(alpha_interp)
        c_data[alpha_fi] = alpha

    # Ajouter la colonne 'Effective charring depth' à c_data
    c_data[d_ef] = c_data[char] + c_data[alpha_fi] * d0
    c_data[d_ef] = c_data[d_ef].round(decimals=EnvB.unit_round('[mm.]'))


# End calculation of effective cross-section (EC5-1-2 Annex A)
#########

#########
#  Files management
def remove_files(chemin, f_extension):
    # Utilisez glob.glob pour trouver tous les fichiers d'extension f_extension dans le répertoire
    f_remove = glob.glob(os.path.join(chemin, f_extension))

    # Supprimez chaque fichier d'extension f_extension trouvé
    for f_to_remove in f_remove:
        os.remove(f_to_remove)
        print(f'Le fichier {f_to_remove} a été supprimé.')


# End of  Files management
#########


#############################################################################
# Implémentation méthode itérative originelle de l'article [1]

def superior(v, x):
    if isinstance(v, (int, float)):
        return v > x
    else:
        return False


def int_superior(v, x):
    if isinstance(v, int):
        return v > x
    else:
        return False


def init_experiment_list(essais: list) -> tuple[pd.DataFrame, list]:
    def mat(row_mat, row_thick):
        if row_mat is not None:
            num_mat = thermal_base[thermal_base.isin([row_mat]).any(axis=1)].index
            if not num_mat.empty:
                mat_value = thermal_base.iloc[num_mat[0]]
                return MaterialComposant.from_thermal(row_thick, mat_value)
        return None

    def dernier_not_none(ma_liste):
        if ma_liste[-1] is not None:
            return
        else:
            # Trouver l'index du dernier élément non-None
            dernier_index = None
            for i in range(len(ma_liste) - 1, -1, -1):
                if ma_liste[i] is not None:
                    dernier_index = i
                    break
            # Supprimer tous les éléments après ce dernier élément non-None
            if dernier_index is not None:
                del ma_liste[dernier_index + 1:]

    def is_inlist(ref, list_essais) -> bool:
        for mot in list_essais:
            if re.search(mot, ref, re.IGNORECASE):
                return True
        return False

    # # compartment_list : à partir de la ligne correspondante du fichier excel (recap_experiment) crée un nouveau
    # # compartiment à calculer et le rajoute à la liste retournée par init_experiment_list()
    # noinspection SpellCheckingInspection
    def compartment_list(cur_row):
        # if 'all' not in essais and not (cur_row.iloc[experiment_list.columns.get_loc('Ref')]) in essais:
        #     return
        if 'all' not in essais and not is_inlist(cur_row.iloc[experiment_list.columns.get_loc('Ref')], essais):
            return
        le = cur_row.iloc[experiment_list.columns.get_loc('Lr')]
        # Ou (équivalent) : le = row_values. Lr
        wi = cur_row.iloc[experiment_list.columns.get_loc('Wr')]
        he = cur_row.iloc[experiment_list.columns.get_loc('Hr')]
        te = ti = cur_row.iloc[experiment_list.columns.get_loc('Tamb')]
        wb = cur_row.iloc[experiment_list.columns.get_loc('wb')]
        hb = cur_row.iloc[experiment_list.columns.get_loc('hb')]
        wco = cur_row.iloc[experiment_list.columns.get_loc('wco')]
        hco = cur_row.iloc[experiment_list.columns.get_loc('hco')]
        if isinstance(hb, (int, float)) and isinstance(wb, (int, float)):
            beam_thick = min(wb, hb)
        else:
            beam_thick = 0
        if isinstance(hco, (int, float)) and isinstance(wco, (int, float)):
            column_thick = min(wco, hco)
        else:
            column_thick = 0
        if te is None:
            te = ti = 17
        ha = cur_row.iloc[experiment_list.columns.get_loc('Hamb')]
        if not isinstance(ha, (int, float)) or ha < EnvB.min_humidity_brisk:
            ha = EnvB.min_humidity_brisk
        tstep = cur_row.iloc[experiment_list.columns.get_loc('tstep')]
        if not isinstance(tstep, (int, float)) or tstep > 1:
            tstep = 1
        erc = cur_row.iloc[experiment_list.columns.get_loc('ercontrol')]
        if isinstance(erc, (int, float)):
            erc = min(erc, 0.1)
        else:
            erc = 0.1
        erv = cur_row.iloc[experiment_list.columns.get_loc('ervent')]
        if isinstance(erv, (int, float)):
            erv = min(erv, 1e-3)
        else:
            erv = 1e-3
        interval = cur_row.iloc[experiment_list.columns.get_loc('interval')]
        if not isinstance(interval, (int, float)) or interval < tstep:
            interval = EnvB.default_excel_interval
        if isinstance(cur_row.tchar_fin, (int, float)):
            tchar_fin = cur_row.tchar_fin * 60
        else:
            tchar_fin = EnvB.default_total_time
        if isinstance(cur_row.dexp, (int, float)):
            tfin = cur_row.dexp * 60
        else:
            tfin = EnvB.default_total_time
        if isinstance(cur_row.db, (int, float)):
            db = cur_row.db
        else:
            db = EnvB.default_wood_density
        if isinstance(cur_row.dco, (int, float)):
            dco = cur_row.dco
        else:
            dco = EnvB.default_wood_density
        if isinstance(cur_row.Patm, (int, float)):
            pr = cur_row.Patm
        else:
            pr = EnvB.default_pressure
        if cur_row.pfire is not None and cur_row.pfire == 'Rear':
            fire = Fire(0, cur_row.Ref, le / 2, wi - 0.3, 0)
        else:
            fire = None
        if all(superior(x, 0) for x in [le, wi, he]):
            new_compartment = Compartment(nom=cur_row.iloc[experiment_list.columns.get_loc('Ref')],
                                          length=le, width=wi, height=he,
                                          wall=cur_row.iloc[experiment_list.columns.get_loc('Wexposed')],
                                          ceiling=cur_row.iloc[experiment_list.columns.get_loc('Cexposed')],
                                          column=cur_row.iloc[experiment_list.columns.get_loc('Coexposed')],
                                          beam=cur_row.iloc[experiment_list.columns.get_loc('Bexposed')],
                                          floor=cur_row.iloc[experiment_list.columns.get_loc('Fexposed')],
                                          interior=ti, exterior=te, humid=ha, time_step=tstep, t_char_fin=tchar_fin,
                                          beam_thick=beam_thick, column_thick=column_thick, bwd=db, co_wd=dco,
                                          error_control=erc, error_vent=erv, excel_interval=interval, pressure=pr,
                                          fire=fire)
        else:
            # ecriture en Rouge
            print(
                f"!! no dimension for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
            return
        nombre_vents = cur_row.iloc[experiment_list.columns.get_loc('Nv')]
        if not int_superior(nombre_vents, 0):
            print(
                f"!! no vents for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
            return
        cur_vent = 1
        for i in range(nombre_vents):
            n_vent = cur_row.iloc[experiment_list.columns.get_loc('Nv' + str(i + 1))]
            w = cur_row.iloc[experiment_list.columns.get_loc('wv' + str(i + 1))]
            h = cur_row.iloc[experiment_list.columns.get_loc('hv' + str(i + 1))]
            s = cur_row.iloc[experiment_list.columns.get_loc('sv' + str(i + 1))]
            o = cur_row.iloc[experiment_list.columns.get_loc('o_hv' + str(i + 1))]
            f = cur_row.iloc[experiment_list.columns.get_loc('Fv' + str(i + 1))]
            if isinstance(n_vent, int) and n_vent > 0:
                d = cur_row.iloc[experiment_list.columns.get_loc('dv' + str(i + 1))]
                for j in range(n_vent + 1):
                    new_compartment.vents.append(Vent(num=cur_vent, room1=new_compartment.id, room2='OUTSIDE', width=w,
                                                      height=h, sill=s, offset=o + j * (w + d), face=f))
                    cur_vent = cur_vent + 1
            else:
                new_compartment.vents.append(Vent(num=cur_vent, room1=new_compartment.id, room2='OUTSIDE',
                                                  width=w, height=h, sill=s, offset=o, face=f))
                cur_vent = cur_vent + 1
            if cur_vent > nombre_vents:
                break
        new_compartment.calc_opening_factor()
        calculated = False
        hf = cur_row.iloc[experiment_list.columns.get_loc('hrrfuel')]  # ou cur_row_values.hrrfuel
        hc = cur_row.iloc[experiment_list.columns.get_loc('Hc')]  # in present version, not use with BRisk, see set_fuel
        if hc is None:
            hc = EnvB.default_structural_wood_heat_of_combustion
        if hf is not None:
            hf = EnvB.CurPath + '\\input\\' + hf
            if os.path.exists(hf):
                test_open_write(hf)
                hrr_f = extract_dataframe(hf, new_compartment.id)
                rowf, colf = find_cell(hrr_f, 't [mn]')
                if rowf is not None:
                    hrr_f = hrr_f.iloc[rowf:, colf:colf + 3]
                    nf_columns = hrr_f.iloc[0].values
                    hrr_f.columns = nf_columns
                    hrr_f = hrr_f.iloc[1:, 0:]
                    abscises = hrr_f['t [s]'].tolist()
                    dernier_not_none(abscises)
                    ordonnees = hrr_f['HRRFuel'].tolist()
                    dernier_not_none(ordonnees)
                    f_hrr = HrrFunction(abscises, ordonnees)
                    new_compartment.hrr_fuel = f_hrr
                    new_compartment.set_fuel(cur_row.Ft, hc)
                    new_compartment.fire.fd = cur_row.Fd
                    print('Hrr Fuel from ' + hf + ' ok')
                    calculated = True

        if isinstance(cur_row.Wdir, (int, float)) and EnvB.allow_wind:
            d_wind = u_round(cur_row.Wdir, '[°]')
        else:
            d_wind = 0
        if isinstance(cur_row.Vwind, (int, float)) and EnvB.allow_wind:
            v_wind = u_round(cur_row.Vwind, '[m/s]')
            new_compartment.fire.wind = [v_wind, d_wind, 1]
        else:
            v_wind = 0
        # seifbois_b1_fd = (['SeifBois B1', (30, 1.0208), (60, 2.27), (90, 3.55), (120, 4.91)])
        # seifbois_b1_brandon = (['SeifBois B1', (30, 0.9, 0.9), (60, 2.08, 2), (90, 3.3, 3), (120, 4.65, 4)])
        if not calculated:
            if new_compartment.calculate_hrr_fuel(cur_row.Fd, tfin, v_wind, d_wind):
                # if new_compartment.calculate_brandon_hrr_fuel(cur_row.Fd, tfin, v_wind, d_wind, 4.65, 4):
                new_compartment.set_fuel(cur_row.Ft, hc)
                print('Hrr Fuel ok')
            else:
                print(
                    f"!! no Hrr Fuel for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
                return
        h_f = cur_row.iloc[experiment_list.columns.get_loc('hfire')]
        if isinstance(h_f, (int, float)):
            new_compartment.fire.elevation = h_f
        # # add materials
        mat1 = mat(cur_row.Wl_mat, cur_row.Wl_thick)
        if mat1 is None:
            print(
                f"!! no wall material for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
            return
        else:
            mat_s1 = mat(cur_row.Ws1_mat, cur_row.Ws1_thick)
            mat_s2 = mat(cur_row.Ws2_mat, cur_row.Ws2_thick)
            mat_wall = Material('wall', mat1, mat_s1, mat_s2)
            dw = cur_row.dw
            if dw is not None and isinstance(dw, (int, float)):
                new_compartment.wall_wood_density = dw
            print(mat_wall)
        mat1 = mat(cur_row.Cl_mat, cur_row.Cl_thick)
        if mat1 is None:
            print(
                f"!! no Ceiling material for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
            return
        else:
            mat_s1 = mat(cur_row.Cs1_mat, cur_row.Cs1_thick)
            mat_s2 = mat(cur_row.Cs2_mat, cur_row.Cs2_thick)
            mat_ceiling = Material('ceiling', mat1, mat_s1, mat_s2)
            dc = cur_row.dc
            if dc is not None and isinstance(dc, (int, float)):
                new_compartment.ceiling_wood_density = dc
            print(mat_ceiling)
        mat1 = mat(cur_row.Fl_mat, cur_row.Fl_thick)
        if mat1 is None:
            print(
                f"!! no Floor material for \033[1;31m{cur_row.iloc[experiment_list.columns.get_loc('Ref')]}\033[0m")  # noqa: E231, E702
            return
        else:
            mat_s1 = mat(cur_row.Fs1_mat, cur_row.Fs1_thick)
            mat_s2 = mat(cur_row.Fs2_mat, cur_row.Fs2_thick)
            mat_floor = Material('floor', mat1, mat_s1, mat_s2)
            df = cur_row.df
            if df is not None and isinstance(df, (int, float)):
                new_compartment.floor_wood_density = df
            print(mat_floor)
        new_compartment.wall_mat = mat_wall
        new_compartment.ceiling_mat = mat_ceiling
        new_compartment.floor_mat = mat_floor
        if EnvB.allow_contribution_protected:
            new_compartment.wall_protected = mat_wall.is_wood_protected()
            new_compartment.ceiling_protected = mat_ceiling.is_wood_protected()
            new_compartment.floor_protected = mat_floor.is_wood_protected()
            tc = cur_row.t_ceiling_gyps_fo
            if tc is not None and isinstance(tc, (int, float)):
                new_compartment.ceiling_time_fo = tc
            tw = cur_row.t_wall_gyps_fo
            if tw is not None and isinstance(tw, (int, float)):
                new_compartment.wall_time_fo = tw
        # TODO Beam & Column protected
        if cur_row.scond is not None:
            if cur_row.scond == 'Only lining involve':
                new_compartment.lining_involved_only = True
            if isinstance(cur_row.scond, (float, int)):
                new_compartment.floor_bloc = cur_row.scond

        new_compartment.wood_thickness()
        new_compartment.calculate_ec_curves_parameters()
        # ### End Add material

        list_compartment.append(new_compartment)
        # print(f"Traitement de la ligne : {cur_row}")
        print(new_compartment)

    # ## End compartment_List

    thermal_base = extract_dataframe(EnvB.Thermal_file, 'Thermal')
    list_compartment = []
    experiment_list = extract_dataframe(EnvB.experiment_excel_file, 'Recap')
    # print(experiment_list)
    row, col = find_cell(experiment_list, 'Inst')
    if row is not None:
        experiment_list = experiment_list.iloc[row:, col:]
        new_columns = experiment_list.iloc[0].values
        experiment_list.columns = new_columns
        row = find_ligne(experiment_list, 'TOTAL')
        if row is not None:
            experiment_list = experiment_list.iloc[:row + 1]
        remove_column_none(experiment_list)
        for colonne in experiment_list.columns:
            experiment_list = arrondir_colonne(experiment_list, colonne)
        experiment_list = fill_none_with_previous(experiment_list, 'Fd')
        experiment_list = fill_none_with_previous(experiment_list, 'Af')
        experiment_list = fill_none_with_previous(experiment_list, 'At(b)')
        # noinspection SpellCheckingInspection
        experiment_list = fill_none_with_previous(experiment_list, 'Aexp/At')
        experiment_list = fill_none_with_previous(experiment_list, 'hv')
        experiment_list = fill_none_with_previous(experiment_list, 'Av')
        experiment_list = fill_none_with_previous(experiment_list, 'O(b)')
        experiment_list = fill_none_with_previous(experiment_list, 'Ft')
        experiment_list = fill_none_with_previous(experiment_list, 'Hc')
        print(experiment_list)

    experiment_char = extract_dataframe(EnvB.experiment_excel_file, 'Char Depth Graph')
    row, col = find_cell(experiment_char, 'Reference')
    # Créez une liste avec les éléments non vides de la colonne à partir de cur_row+1
    # result_list = df['Colonne'][cur_row + 1:].#dropna().tolist()
    result_list = experiment_char.iloc[row + 1:, col].dropna().tolist()
    EnvB.result_list_char = experiment_char.iloc[:row - 1:, col].dropna().tolist()
    current_date = date.today()
    EnvB.current_experiment_excel_file = EnvB.CurPath + '\\' + current_date.strftime(
        "%Y%m%d") + EnvB.Results_file
    if not os.path.exists(EnvB.current_experiment_excel_file):
        shutil.copy(EnvB.experiment_excel_file, EnvB.current_experiment_excel_file)
    # Parcourir les éléments de la liste
    for element in result_list:
        # Trouver l'index de la ligne contenant l'élément
        row_index = experiment_list[experiment_list.isin([element]).any(axis=1)].index
        if not row_index.empty:
            # Appeler la fonction avec la ligne du DataFrame
            # row_values = experiment_list.loc[row_index[0]]
            # compartment_list(row_values)
            compartment_list(experiment_list.loc[row_index[0]])
            # Attention, loc(r) i.e la ligne d'index r plutôt que #iloc(i) i_ème ligne
    print(experiment_char)

    return experiment_list, list_compartment


courbe_iso = cumul_ingberg_iso = mlr_function = TimeFunction([0, 1], [0, 1])


def init_courbe_iso_and_mlr():
    global mlr_function, courbe_iso, cumul_ingberg_iso
    f = extract_dataframe(EnvB.courbe_iso_file, 'Feuil1')
    temp = np.array(f['courbe ISO'].tolist())
    courbe_iso = TimeFunction(f['t min'].tolist(), temp)
    f['Ingberg surface'] = f.apply(
        lambda row: s_ingberg(row, f, 't min', 'courbe ISO'), axis=1)
    f['Ingberg surface equivalency (C.min)'] = f['Ingberg surface'].cumsum()
    cumul_ingberg_iso = TimeFunction(f['t min'].tolist(), f['Ingberg surface equivalency (C.min)'].tolist())
    if EnvB.default_method == 'Mlr':
        mlr_f = extract_dataframe(EnvB.mlr_file, 'MLR Function')
        # mlr_function = TimeFunction(mlr_f['e (mm)'].tolist(), mlr_f['MLR* (g kJ-1)'].tolist())
        mlr_function = TimeFunction(mlr_f['e (mm)'].tolist(), mlr_f['Moyenne+2xSD (g/kJ)'].tolist())
    else:
        mlr_function = None


def boucle_calculs(compartment1: Compartment):
    start_time = time.time()
    print("Heure de début:", time.strftime("%H:%M:%S", time.localtime(start_time)))

    current_iter = 0
    current_data = pd.DataFrame()
    hrr_fuel = pd.DataFrame()
    hrr_initial = None
    prev_iter_data = pd.DataFrame()
    fire_name = ''
    max_char_wall = compartment1.wall_wood_thick
    max_char_ceiling = compartment1.ceiling_wood_thick
    max_char_floor = compartment1.floor_wood_thick
    max_char_beam = compartment1.beam_thick
    max_char_column = compartment1.column_thick

    # if EnvB.default_method == 'Mlr':
    #     mlr_f = extract_dataframe(EnvB.mlr_file, 'MLR Function')
    #     # mlr_function = TimeFunction(mlr_f['e (mm)'].tolist(), mlr_f['MLR* (g kJ-1)'].tolist())
    #     mlr_function = TimeFunction(mlr_f['e (mm)'].tolist(), mlr_f['Moyenne+2xSD (g/kJ)'].tolist())
    # else:
    #     mlr_function = None

    EnvB.ModelName = f"{compartment1.id}"
    EnvB.ModelPaths = EnvB.CurPath + '\\' + EnvB.ModelName
    EnvB.long_line.append(EnvB.ModelPaths)
    destination_racine_dir = EnvB.ModelPaths + '\\' + f"{EnvB.BaseModel}{current_iter + 1}"
    destination_dir = destination_racine_dir
    while not converge(current_iter, current_data, prev_iter_data, destination_racine_dir, compartment1):
        current_model = f"{EnvB.BaseModel}{current_iter}"
        base_model_path = EnvB.ModelPaths + '\\' + current_model
        if current_iter > 0:
            # New EC5 prev_hrr = current_data[['New HRR (kW)']].copy()
            free_mem(prev_iter_data)
            prev_iter_data = current_data.copy()

        if EnvB.current_zone_model[1] == 'Brisk':
            list_of_models = os.listdir(base_model_path)
            sub_model_folder_path = base_model_path + '\\' + list_of_models[0]
            # print(base_model_path)
            print(f"Répertoire de calcul: {sub_model_folder_path}")

            if current_iter == 0:
                print(compartment1)
                remove_files(sub_model_folder_path, '*.csv')
                remove_files(sub_model_folder_path, '*.txt')
                remove_files(sub_model_folder_path, '*.pdf')
                remove_files(sub_model_folder_path, '*.xlsx')
                remove_files(sub_model_folder_path, 'output1.xml')
                # noinspection SpellCheckingInspection
                remove_files(sub_model_folder_path, 'dumpdata.dat')

            # Copie des fichiers .xlm et .dat pour l'itération suivante
            destination_racine_dir = EnvB.ModelPaths + '\\' + f"{EnvB.BaseModel}{current_iter + 1}"
            if not os.path.exists(destination_racine_dir):
                os.makedirs(destination_racine_dir)  # Créer le répertoire de destination s'il n'existe pas
            else:
                shutil.rmtree(destination_racine_dir)  # vide le répertoire de destination s'il existe
                os.makedirs(destination_racine_dir)
            destination_dir = destination_racine_dir + '\\' + f"{EnvB.BaseModel}{current_iter + 1}"
            if not os.path.exists(destination_dir):
                os.makedirs(destination_dir)
            for file_name in os.listdir(sub_model_folder_path):  # Parcourir les fichiers dans le répertoire source
                if file_name.endswith('.xml') or file_name.endswith('.dat'):
                    # Construire les chemins complets source et destination
                    source_file = os.path.join(sub_model_folder_path, file_name)
                    destination_file = os.path.join(destination_dir, file_name)
                    # Copier le fichier
                    shutil.copy2(source_file, destination_file)
                    # print(f'Copié: {file_name}')
            os.rename(destination_dir + '\\' + current_model + '.xml',
                      destination_dir + '\\' + f"{EnvB.BaseModel}{current_iter + 1}.xml")
            # (f"Tous les fichiers .xml et .dat ont été copiés dans{destination_dir}")

            # lancement du calcul pour l'itération courante
            # processus =
            subprocess.Popen([EnvB.BriskPath, base_model_path])

            base_model_time = 0
            while not os.path.exists(sub_model_folder_path + '\\' + 'output1.xml'):
                # while the output isn't produced ye#t
                base_model_time = + 1  # just something to make the code wait for B-RISK
            while not os.path.exists(sub_model_folder_path + '\\' + current_model + '_results.xlsx'):
                # while the results isn't produced ye#t
                base_model_time = + 1  # just something to make the code wait for B-RISK
            while not os.path.exists(sub_model_folder_path + '\\' + current_model + '_zone.csv'):
                # while the zone isn't produced ye#t
                base_model_time = + 1  # just something to make the code wait for B-RISK
            print(base_model_time)
            time.sleep(1)  # just waiting to make sure no data is lost or corrupted
            os.system(r"taskkill /F /IM BRISK.exe")

            # Warning kill all excel...
            # if EnvB.Machine == 'Fanfan2':
            #     os.system(r"taskkill /F /IM EXCEL.exe")
            # input("Appuyez sur n'importe quelle touche pour continuer...")

            # TODO read output.xml and check 'flag stop'
            output_file = sub_model_folder_path + '\\' + 'output1.xml'
            # noinspection SpellCheckingInspection
            compartment1.calculated_ok = find_attribute_value(output_file, 'flagstop')
            if compartment1.calculated_ok == '1':
                print(f"\n \033[1;31mLe programme Brisk s'est terminé prématurément\033[0m")
                if current_iter > 0:
                    write_xls_char_results(current_data, compartment1.id, compartment1.time_char_fin / 60,
                                           compartment1.model_extinction(), True, compartment1.warning)
                compartment1.dump_to_first_sheet_xlsx(hrr_initial)
                return

            # Lecture du fichier excel de résultat et création du dataframe de calcul Curent_Iteration
            excel_file = sub_model_folder_path + '\\' + current_model + '_results.xlsx'
            current_data = extract_dataframe(excel_file, 'Room 1')
            # time_of_flashover = find_attribute_value(output_file, 'time_at_FO')
        elif EnvB.current_zone_model[1] == 'Cfast':
            print(f"Répertoire de calcul: {base_model_path}")

            if current_iter == 0:
                print(compartment1)
                remove_files(base_model_path, '*.csv')
                remove_files(base_model_path, '*.log')
                remove_files(base_model_path, '*.out')
                remove_files(base_model_path, '*.status')
                remove_files(base_model_path, '*.plt')
                remove_files(base_model_path, '*.smv')
                hrr_initial = compartment1.hrr_fuel

            # Copie des fichiers .in pour l'itération suivante
            destination_racine_dir = EnvB.ModelPaths + '\\' + f"{EnvB.BaseModel}{current_iter + 1}"
            destination_dir = destination_racine_dir
            if not os.path.exists(destination_dir):
                os.makedirs(destination_dir)  # Créer le répertoire de destination s'il n'existe pas
            else:
                shutil.rmtree(destination_dir)  # vide le répertoire de destination s'il existe
                os.makedirs(destination_dir)

            # Lancer le programme avec Popen
            processus = subprocess.Popen(
                [EnvB.CfastPath, base_model_path + '\\' + current_model],
                stdout=subprocess.PIPE,  # Capturer la sortie standard
                stderr=subprocess.PIPE,  # Capturer la sortie d'erreur
                text=True  # Interpréter les sorties en texte
            )

            # Lire la sortie standard au fur et à mesure
            cfast_error = False
            for ligne in processus.stdout:
                print(ligne, end='')  # Affiche la sortie au fur et à mesure
                if re.search('Error', ligne, re.IGNORECASE):
                    cfast_error = True
            for ligne in processus.stderr:
                print(f"\n \033[1;31m{ligne}\033[0m", end='')  # Affiche la sortie au fur et à mesure

            # Attendre que le processus se termine
            processus.wait()

            # Vérifier le code de retour
            if (processus.returncode == 0) and not cfast_error:
                print("\nLe programme Cfast s'est terminé avec succès.")
            else:
                print(f"\n \033[1;31mLe programme CFast s'est terminé prématurément avec le code de retour "
                      f"{processus.returncode}\033[0m")
                compartment1.calculated_ok = '1'
                if current_iter > 0:
                    write_xls_char_results(current_data, compartment1.id, compartment1.time_char_fin / 60,
                                           compartment1.model_extinction(), True, compartment1.warning)
                compartment1.dump_to_first_sheet_xlsx(hrr_initial)
                return

            # Lecture du fichier excel de résultat et création du dataframe pour calcul Curent_Iteration
            csv_file = base_model_path + '\\' + current_model + '_compartments.csv'
            # excel_file = base_model_path + '\\' + current_model + '_zone.xlsx'
            # convert_csv_to_xlsx(csv_file, excel_file)
            # current_data = csv_to_dataframe(csv_file)
            current_data = pd.read_csv(csv_file)
            current_data = current_data.iloc[3:]  # Warning, first line index is 3
            current_data = current_data.reset_index(drop=True)
            current_data.rename(columns=EnvB.mapping_cfast_brisk, inplace=True)

        # iteration for EC5-1-2 Annex A: A.4.4 Design model for parametric temperature‐time curves
        # TODO check if it is better to keep independent or to insert in next part (if current_data is not None)
        elif EnvB.current_zone_model[1] == 'PC':
            print(f"Répertoire de calcul: {base_model_path}")
            if current_iter == 0:
                print(compartment1)
                hrr_initial = compartment1.hrr_fuel
                tf = int(compartment1.exposed_time() / 60)
                current_data['Time (min)'] = range(0, tf + 1)
                current_data['Time (sec)'] = current_data['Time (min)'] * 60
                column_factor = 1 * int(compartment1.a_column_exposed > 0)
                beam_factor = 1 * int(compartment1.a_beam_exposed > 0)
                ceiling_factor = 1 * int(compartment1.a_ceiling_exposed > 0)
                wall_factor = 1 * int(compartment1.a_wall_exposed > 0)
                floor_factor = 1 * int(compartment1.a_floor_exposed > 0)
                beta_n1 = Ec.beta_n(False)
                beta_n2 = Ec.beta_n(True)
                # Brandon annex value
                # qd_fi = 550 * compartment1.a_f / compartment1.a_t
                qd_fi = compartment1.fire.fd * compartment1.a_f / compartment1.a_t
                # d_char_t1_prev = 0
                # d_char_t2_prev = 0
                qd_st = 0
                qd_tot_t = qd_fi
            # noinspection PyUnboundLocalVariable
            compartment1.calculate_ec_curves_parameters(qd_st)
            compartment1.add_parametric_temperature_time_curve(current_data, qd_st)
            t_temp_end = compartment1.t_end
            t_temp_max = compartment1.t_max
            if current_iter == 0:
                t1_max = t_temp_max

            if EnvB.parametric_curves == 'B':
                beta_par1 = beta_par2 = float(1.5 * 0.67 * ((0.2 * sqrt(compartment1.Gamma) - 0.04) /
                                                            (0.16 * sqrt(compartment1.Gamma) + 0.08)))
                # noinspection PyUnboundLocalVariable
                beta_par_eff_1 = beta_n1 * compartment1.Gamma ** 0.28  # [mm/mn] (A.9)
                # noinspection PyUnboundLocalVariable
                beta_par_eff_2 = beta_n2 * compartment1.Gamma ** 0.28  # TODO circular sections
                compartment1.d_sect_eff_surf = 0.5 * beta_par_eff_1 * (t_temp_end + t_temp_max)
                compartment1.d_sect_eff_lin = 0.5 * beta_par_eff_2 * (t_temp_end + t_temp_max)
                d0 = 0
            else:
                beta_par1 = beta_n1 * compartment1.Gamma ** 0.25  # [mm/mn] (A.9)
                beta_par2 = beta_n2 * compartment1.Gamma ** 0.25  # TODO circular sections
                beta_par_eff_1 = beta_par1
                beta_par_eff_2 = beta_par2
                d0 = 8 + compartment1.Gamma / 50 - min(compartment1.Gamma, 9) ** 2 / 20
                compartment1.d_sect_eff_surf = 0.5 * beta_par1 * (t_temp_end + t_temp_max) + d0
                compartment1.d_sect_eff_lin = 0.5 * beta_par2 * (t_temp_end + t_temp_max) + d0

            compartment1.beta_par_surf = beta_par1
            compartment1.beta_par_lin = beta_par2
            # noinspection PyUnboundLocalVariable
            t0 = 0.009 * qd_tot_t / compartment1.opening_factor
            compartment1.t_constant_char = t0
            d_char_t1 = 2 * beta_par1 * t0
            d_char_t2 = 2 * beta_par2 * t0
            a_surf_exposed = compartment1.a_ceiling_exposed + compartment1.a_wall_exposed + compartment1.a_floor_exposed
            a_lin_exposed = compartment1.a_beam_exposed + compartment1.a_column_exposed
            if EnvB.parametric_curves == 'B':
                # noinspection PyUnboundLocalVariable
                qd_st = (a_surf_exposed + a_lin_exposed) * 5.39 * (
                        d_char_t1 - 0.7 * beta_par1 * t1_max) / compartment1.a_t
            else:
                s_10 = 120  # * 17.5 / compartment1.heat_of_combustion
                # TODO check for O2 reduction (of to calculate it in parametric curve case)
                #  if EnvB.allow_char_energy_storage
                a_st = 1
                m = 0.8  # EC1-2 E.5(2)
                qd_st = (m * 60 * s_10 * a_st * (d_char_t1 * a_surf_exposed + d_char_t2 * a_lin_exposed)
                         / compartment1.a_t / 1000)
            # noinspection PyUnboundLocalVariable
            qd_tot_t = qd_fi + qd_st
            # d_char_t1_prev = d_char_t1
            # d_char_t2_prev = d_char_t2
            # d_char_res = max(d_char_t1 - d_char_t1_prev, d_char_t2 - d_char_t2_prev)
            current_data['Cumul e_char_eff ceiling (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_eff_ec5_law(x, beta_par_eff_1, t_temp_end, t_temp_max, d0, ceiling_factor)))
            current_data['Cumul e_char_eff wall (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_eff_ec5_law(x, beta_par_eff_1, t_temp_end, t_temp_max, d0, wall_factor)))
            current_data['Cumul e_char_eff floor (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_eff_ec5_law(x, beta_par_eff_1, t_temp_end, t_temp_max, d0, floor_factor)))
            current_data['Cumul e_char_eff beam (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_eff_ec5_law(x, beta_par_eff_2, t_temp_end, t_temp_max, d0, beam_factor)))
            current_data['Cumul e_char_eff column (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_eff_ec5_law(x, beta_par_eff_2, t_temp_end, t_temp_max, d0, column_factor)))
            current_data['Cumul e_char ceiling (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_ec5_law(x, beta_par1, t0, ceiling_factor)))
            current_data['Cumul e_char wall (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_ec5_law(x, beta_par1, t0, wall_factor)))
            current_data['Cumul e_char floor (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_ec5_law(x, beta_par1, t0, floor_factor)))
            current_data['Cumul e_char beam (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_ec5_law(x, beta_par2, t0, beam_factor)))
            current_data['Cumul e_char column (mm)'] = (current_data['Time (min)'].apply(
                lambda x: charring_ec5_law(x, beta_par2, t0, column_factor)))
            current_data['e_char ceiling (mm)'] = current_data['Cumul e_char ceiling (mm)'].diff().fillna(
                current_data['Cumul e_char ceiling (mm)'])
            current_data['e_char wall (mm)'] = current_data['Cumul e_char wall (mm)'].diff().fillna(
                current_data['Cumul e_char wall (mm)'])
            current_data['e_char floor (mm)'] = current_data['Cumul e_char floor (mm)'].diff().fillna(
                current_data['Cumul e_char floor (mm)'])
            current_data['e_char beam (mm)'] = current_data['Cumul e_char beam (mm)'].diff().fillna(
                current_data['Cumul e_char beam (mm)'])
            current_data['e_char column (mm)'] = current_data['Cumul e_char column (mm)'].diff().fillna(
                current_data['Cumul e_char column (mm)'])
            current_data['Cumul e_char protected_wall (mm)'] = 0
            current_data['Cumul e_char protected_ceiling (mm)'] = 0
            # current_data = ajouter_colonne_modif(current_data, 'Time (min)', 'Cumul e_char ceiling (mm)',
            #                                      lambda x: charring_rate_ec5_parametric_curves_law(x, t0, d_char_t1,
            #                                                                                        ceiling_factor))
            # current_data['e_char ceiling (mm)'] = current_data['Cumul e_char ceiling (mm)'].diff().fillna(
            #     current_data['Cumul e_char ceiling (mm)'])

            # écriture des fichiers excel d'itération et de synthèse (avec graphiques)
            temps = 'Time (min)'
            if current_iter == 0:
                write_xls_from_dataframe(current_data, current_iter, [(temps, 'Cumul e_char_eff wall (mm)'),
                                                                      (temps, 'Cumul e_char_eff ceiling (mm)'),
                                                                      (temps, 'Cumul e_char_eff column (mm)'),
                                                                      (temps, 'Cumul e_char_eff beam (mm)'),
                                                                      (temps, 'Cumul e_char wall (mm)'),
                                                                      (temps, 'Cumul e_char ceiling (mm)'),
                                                                      (temps, 'Cumul e_char column (mm)'),
                                                                      (temps, 'Cumul e_char beam (mm)'),
                                                                      (temps, 'Temp EC param curve (C)'), ],
                                         compartment1, True)
            else:
                add_data_and_series(current_data, current_iter, [(temps, 'Cumul e_char_eff wall (mm)'),
                                                                 (temps, 'Cumul e_char_eff ceiling (mm)'),
                                                                 (temps, 'Cumul e_char_eff column (mm)'),
                                                                 (temps, 'Cumul e_char_eff beam (mm)'),
                                                                 (temps, 'Cumul e_char wall (mm)'),
                                                                 (temps, 'Cumul e_char ceiling (mm)'),
                                                                 (temps, 'Cumul e_char column (mm)'),
                                                                 (temps, 'Cumul e_char beam (mm)'),
                                                                 (temps, 'Temp EC param curve (C)')], compartment1,
                                    True)

            current_iter = current_iter + 1

        # End iteration for EC5-1-2 Annex A: A.4.4 Design model for parametric temperature‐time curves

        if current_data is not None and not EnvB.current_zone_model[1] == 'PC':
            current_data = filter_columns_by_keywords(current_data, EnvB.liste_lue)
            # Convertir toutes les valeurs en float, remplacer les non-convertibles par zéro
            current_data = current_data.apply(pd.to_numeric, errors='coerce')
            missing_values_count = current_data.isna().sum().sum()
            current_data = current_data.fillna(0).astype(float)
            if missing_values_count > 0:
                print(f"\033[1;33mAvertissement\033[0m : {missing_values_count} valeurs de "
                      f"du fichier de résultat dans {base_model_path}"  # {excel_file}"
                      f" ont été remplacées par 0.")
            if EnvB.current_zone_model[1] == 'Cfast':
                # Results Cfast (W) -> (kW)
                current_data['Unconstrained HRR (kW)'] = current_data['Unconstrained HRR (kW)'] / 1000
                current_data['HRR (kW)'] = current_data['HRR (kW)'] / 1000
                # limit layer temp to 1300°C
                max_temp = 1300
                current_data.loc[current_data['Upper Layer Temp (C)'] > max_temp, 'Upper Layer Temp (C)'] = max_temp
                current_data.loc[current_data['Lower Layer Temp (C)'] > max_temp, 'Lower Layer Temp (C)'] = max_temp
                # current_data['Upper Layer Temp (C)'] = current_data.apply(
                #     lambda row: max_temp if row.name > max_temp else row['Upper Layer Temp (C)'],
                #     axis=1)
                # current_data['Lower Layer Temp (C)'] = current_data.apply(
                #     lambda row: max_temp if row.name > max_temp else row['Lower Layer Temp (C)'],
                #     axis=1)

            if current_iter == 0:
                hrr_fuel = current_data[['Unconstrained HRR (kW)']].copy()
                # L'utilisation des crochets doubles [['HRR (kW)']] garantit que le résultat est un DataFrame et non une
                # série

            # Set factors from exposed part and contribution law
            if EnvB.default_method == 'EC5':
                c_linear = 1  # instead 1.08 DO EC5-1-2 A.4.1 (3) (EC5)
                c_ceiling = 1  # Instead of 0.75 (LG Article)
            else:  # 'LG', 'Flux' ou 'Mlr'
                c_linear = 1.08
                c_ceiling = EnvB.default_ceiling_factor
            column_ll_factor = c_linear * int(compartment1.a_column_exposed > 0)
            column_ul_factor = column_ll_factor
            beam_factor = c_linear * int(compartment1.a_beam_exposed > 0)
            ceiling_factor = c_ceiling * int(compartment1.a_ceiling_exposed > 0)
            ul_factor = 1 * int(compartment1.a_wall_exposed > 0)
            ll_factor = 1 * int(compartment1.a_wall_exposed > 0)
            floor_factor = 1 * int(compartment1.a_floor_exposed > 0)

            # Add column Time in mn
            current_data.insert(0, 'Time (min)', current_data['Time (sec)'] / 60)
            # UL = Upper Layer (hot gas), LL = Lower Layer
            # Add height of hot gas layer (h_UL) from height of interface layer
            current_data = ajouter_colonne_modif(current_data, 'Layer (m)', 'h_UL (m)',
                                                 lambda x: h_ul(x, compartment1.height))
            # Calculate view factors (for all method even if some as 'EC5' doesn't use it)
            current_data = ajouter_colonne_modif(current_data, 'h_UL (m)', 'Phi_LL ()',
                                                 lambda x: phi_factor(compartment1, x))
            current_data = ajouter_colonne_modif(current_data, 'h_UL (m)', 'Phi_Floor ()',
                                                 lambda x: floor_view_factor(compartment1, x))
            # Calculate reduction factors from oxygen concentration
            current_data = ajouter_colonne_modif(current_data, 'O2 Upper (%)', 'RR_UL ()',
                                                 oxygen_reduction_factor)
            current_data = ajouter_colonne_modif(current_data, 'O2 Lower (%)', 'RR_LL ()',
                                                 oxygen_reduction_factor)
            # return wood density of wall,ceiling,floor, beam & column
            wd, cd, fd, bd, cod = compartment1.wood_density()

            if EnvB.default_method == 'LG':
                # Calculate incident flux from hot gas temperature
                # (Did char law with wood density)
                current_data = ajouter_colonne_modif(current_data, 'Upper Layer Temp (C)', 'qUL (kW/m²)',
                                                     flux_from_ul_temp)
                current_data = ajouter_colonne_multiplication(current_data, 'qUL (kW/m²)', 'Phi_LL ()',
                                                              'qLL (kW/m²)', 1)
                current_data = ajouter_colonne_multiplication(current_data, 'qUL (kW/m²)',
                                                              'Phi_Floor ()', 'qFloor (kW/m²)', 1)
                current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_ceiling (mm/min)',
                                                     lambda x: charring_rate_law(x, cd))
                current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_beam (mm/min)',
                                                     lambda x: charring_rate_law(x, bd))  # or
                current_data = ajouter_colonne_modif(current_data, 'qLL (kW/m²)', 'Beta_LL (mm/min)',
                                                     lambda x: charring_rate_law(x, wd))
                current_data = ajouter_colonne_modif(current_data, 'qLL (kW/m²)', 'Beta_column_LL (mm/min)',
                                                     lambda x: charring_rate_law(x, cod))
                current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_UL (mm/min)',
                                                     lambda x: charring_rate_law(x, wd))
                current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_column_UL (mm/min)',
                                                     lambda x: charring_rate_law(x, cod))
                current_data = ajouter_colonne_modif(current_data, 'qFloor (kW/m²)', 'Beta_floor (mm/min)',
                                                     lambda x: charring_rate_law(x, fd))
                current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_ceiling (mm/min)', ceiling_factor)
                current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_beam (mm/min)', beam_factor)
                current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_LL (mm/min)', ll_factor)
                current_data = colonne_multiplication(current_data, 'RR_LL ()',
                                                      'Beta_column_LL (mm/min)', column_ll_factor)
                current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_UL (mm/min)', ul_factor)
                current_data = colonne_multiplication(current_data, 'RR_UL ()',
                                                      'Beta_column_UL (mm/min)', column_ul_factor)
                current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_floor (mm/min)', floor_factor)
                # calculate char thickness
                current_data['e_char LL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_LL (mm/min)'), axis=1)
                current_data['e_char UL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_UL (mm/min)'), axis=1)
                current_data['e_char ceiling (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_ceiling (mm/min)'), axis=1)
                current_data['e_char floor (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_floor (mm/min)'), axis=1)
                current_data['e_char beam (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_beam (mm/min)'), axis=1)
                current_data['e_char column LL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_column_LL (mm/min)'), axis=1)
                current_data['e_char column UL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_column_UL (mm/min)'), axis=1)
            elif (EnvB.default_method == 'Flux') or (EnvB.default_method == 'Mlr'):
                # retrieve flux from zone model results
                current_data['Convective upper wall heat flux (kW/m2)'] = (current_data['Conv HT Coef UWall (W/m2.K)'] *
                                                                           (current_data['Upper Layer Temp (C)'] -
                                                                            current_data['Upper Wall Temp (C)']) / 1000)
                current_data['Convective lower Wall heat flux (kW/m2)'] = (current_data['Conv HT Coef LWall (W/m2.K)'] *
                                                                           (current_data['Lower Layer Temp (C)'] -
                                                                            current_data['Lower Wall Temp (C)']) / 1000)
                current_data['Convective ceiling heat flux (kW/m2)'] = (current_data['Conv HT Coef Ceil (W/m2.K)'] *
                                                                        (current_data['Upper Layer Temp (C)'] -
                                                                         current_data['Ceiling Temp (C)']) / 1000)
                current_data['Convective floor heat flux (kW/m2)'] = (current_data['Conv HT Coef Floor (W/m2.K)'] *
                                                                      (current_data['Lower Layer Temp (C)'] -
                                                                       current_data['Floor Temp (C)']) / 1000)
                current_data['qUL (kW/m²)'] = (current_data['Incident upper wall radiant flux (kW/m2)'] +
                                               current_data['Convective upper wall heat flux (kW/m2)'])
                current_data['qCeiling (kW/m²)'] = (current_data['Incident ceiling radiant flux (kW/m2)'] +
                                                    current_data['Convective ceiling heat flux (kW/m2)'])
                current_data['qLL (kW/m²)'] = (current_data['Incident lower Wall radiant flux (kW/m2)'] +
                                               current_data['Convective lower Wall heat flux (kW/m2)'])
                current_data['qFloor (kW/m²)'] = (current_data['Incident floor radiant flux (kW/m2)'] +
                                                  current_data['Convective floor heat flux (kW/m2)'])
                if EnvB.default_method == 'Flux':
                    current_data = ajouter_colonne_modif(current_data, 'qCeiling (kW/m²)', 'Beta_ceiling (mm/min)',
                                                         lambda x: charring_rate_law(x, cd))
                    # TODO check if better ceiling or UL for beam ? i.e or
                    # or # current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_beam (mm/min)',
                    #                                      lambda x: charring_rate_law(x, bd)) # or #
                    current_data = ajouter_colonne_modif(current_data, 'qCeiling (kW/m²)', 'Beta_beam (mm/min)',
                                                         lambda x: charring_rate_law(x, bd))
                    current_data = ajouter_colonne_modif(current_data, 'qLL (kW/m²)', 'Beta_LL (mm/min)',
                                                         lambda x: charring_rate_law(x, wd))
                    current_data = ajouter_colonne_modif(current_data, 'qLL (kW/m²)', 'Beta_column_LL (mm/min)',
                                                         lambda x: charring_rate_law(x, cod))
                    current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_UL (mm/min)',
                                                         lambda x: charring_rate_law(x, wd))
                    current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_column_UL (mm/min)',
                                                         lambda x: charring_rate_law(x, cod))
                    current_data = ajouter_colonne_modif(current_data, 'qFloor (kW/m²)', 'Beta_floor (mm/min)',
                                                         lambda x: charring_rate_law(x, fd))
                    # Apply Reduction factor due to lack of oxygen for char speed
                    current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_ceiling (mm/min)',
                                                          ceiling_factor)
                    current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_beam (mm/min)', beam_factor)
                    current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_LL (mm/min)', ll_factor)
                    current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_column_LL (mm/min)',
                                                          column_ll_factor)
                    current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_UL (mm/min)', ul_factor)
                    current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_column_UL (mm/min)',
                                                          column_ul_factor)
                    current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_floor (mm/min)', floor_factor)
                else:  # i.e (EnvB.default_method == 'Mlr')
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qCeiling (kW/m²)', 'e_mlr_ceiling (mm)',
                                   'mlr*_ceiling  (g kJ-1)', 'mlr_ceiling (g s-1 m-2)', 'Beta_ceiling (mm/min)',
                                   mlr_function, cd)
                    # TODO check if better ceiling or UL for beam ?
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qCeiling (kW/m²)', 'e_mlr_beam (mm)',
                                   'mlr*_beam  (g kJ-1)', 'mlr_beam (g s-1 m-2)', 'Beta_beam (mm/min)',
                                   mlr_function, bd)
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qLL (kW/m²)', 'e_mlr_LL (mm)',
                                   'mlr*_LL  (g kJ-1)', 'mlr_LL (g s-1 m-2)', 'Beta_LL (mm/min)',
                                   mlr_function, wd)
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qLL (kW/m²)', 'e_mlr_column_LL (mm)',
                                   'mlr*_column_LL  (g kJ-1)', 'mlr_column_LL (g s-1 m-2)', 'Beta_column_LL (mm/min)',
                                   mlr_function, cod)
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qUL (kW/m²)', 'e_mlr_UL (mm)',
                                   'mlr*_UL  (g kJ-1)', 'mlr_UL (g s-1 m-2)', 'Beta_UL (mm/min)',
                                   mlr_function, wd)
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qUL (kW/m²)', 'e_mlr_column_UL (mm)',
                                   'mlr*_column_UL  (g kJ-1)', 'mlr_column_UL (g s-1 m-2)', 'Beta_column_UL (mm/min)',
                                   mlr_function, cod)
                    ajoute_mlr_law(current_data, 'Time (sec)', 'qFloor (kW/m²)', 'e_mlr_floor (mm)',
                                   'mlr*_floor  (g kJ-1)', 'mlr_floor (g s-1 m-2)', 'Beta_floor (mm/min)',
                                   mlr_function, fd)
                    # Don't apply Reduction factor due to lack of oxygen for char speed for mlr method ?
                    current_data['Beta_ceiling (mm/min)'] = current_data['Beta_ceiling (mm/min)'] * ceiling_factor
                    current_data['Beta_beam (mm/min)'] = current_data['Beta_beam (mm/min)'] * beam_factor
                    current_data['Beta_LL (mm/min)'] = current_data['Beta_LL (mm/min)'] * ll_factor
                    current_data['Beta_column_LL (mm/min)'] = current_data['Beta_column_LL (mm/min)'] * column_ll_factor
                    current_data['Beta_column_UL (mm/min)'] = current_data['Beta_column_UL (mm/min)'] * column_ul_factor
                    current_data['Beta_UL (mm/min)'] = current_data['Beta_UL (mm/min)'] * ul_factor
                    current_data['Beta_floor (mm/min)'] = current_data['Beta_floor (mm/min)'] * floor_factor

                # calculate char thickness (for both 'Flux' & 'Mlr' methods
                current_data['e_char LL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_LL (mm/min)'), axis=1)
                current_data['e_char UL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_UL (mm/min)'), axis=1)
                current_data['e_char ceiling (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_ceiling (mm/min)'), axis=1)
                current_data['e_char floor (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_floor (mm/min)'), axis=1)
                current_data['e_char beam (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_beam (mm/min)'), axis=1)
                current_data['e_char column LL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_column_LL (mm/min)'), axis=1)
                current_data['e_char column UL (mm)'] = current_data.apply(
                    lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_column_UL (mm/min)'), axis=1)
            elif EnvB.default_method == 'EC5':
                # Methode EC5 e_char first calculated
                # Calculer l'intégrale de kelvin_2(T° gas) de t0 à ti en utilisant l'intégration cumulative par trapèzes
                current_data['Cumul e_char UL (mm)'] = ((sc.integrate.cumulative_trapezoid(
                    current_data['Upper Layer Temp (C)'].apply(kelvin_2), current_data['Time (min)'],
                    initial=0) / 1.35E5) ** (1 / 1.6))
                # Calculer la différence entre les éléments consécutifs de Cumul e_char
                # pour avoir l'épaisseur de carbonisation sur le pas de temps
                current_data['e_char UL (mm)'] = current_data['Cumul e_char UL (mm)'].diff().fillna(
                    current_data['Cumul e_char UL (mm)'])
                if EnvB.allow_char_energy_storage:
                    current_data = colonne_multiplication(current_data, 'RR_UL ()', 'e_char UL (mm)', 1)

                # Si Calcul "2 zones"
                current_data['Cumul e_char LL (mm)'] = ((sc.integrate.cumulative_trapezoid(
                    current_data['Lower Layer Temp (C)'].apply(kelvin_2), current_data['Time (min)'],
                    initial=0) / 1.35E5) ** (1 / 1.6))
                current_data['e_char LL (mm)'] = current_data['Cumul e_char LL (mm)'].diff().fillna(
                    current_data['Cumul e_char LL (mm)'])
                if EnvB.allow_char_energy_storage:
                    current_data = colonne_multiplication(current_data, 'RR_LL ()', 'e_char LL (mm)', 1)
                # fin Si Calcul "2 zones"

                # # Si Calcul "1 zone"
                # current_data['e_char LL (mm)'] = current_data['e_char UL (mm)'] * 1
                # if EnvB.allow_char_energy_storage:
                #     current_data = colonne_multiplication(current_data, 'RR_LL ()', 'e_char LL (mm)', 1)
                # current_data['Cumul e_char LL (mm)'] = current_data['e_char LL (mm)'].cumsum()
                # # fin Si Calcul "1 zone"

                current_data['e_char ceiling (mm)'] = current_data['e_char UL (mm)'] * ceiling_factor
                current_data['e_char floor (mm)'] = current_data['e_char LL (mm)'] * floor_factor
                current_data['e_char beam (mm)'] = current_data['e_char UL (mm)'] * beam_factor
                current_data['e_char column LL (mm)'] = current_data['e_char LL (mm)'] * column_ll_factor
                current_data['e_char column UL (mm)'] = current_data['e_char UL (mm)'] * column_ul_factor
                current_data['e_char LL (mm)'] = current_data['e_char LL (mm)'] * ll_factor
                current_data['e_char UL (mm)'] = current_data['e_char UL (mm)'] * ul_factor

                # Methode EC5 e_char first calculated and beta deducted
                current_data['Beta_LL (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char LL (mm)'), axis=1)
                current_data['Beta_UL (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char UL (mm)'), axis=1)
                current_data['Beta_ceiling (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char ceiling (mm)'), axis=1)
                current_data['Beta_floor (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char floor (mm)'), axis=1)
                current_data['Beta_beam (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char beam (mm)'), axis=1)
                current_data['Beta_column_LL (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char column LL (mm)'), axis=1)
                current_data['Beta_column_UL (mm/min)'] = current_data.apply(
                    lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char column UL (mm)'), axis=1)

            current_data['Cumul e_char ceiling (mm)'] = current_data['e_char ceiling (mm)'].cumsum()
            current_data['Cumul e_char floor (mm)'] = current_data['e_char floor (mm)'].cumsum()
            current_data['Cumul e_char beam (mm)'] = current_data['e_char beam (mm)'].cumsum()
            current_data['Cumul e_char column LL (mm)'] = current_data['e_char column LL (mm)'].cumsum()
            current_data['Cumul e_char column UL (mm)'] = current_data['e_char column UL (mm)'].cumsum()
            current_data['Cumul e_char LL (mm)'] = current_data['e_char LL (mm)'].cumsum()
            current_data['Cumul e_char UL (mm)'] = current_data['e_char UL (mm)'].cumsum()

            # Test if all wood is consumed and update
            if ceiling_factor > 0:
                if max_charring_update(current_data, max_char_ceiling, 'Cumul e_char ceiling (mm)',
                                       'Beta_ceiling (mm/min)', 'e_char ceiling (mm)'):
                    compartment1.wood_consumed = True
                    compartment1.ceiling_mat.wood_consumed = True
            if floor_factor > 0:
                if max_charring_update(current_data, max_char_floor, 'Cumul e_char floor (mm)',
                                       'Beta_floor (mm/min)', 'e_char floor (mm)'):
                    compartment1.wood_consumed = True
                    compartment1.floor_mat.wood_consumed = True
            if ul_factor > 0:
                if max_charring_update(current_data, max_char_wall, 'Cumul e_char UL (mm)',
                                       'Beta_UL (mm/min)', 'e_char UL (mm)'):
                    compartment1.wood_consumed = True
                    compartment1.wall_mat.wood_consumed = True
            if ll_factor > 0:
                if max_charring_update(current_data, max_char_wall, 'Cumul e_char LL (mm)',
                                       'Beta_LL (mm/min)', 'e_char LL (mm)'):
                    compartment1.wood_consumed = True
                    compartment1.wall_mat.wood_consumed = True
            if beam_factor > 0:
                if max_charring_update(current_data, max_char_beam, 'Cumul e_char beam (mm)',
                                       'Beta_beam (mm/min)', 'e_char beam (mm)'):
                    compartment1.wood_consumed = True
            if column_ul_factor > 0:
                if max_charring_update(current_data, max_char_column, 'Cumul e_char column UL (mm)',
                                       'Beta_column_UL (mm/min)', 'e_char column UL (mm)'):
                    compartment1.wood_consumed = True
            if column_ll_factor > 0:
                if max_charring_update(current_data, max_char_column, 'Cumul e_char column LL (mm)',
                                       'Beta_column_LL (mm/min)', 'e_char column LL (mm)'):
                    compartment1.wood_consumed = True
            # End Test all wood consumed and update

            if EnvB.default_method == 'EC5':
                # Calculate timber contribution Modif EC5-1-2 A.3 5390/60=89.3 (LG article) -> 120
                # TODO replace 120 or (5360/60) by a global variable (Environnement)
                f_linear_member_ul = 1 + (current_data['Cumul e_char UL (mm)'].max() / 1000) ** 2 * (4 - np.pi)
                # TODO ul / ll for linear element ?
                # f_linear_member_ll = 1 + (current_data['Cumul e_char LL (mm)'].max() / 1000) ** 2 * (4 - np.pi)
                s_10 = 120 * 17.5 / compartment1.fire.fuel.heat_of_combustion  # EC5-1-2 [kW/m²]
            else:
                # DO adjust for fuel heat_of_combustion
                # if re.search('Pine', compartment1.fire.fuel.fuel_type, re.IGNORE CASE):
                #     f_comb = 1
                # else:
                #     f_comb = Pine.heat_of_combustion / compartment1.fire.fuel.heat_of_combustion
                #     # note: hc_wood_value = 17.5 for EC5
                # # 5390/60 = LG Article value
                # s_10 = 5390 / 60 * f_comb
                s_10 = (5390 / 60 * EnvB.default_structural_wood_heat_of_combustion
                        / compartment1.fire.fuel.heat_of_combustion)
                f_linear_member_ul = 1
            qt_wall_factor = s_10 * compartment1.a_wall_exposed / compartment1.height
            qt_ceiling_factor = s_10 * compartment1.a_ceiling_exposed
            qt_column_factor = s_10 * compartment1.a_column_exposed / compartment1.height * f_linear_member_ul
            qt_beam_factor = s_10 * compartment1.a_beam_exposed * f_linear_member_ul
            qt_floor_factor = s_10 * compartment1.a_floor_exposed
            current_data = ajouter_colonne_multiplication(current_data, 'Layer (m)', 'Beta_LL (mm/min)',
                                                          'Q_timber_wall_LL (kW)', qt_wall_factor)
            current_data = ajouter_colonne_multiplication(current_data, 'h_UL (m)', 'Beta_UL (mm/min)',
                                                          'Q_timber_wall_UL (kW)', qt_wall_factor)
            current_data = ajouter_colonne_multiplication(current_data, 'Layer (m)', 'Beta_column_LL (mm/min)',
                                                          'Q_timber_column_LL (kW)', qt_column_factor)
            current_data = ajouter_colonne_multiplication(current_data, 'h_UL (m)', 'Beta_column_UL (mm/min)',
                                                          'Q_timber_column_UL (kW)', qt_column_factor)
            current_data['Q_timber_ceiling (kW)'] = current_data['Beta_ceiling (mm/min)'] * qt_ceiling_factor
            current_data['Q_timber_beam (kW)'] = current_data['Beta_beam (mm/min)'] * qt_beam_factor
            current_data['Q_timber_floor (kW)'] = current_data['Beta_floor (mm/min)'] * qt_floor_factor

            # Contribution_Brisk 2: Ajout Contribution fall of protection EC5-> False TODO change if True
            if EnvB.allow_contribution_protected:
                # # filtered_data = current_data[current_data['Upper Wall Interface Temp (C)'] >= EnvB.fall_off_temp]
                filtered_data = current_data[current_data['Time (min)'] >= compartment1.wall_time_fo]
                if compartment1.wall_protected and compartment1.wall_time_fo > 0 and not filtered_data.empty:
                    index_fall_off = filtered_data.index[0]
                    # Pas utile, pour un besoin de présentation dans l'Excel résultat seulement
                    # Remplacer les valeurs inférieures à l'index fall_off par 0.
                    # # current_data.loc[:index_fall_off - 1, 'Upper Wall Interface Temp (C)'] = (
                    # #     current_data.loc[:index_fall_off - 1, 'Upper Wall Interface Temp (C)']
                    # #    .apply(lambda x: 0 if x < EnvB.fall_off_temp else x))
                    if EnvB.default_method == 'LG':
                        if EnvB.current_zone_model[1] == 'Brisk':
                            current_data['Upper Wall Interface Temp (C)'] = current_data.apply(
                                lambda row: 0 if row.name < index_fall_off else row['Upper Wall Interface Temp (C)'],
                                axis=1)
                        # Créer la nouvelle colonne 'Beta_protected_UL (mm/min)'
                        current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_protected_UL (mm/min)',
                                                             lambda x: charring_rate_law(x, wd))
                        current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_protected_UL (mm/min)', 1)
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_protected_UL (mm/min)'), axis=1)
                        current_data['Beta_protected_UL (mm/min)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['Beta_protected_UL (mm/min)'], axis=1)
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_UL (mm)'], axis=1)
                    elif EnvB.default_method == 'EC5':
                        # New EC5:
                        tul_protected = current_data[['Upper Layer Temp (C)']].copy().apply(
                            lambda row: 0 if row.name < index_fall_off else row['Upper Layer Temp (C)'], axis=1)
                        current_data['Cumul e_char protected_UL (mm)'] = (sc.integrate.cumulative_trapezoid(
                            tul_protected.apply(kelvin_2), current_data['Time (min)'],
                            initial=index_fall_off) / 1.35E5) ** (1 / 1.6)
                        # / ! \ attention (défaut méthode EC5 ?) la carbonisation commence même avec une température
                        # de gaz nulle : Loi en K²
                        current_data['e_char protected_UL (mm)'] = (
                            current_data['Cumul e_char protected_UL (mm)'].diff()
                            .fillna(current_data['Cumul e_char protected_UL (mm)']))
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_UL (mm)'], axis=1)
                        if EnvB.allow_char_energy_storage:
                            current_data = colonne_multiplication(
                                current_data, 'RR_UL ()', 'e_char protected_UL (mm)', 1)
                        current_data['Beta_protected_UL (mm/min)'] = current_data.apply(
                            lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char protected_UL (mm)'), axis=1)
                        # End New EC5
                    elif EnvB.default_method == 'Mlr':
                        # New version Brisk_calc4:
                        current_data['qULp (kW/m²)'] = current_data['qUL (kW/m²)']
                        current_data['qULp (kW/m²)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['qULp (kW/m²)'], axis=1)
                        ajoute_mlr_law(current_data, 'Time (sec)', 'qULp (kW/m²)', 'e_mlr_ULp (mm)',
                                       'mlr*_ULp  (g kJ-1)', 'mlr_ULp (g s-1 m-2)', 'Beta_protected_UL (mm/min)',
                                       mlr_function, wd)
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_protected_UL (mm/min)'), axis=1)
                    elif EnvB.default_method == 'Flux':
                        current_data['Upper Wall Interface Temp (C)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['Upper Wall Interface Temp (C)'],
                            axis=1)
                        # Créer la nouvelle colonne 'Beta_protected_UL (mm/min)'
                        current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)', 'Beta_protected_UL (mm/min)',
                                                             lambda x: charring_rate_law(x, wd))
                        current_data = colonne_multiplication(current_data, 'RR_UL ()', 'Beta_protected_UL (mm/min)', 1)
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_protected_UL (mm/min)'), axis=1)
                        current_data['Beta_protected_UL (mm/min)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['Beta_protected_UL (mm/min)'], axis=1)
                        current_data['e_char protected_UL (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_UL (mm)'], axis=1)

                    current_data['Cumul e_char protected_UL (mm)'] = current_data['e_char protected_UL (mm)'].cumsum()
                    if max_charring_update(current_data, max_char_wall, 'Cumul e_char protected_UL (mm)',
                                           'Beta_protected_UL (mm/min)', 'e_char protected_UL (mm)'):
                        compartment1.wood_consumed = True
                        compartment1.wall_mat.wood_consumed = True
                    qt_protected_wall_factor = (s_10 * (compartment1.a_wall - compartment1.a_v -
                                                        compartment1.a_wall_exposed) / compartment1.height)
                    current_data = ajouter_colonne_multiplication(current_data, 'h_UL (m)',
                                                                  'Beta_protected_UL (mm/min)',
                                                                  'Q_protected_wall_UL (kW)',
                                                                  qt_protected_wall_factor)
                    if (EnvB.default_method == 'LG') or (EnvB.default_method == 'Flux'):
                        if EnvB.current_zone_model[1] == 'Brisk':
                            current_data['Lower Wall Interface Temp (C)'] = current_data.apply(
                                lambda row: 0 if row.name < index_fall_off else row['Lower Wall Interface Temp (C)'],
                                axis=1)
                        # Créer la nouvelle colonne 'new'
                        current_data = ajouter_colonne_modif(current_data, 'qLL (kW/m²)', 'Beta_protected_LL (mm/min)',
                                                             lambda x: charring_rate_law(x, wd))
                        current_data = colonne_multiplication(current_data, 'RR_LL ()', 'Beta_protected_LL (mm/min)', 1)
                        current_data['e_char protected_LL (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_protected_LL (mm/min)'), axis=1)
                        current_data['Beta_protected_LL (mm/min)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['Beta_protected_LL (mm/min)'], axis=1)
                        current_data['e_char protected_LL (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_LL (mm)'], axis=1)
                    elif EnvB.default_method == 'EC5':
                        # New EC5:
                        tll_protected = current_data[['Lower Layer Temp (C)']].copy().apply(
                            lambda row: 0 if row.name < index_fall_off else row['Lower Layer Temp (C)'], axis=1)
                        current_data['Cumul e_char protected_LL (mm)'] = (sc.integrate.cumulative_trapezoid(
                            tll_protected.apply(kelvin_2), current_data['Time (min)'],
                            initial=index_fall_off) / 1.35E5) ** (1 / 1.6)
                        # / ! \ attention (défaut méthode EC5 ?) la carbonisation commence même avec une température
                        # de gaz nulle : Loi en K²
                        current_data['e_char protected_LL (mm)'] = (
                            current_data['Cumul e_char protected_LL (mm)'].diff().
                            fillna(current_data['Cumul e_char protected_LL (mm)']))
                        current_data['e_char protected_LL (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_LL (mm)'], axis=1)
                        if EnvB.allow_char_energy_storage:
                            current_data = colonne_multiplication(current_data, 'RR_LL ()',
                                                                  'e_char protected_LL (mm)', 1)
                        current_data['Beta_protected_LL (mm/min)'] = current_data.apply(
                            lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char protected_LL (mm)'), axis=1)
                        # End New EC5
                    elif EnvB.default_method == 'Mlr':
                        # New version Brisk_calc4:
                        current_data['qLLp (kW/m²)'] = current_data['qLL (kW/m²)']
                        current_data['qLLp (kW/m²)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['qLLp (kW/m²)'], axis=1)
                        ajoute_mlr_law(current_data, 'Time (sec)', 'qLLp (kW/m²)', 'e_mlr_LLp (mm)',
                                       'mlr*_LLp  (g kJ-1)', 'mlr_LLp (g s-1 m-2)', 'Beta_protected_LL (mm/min)',
                                       mlr_function, wd)
                        current_data['e_char protected_LL (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)', 'Beta_protected_LL (mm/min)'), axis=1)

                    current_data['Cumul e_char protected_LL (mm)'] = current_data['e_char protected_LL (mm)'].cumsum()
                    if max_charring_update(current_data, max_char_wall, 'Cumul e_char protected_LL (mm)',
                                           'Beta_protected_LL (mm/min)', 'e_char protected_LL (mm)'):
                        compartment1.wood_consumed = True
                        compartment1.wall_mat.wood_consumed = True
                    qt_protected_wall_factor = (s_10 * (compartment1.a_wall - compartment1.a_v -
                                                        compartment1.a_wall_exposed) / compartment1.height)
                    current_data = ajouter_colonne_multiplication(current_data, 'Layer (m)',
                                                                  'Beta_protected_LL (mm/min)',
                                                                  'Q_protected_wall_LL (kW)',
                                                                  qt_protected_wall_factor)
                else:
                    current_data['Beta_protected_UL (mm/min)'] = 0
                    current_data['e_char protected_UL (mm)'] = 0
                    current_data['Cumul e_char protected_UL (mm)'] = 0
                    current_data['Q_protected_wall_UL (kW)'] = 0
                    current_data['Beta_protected_LL (mm/min)'] = 0
                    current_data['e_char protected_LL (mm)'] = 0
                    current_data['Cumul e_char protected_LL (mm)'] = 0
                    current_data['Q_protected_wall_LL (kW)'] = 0
                # free_mem(filtered_data)

                # # filtered_data = current_data[current_data['Ceiling Interface Temp (C)'] >= EnvB.fall_off_temp]
                filtered_data = current_data[current_data['Time (min)'] >= compartment1.ceiling_time_fo]
                if compartment1.ceiling_protected and compartment1.ceiling_time_fo > 0 and not filtered_data.empty:
                    index_fall_off = filtered_data.index[0]
                    # current_data['Ceiling Interface Temp (C)'[] = current_data.apply(
                    #     lambda row: 0 if row.name < index_fall_off else row['Ceiling Interface Temp (C)'], axis=1)
                    if (EnvB.default_method == 'LG') or (EnvB.default_method == 'Flux'):
                        # Créer la nouvelle colonne 'new'
                        current_data = ajouter_colonne_modif(current_data, 'qUL (kW/m²)',
                                                             'Beta_protected_ceiling (mm/min)',
                                                             lambda x: charring_rate_law(x, cd))
                        current_data = colonne_multiplication(current_data, 'RR_UL ()',
                                                              'Beta_protected_ceiling (mm/min)',
                                                              EnvB.default_ceiling_factor)
                        current_data['e_char protected_ceiling (mm)'] = current_data.apply(
                            lambda row: e_char(row, current_data, 'Time (sec)',
                                               'Beta_protected_ceiling (mm/min)'), axis=1)
                        current_data['Beta_protected_ceiling (mm/min)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else
                            row['Beta_protected_ceiling (mm/min)'], axis=1)
                        current_data['e_char protected_ceiling (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_ceiling (mm)'],
                            axis=1)
                    elif EnvB.default_method == 'Mlr':
                        # New version Brisk_calc4:
                        current_data['qCeiling p (kW/m²)'] = current_data['qCeiling (kW/m²)']
                        current_data['qCeiling p (kW/m²)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['qCeiling p (kW/m²)'], axis=1)
                        ajoute_mlr_law(current_data, 'Time (sec)', 'qCeiling p (kW/m²)', 'e_mlr_cp (mm)',
                                       'mlr*_cp  (g kJ-1)', 'mlr_cp (g s-1 m-2)', 'Beta_protected_ceiling (mm/min)',
                                       mlr_function, cd)
                        current_data['Beta_protected_ceiling (mm/min)'] = (
                                current_data['Beta_protected_ceiling (mm/min)'] * EnvB.default_ceiling_factor)
                        current_data['e_char protected_ceiling (mm)'] = (
                            current_data.apply(lambda row: e_char(row, current_data, 'Time (sec)',
                                                                  'Beta_protected_ceiling (mm/min)'), axis=1))
                    elif EnvB.default_method == 'EC5':
                        # New EC5:
                        tul_protected = current_data[['Upper Layer Temp (C)']].copy().apply(
                            lambda row: 0 if row.name < index_fall_off else row['Upper Layer Temp (C)'], axis=1)
                        current_data['Cumul e_char protected_ceiling (mm)'] = (sc.integrate.cumulative_trapezoid(
                            tul_protected.apply(kelvin_2), current_data['Time (min)'],
                            initial=index_fall_off) / 1.35E5) ** (1 / 1.6)
                        # / ! \ attention (défaut méthode EC5 ?) la carbonisation commence même avec une température
                        # de gaz nulle : Loi en K²
                        current_data['e_char protected_ceiling (mm)'] = (
                            current_data['Cumul e_char protected_ceiling (mm)'].diff()
                            .fillna(current_data['Cumul e_char protected_ceiling (mm)']))
                        current_data['e_char protected_ceiling (mm)'] = current_data.apply(
                            lambda row: 0 if row.name < index_fall_off else row['e_char protected_ceiling (mm)'],
                            axis=1)
                        if EnvB.allow_char_energy_storage:
                            current_data = colonne_multiplication(current_data, 'RR_UL ()',
                                                                  'e_char protected_ceiling (mm)', 1)
                        current_data['Beta_protected_ceiling (mm/min)'] = current_data.apply(
                            lambda row: beta_char(row, current_data, 'Time (sec)', 'e_char protected_ceiling (mm)'),
                            axis=1)
                        # End New EC5

                    current_data['Cumul e_char protected_ceiling (mm)'] = (
                        current_data['e_char protected_ceiling (mm)'].cumsum())
                    if max_charring_update(current_data, max_char_ceiling, 'Cumul e_char protected_ceiling (mm)',
                                           'Beta_protected_ceiling (mm/min)', 'e_char protected_ceiling (mm)'):
                        compartment1.wood_consumed = True
                        compartment1.wall_mat.wood_consumed = True
                    qt_protected_ceiling_factor = s_10 * (compartment1.a_f - compartment1.a_ceiling_exposed)
                    current_data['Q_protected_ceiling (kW)'] = (
                            current_data['Beta_protected_ceiling (mm/min)'] * qt_protected_ceiling_factor)
                else:
                    current_data['Beta_protected_ceiling (mm/min)'] = 0
                    current_data['e_char protected_ceiling (mm)'] = 0
                    current_data['Cumul e_char protected_ceiling (mm)'] = 0
                    current_data['Q_protected_ceiling (kW)'] = 0
            else:
                current_data['Beta_protected_UL (mm/min)'] = 0
                current_data['e_char protected_UL (mm)'] = 0
                current_data['Cumul e_char protected_UL (mm)'] = 0
                current_data['Beta_protected_LL (mm/min)'] = 0
                current_data['e_char protected_LL (mm)'] = 0
                current_data['Cumul e_char protected_LL (mm)'] = 0
                current_data['Beta_protected_ceiling (mm/min)'] = 0
                current_data['e_char protected_ceiling (mm)'] = 0
                current_data['Cumul e_char protected_ceiling (mm)'] = 0
                current_data['Q_protected_wall_UL (kW)'] = 0
                current_data['Q_protected_wall_LL (kW)'] = 0
                current_data['Q_protected_ceiling (kW)'] = 0
                # free_mem(filtered_data)
                # TODO protected floor, beam, column ?

            # Calculate new HRR Ajout de la colonne Hrr_Iter+1 (HRR_Fuel+Q_timber) à Curent_Iteration
            current_data['Q_timber (kW)'] = (current_data['Q_timber_ceiling (kW)'] +
                                             current_data['Q_timber_wall_UL (kW)'] +
                                             current_data['Q_timber_wall_LL (kW)'] +
                                             current_data['Q_timber_column_LL (kW)'] +
                                             current_data['Q_timber_column_UL (kW)'] +
                                             current_data['Q_timber_beam (kW)'] +
                                             current_data['Q_timber_floor (kW)'] +
                                             current_data['Q_protected_wall_UL (kW)'] +
                                             current_data['Q_protected_wall_LL (kW)'] +
                                             current_data['Q_protected_ceiling (kW)'])

            current_data['New HRR (kW)'] = hrr_fuel['Unconstrained HRR (kW)'] + current_data['Q_timber (kW)']

            # Methode EC5 calcul effective cross-section
            d0w, d0c, d0f, d0b, d0co = compartment1.wood_zero_strength_layer_depth()
            if ceiling_factor > 0:
                add_effective_char_depth_column(current_data, 'Upper Layer Temp (C)', 'Cumul e_char ceiling (mm)',
                                                'UL', 'Ceiling', d0c)
            if floor_factor > 0:
                add_effective_char_depth_column(current_data, 'Lower Layer Temp (C)', 'Cumul e_char floor (mm)',
                                                'LL', 'Floor', d0f)
            if ul_factor > 0:
                add_effective_char_depth_column(current_data, 'Upper Layer Temp (C)', 'Cumul e_char UL (mm)',
                                                'UL', 'Wall UL', d0w)
            if ll_factor > 0:
                add_effective_char_depth_column(current_data, 'Lower Layer Temp (C)', 'Cumul e_char LL (mm)',
                                                'LL', 'Wall LL', d0w)
            if beam_factor > 0:
                add_effective_char_depth_column(current_data, 'Upper Layer Temp (C)', 'Cumul e_char beam (mm)',
                                                'UL', 'Beam', d0b)
            if column_ul_factor > 0:
                add_effective_char_depth_column(current_data, 'Upper Layer Temp (C)', 'Cumul e_char column UL (mm)',
                                                'UL', 'Column UL', d0co)
            if column_ll_factor > 0:
                add_effective_char_depth_column(current_data, 'Lower Layer Temp (C)', 'Cumul e_char column LL (mm)',
                                                'LL', 'Column LL', d0co)
            # End Methode EC5 calcul effective cross-section

            z_wall = 'Z Wall (m)'
            if compartment1.time_char_fin < compartment1.exposed_time():
                index = (current_data['Time (sec)'] - compartment1.time_char_fin).abs().idxmin()
                # .abs().id x min() Trouve le 0 de l'opération dans la colonne ;-)
            else:
                index = 0
            ajouter_colonne_char_z_wall(current_data, z_wall, 'e_char(z_wall) (mm)', compartment1, False, index)
            if (current_data['Q_protected_wall_LL (kW)'] + current_data['Q_protected_wall_UL (kW)']).any() != 0:
                ajouter_colonne_char_z_wall(current_data, z_wall, 'e_char protected(z_wall) (mm)',
                                            compartment1, True, index)
            else:
                current_data['e_char protected(z_wall) (mm)'] = 0
            # else:
            #     ajouter_colonne_char_z_wall(current_data, z_wall, 'e_char(z_wall) (mm)', compartment1, False)
            print(current_data)

            # écriture des fichiers input de la prochaine itération avec le nouvel HRR
            if EnvB.current_zone_model[1] == 'Brisk':
                fichier_xml_entree = destination_dir + '\\' + f"{EnvB.BaseModel}{current_iter + 1}.xml"
                remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
                                       'base_name', f"{EnvB.BaseModel}{current_iter + 1}")
                fichier_xml_entree = destination_dir + '\\' + 'items.xml'
                if current_iter == 0:
                    fire_name = lire_texte_xml(fichier_xml_entree, 'description')
                remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
                                       'description', fire_name + f"_iter{current_iter + 1}")
                # noinspection SpellCheckingInspection
                remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree,
                                       'detaileddescription', fire_name + f"_iter{current_iter + 1}")
                remplacer_double_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "hrr",
                                                  current_data['Time (sec)'], current_data['New HRR (kW)'])

                # Mise à jour de l'objet Fire pour avoir la bonne masse de combustible en function du nouvel HRR
                new_hrr = current_data['New HRR (kW)']
                f_new = HrrFunction(current_data['Time (sec)'].tolist(), new_hrr.tolist())
                mass = f_new.integral_total() / (compartment1.fire.fuel.heat_of_combustion * 1000)
                remplacer_str_dans_xml(fichier_xml_entree, fichier_xml_entree, 'mass', str(mass))
                fichier_xml_entree = destination_dir + '\\' + 'input1.xml'
                remplacer_nombres_dans_xml(fichier_xml_entree, fichier_xml_entree, "HRR", new_hrr)
                modif_fire_description_in_xml(fichier_xml_entree, fire_name + f"_iter{current_iter + 1}")
            elif EnvB.current_zone_model[1] == 'Cfast':
                new_hrr = current_data['New HRR (kW)']
                f_new = HrrFunction(current_data['Time (sec)'].tolist(), new_hrr.tolist())
                compartment1.hrr_fuel = f_new
                compartment1.to_cfast(destination_dir, current_iter + 1)

            # écriture des fichiers excel d'itération et de synthèse (avec graphiques)
            temps = 'Time (min)'

            # Add Ingberg time equivalency of fuel fire
            # if current_iter == 0:
            if EnvB.current_zone_model[1] == 'PC':
                temp_curve = 'Temp EC param curve (C)'
            else:
                temp_curve = 'Upper Layer Temp (C)'
            current_data['Ingberg surface'] = current_data.apply(
                lambda row: s_ingberg(row, current_data, temps, temp_curve), axis=1)
            current_data['Ingberg surface equivalency (C.min)'] = current_data['Ingberg surface'].cumsum()
            # cumul_ingberg_iso = TimeFunction(current_data[temps].tolist(), current_data['cumul ingberg'].tolist())
            s = current_data['Ingberg surface equivalency (C.min)'].max()
            xt = cumul_ingberg_iso.abscisse(s)
            if not xt == 0:
                fd_factor = 120 / xt
            if current_iter == 0:
                compartment1.fire.fuel_ingberg_equivalency = u_round(xt, '[mn]')
                compartment1.fire.total_fire_ingberg_equivalency = compartment1.fire.fuel_ingberg_equivalency
            else:
                compartment1.fire.total_fire_ingberg_equivalency = u_round(xt, '[mn]')
            # End of adding Ingberg time equivalency of fuel fire

            if current_iter == 0:
                write_xls_from_dataframe(current_data, current_iter, [(temps, 'Cumul e_char LL (mm)'),
                                                                      (temps, 'Cumul e_char UL (mm)'),
                                                                      (temps, 'Cumul e_char ceiling (mm)'),
                                                                      (temps, 'Cumul e_char column LL (mm)'),
                                                                      (temps, 'Cumul e_char column UL (mm)'),
                                                                      (temps, 'Cumul e_char beam (mm)'),
                                                                      (temps, 'Upper Layer Temp (C)'),
                                                                      (temps, 'Unconstrained HRR (kW)'),
                                                                      (temps, 'HRR (kW)'),
                                                                      (temps, 'Upper Layer '),
                                                                      (temps, 'O2 Upper (%)'),
                                                                      (z_wall, 'e_char(z_wall) (mm)')], compartment1)
            else:
                add_data_and_series(current_data, current_iter, [(temps, 'Cumul e_char LL (mm)'),
                                                                 (temps, 'Cumul e_char UL (mm)'),
                                                                 (temps, 'Cumul e_char ceiling (mm)'),
                                                                 (temps, 'Cumul e_char column LL (mm)'),
                                                                 (temps, 'Cumul e_char column UL (mm)'),
                                                                 (temps, 'Cumul e_char beam (mm)'),
                                                                 (temps, 'Upper Layer Temp (C)'),
                                                                 (temps, 'Unconstrained HRR (kW)'),
                                                                 (temps, 'HRR (kW)'),
                                                                 (temps, 'Upper Layer '),
                                                                 (temps, 'O2 Upper (%)'),
                                                                 (z_wall, 'e_char(z_wall) (mm)')], compartment1)
            current_iter = current_iter + 1

        # input("Appuyez sur n'importe quelle touche pour continuer...")

    # Fin Implémentation méthode itérative originelle de l'article [1]
    # ###########################################################################

    end_time1 = time.time()
    elapsed_time = end_time1 - start_time
    EnvB.CalculationTime = round(elapsed_time)  # str(timedelta(seconds=round(elapsed_time))
    compartment1.calculated_ok = '0'
    # Ajouter la première Feuille avec description du compartiment
    compartment1.dump_to_first_sheet_xlsx(hrr_initial)
    #  Automatisation à partir du fichier excel récap essais et ajout des résultats de calculs pour chaque essai
    write_xls_char_results(current_data, compartment1.id, compartment1.time_char_fin / 60,
                           compartment1.model_extinction(), False, compartment1.warning)

    free_mem(current_data)
    # DO implémentation de la méthode "Flux incident"
    # DO implémentation de la méthode "MLR"
    # DO création de la version CFAST

    end_time = time.time()
    elapsed_time = end_time - start_time
    # print("Temps de calcul : {:.2f} secondes".format(elapsed_time))
    # Convertir en heures, minutes et secondes
    elapsed_time_formatted = str(timedelta(seconds=round(elapsed_time)))
    print(f"Temps de calcul: {elapsed_time_formatted}")
